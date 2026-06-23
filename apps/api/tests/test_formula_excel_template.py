from io import BytesIO
from datetime import date

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel, create_engine

from formulia_api.formula_excel_template import FormulaExcelParameter, sorted_export_parameters
from formulia_api.main import create_app
from formulia_api.parameter_order import PARAMETER_ORDER


USER_A = "10000000-0000-0000-0000-000000000001"


def test_canonical_parameter_order_matches_atlantica_order() -> None:
    assert PARAMETER_ORDER == [
        "Ntotal",
        "Norg",
        "Nnitr",
        "Nure",
        "Namo",
        "K2O",
        "P2O5",
        "CaO",
        "MgO",
        "SO3",
        "Zn",
        "Mn",
        "Fe",
        "Cu",
        "B",
        "Mo",
        "Co",
        "SiO2",
        "Mseca",
        "Morg",
        "Corg",
        "Extracto Húmico total",
        "Acidos fulvicos",
        "Acidos húmicos",
        "Extracto de Algas",
        "Polisacaridos",
        "Sum AA totales",
        "Sum AA libres",
        "Ac aspartico",
        "Ac glutamico",
        "Alanina",
        "Glicina",
        "Histidina",
        "Isoleucina",
        "Leucina",
        "Lisina",
        "Serina",
        "Tirosina",
        "Treonina",
        "Valina",
        "Arginina",
        "Fenilalanina",
        "Metionina",
        "Prolina",
        "Hidroxiprolina",
        "Triptofano",
        "As",
        "Hg",
        "Pb",
        "Cd",
        "Cr",
        "Ni",
    ]


def test_export_parameters_use_canonical_parameter_order() -> None:
    parameters = [
        FormulaExcelParameter(code="Ni", label="Ni"),
        FormulaExcelParameter(code="K2O", label="K2O"),
        FormulaExcelParameter(code="Acidos humicos", label="Acidos humicos"),
        FormulaExcelParameter(code="Ntotal", label="Ntotal"),
        FormulaExcelParameter(code="Custom", label="Custom"),
        FormulaExcelParameter(code="Zn", label="Zn"),
    ]

    ordered = sorted_export_parameters(parameters)

    assert [parameter.label for parameter in ordered] == [
        "Ntotal",
        "K2O",
        "Zn",
        "Acidos humicos",
        "Ni",
        "Custom",
    ]


def make_client() -> TestClient:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    app = create_app(engine)
    SQLModel.metadata.create_all(engine)
    return TestClient(app)


def create_tenant(client: TestClient, user_id: str, slug: str) -> str:
    response = client.post(
        "/api/v1/tenants",
        headers={"X-User-Id": user_id},
        json={"name": slug.title(), "slug": slug},
    )
    assert response.status_code == 201
    return response.json()["id"]


def test_saved_formula_exports_atlantica_id_lab_template() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    ntotal = client.post(
        "/api/v1/parameters",
        headers=headers,
        json={"code": "Ntotal", "name": "Nitrogen total", "unit": "% p/p"},
    ).json()
    k2o = client.post(
        "/api/v1/parameters",
        headers=headers,
        json={"code": "K2O", "name": "Potassium", "unit": "% p/p"},
    ).json()
    water = _create_material(client, headers, "AGUA", "WATER", ntotal["id"], 0, 0)
    glycine = _create_material(client, headers, "GLICINA", "GLY", ntotal["id"], 1, 18)
    client.post(
        f"/api/v1/raw-materials/{glycine['id']}/parameter-values",
        headers=headers,
        json={"parameter_id": k2o["id"], "value": 0},
    )
    formula = client.post(
        "/api/v1/formulas",
        headers=headers,
        json={
            "name": "MICROCAT BON + AA - MUESTRA: F1",
            "objective": "Saved formula export template test.",
            "items": [
                {"raw_material_id": water["id"], "percentage": 50},
                {"raw_material_id": glycine["id"], "percentage": 50},
            ],
        },
    ).json()

    response = client.get(
        f"/api/v1/formulas/{formula['id']}/exports/atlantica-id-lab.xlsx",
        headers={**headers, "Origin": "http://127.0.0.1:3000"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    expected_file_name = (
        f"DT_MICROCAT_BON_AA_-_MUESTRA_F1_{date.today().isoformat()}.xlsx"
    )
    assert response.headers["content-disposition"] == (
        f'attachment; filename="{expected_file_name}"'
    )
    assert response.headers["access-control-expose-headers"] == "Content-Disposition"
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == ["Calculadora", "Hoja Lab", "Composición"]

    calculator = workbook["Calculadora"]
    assert calculator["A1"].value == "Materia Prima"
    assert calculator["D1"].value == "Ntotal"
    assert calculator["E1"].value == "K2O"
    assert calculator["A2"].value == "AGUA"
    assert calculator["A3"].value == "GLICINA"
    assert calculator["B5"].value == "=SUM(B2:B4)"
    assert calculator["C5"].value == "=SUMPRODUCT($B$2:$B$3,C2:C3)/100"
    assert calculator["D5"].value == "=SUMPRODUCT($B$2:$B$3,D2:D3)/100"

    lab = workbook["Hoja Lab"]
    assert lab["C3"].value == "MICROCAT BON + AA - MUESTRA: F1"
    assert lab["E6"].value == "=D6*10/2"
    assert lab["F7"].value == "=D7*10*2"
    assert lab["D9"].value == "=SUM(D6:D7)"
    assert len(lab._images) == 1

    composition = workbook["Composición"]
    assert composition["A2"].value == "Parámetro"
    assert composition["A3"].value == "Precio"
    assert composition["B3"].value == "='Calculadora'!C5"
    assert composition["A4"].value == "Ntotal"
    assert composition["B4"].value == "='Calculadora'!D5"
    assert composition["A5"].value == "K2O"
    assert composition["B5"].value == "='Calculadora'!E5"


def test_builder_draft_exports_atlantica_id_lab_template() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    parameter = client.post(
        "/api/v1/parameters",
        headers=headers,
        json={"code": "B", "name": "Boron", "unit": "% p/p"},
    ).json()
    boron = _create_material(client, headers, "BORO", "BOR", parameter["id"], 2, 5)

    response = client.post(
        "/api/v1/formulas/exports/atlantica-id-lab.xlsx",
        headers=headers,
        json={
            "name": "Draft Formula",
            "objective": "Draft formula export template test.",
            "items": [{"raw_material_id": boron["id"], "percentage": 100}],
            "metadata": {
                "sample_code": "F1",
                "lab_date": "2025-10-02",
                "density": 1.01592,
                "ph": 10.48,
            },
        },
    )

    assert response.status_code == 200
    expected_file_name = f"DT_Draft_Formula_{date.today().isoformat()}.xlsx"
    assert response.headers["content-disposition"] == (
        f'attachment; filename="{expected_file_name}"'
    )
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook["Calculadora"]["D1"].value == "B"
    assert workbook["Hoja Lab"]["C3"].value == "Draft Formula - MUESTRA: F1"
    assert workbook["Hoja Lab"]["E6"].value == "=D6*10/2"
    assert workbook["Hoja Lab"]["E10"].value == 1.01592
    assert workbook["Hoja Lab"]["E11"].value == 10.48


def test_builder_draft_export_requires_description() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")

    response = client.post(
        "/api/v1/formulas/exports/atlantica-id-lab.xlsx",
        headers={"X-User-Id": USER_A, "X-Tenant-Id": tenant_id},
        json={"name": "No Description Draft", "objective": " ", "items": []},
    )

    assert response.status_code == 422


def test_builder_draft_export_requires_name() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")

    response = client.post(
        "/api/v1/formulas/exports/atlantica-id-lab.xlsx",
        headers={"X-User-Id": USER_A, "X-Tenant-Id": tenant_id},
        json={"name": " ", "objective": "Draft formula description.", "items": []},
    )

    assert response.status_code == 422


def _create_material(
    client: TestClient,
    headers: dict[str, str],
    name: str,
    code: str,
    parameter_id: str,
    price: float,
    parameter_value: float,
) -> dict[str, object]:
    material = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": name, "code": code},
    ).json()
    assert client.post(
        f"/api/v1/raw-materials/{material['id']}/prices",
        headers=headers,
        json={"price": price, "currency": "EUR", "unit": "kg"},
    ).status_code == 201
    assert client.post(
        f"/api/v1/raw-materials/{material['id']}/parameter-values",
        headers=headers,
        json={"parameter_id": parameter_id, "value": parameter_value},
    ).status_code == 201
    return material
