from io import BytesIO
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel, create_engine

from formulia_api.main import create_app


USER_A = "10000000-0000-0000-0000-000000000001"
USER_B = "20000000-0000-0000-0000-000000000001"


def make_client() -> TestClient:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    app = create_app(engine)
    SQLModel.metadata.create_all(engine)
    return TestClient(app)


def create_tenant(client: TestClient, user_id: str, slug: str) -> str:
    response = client.post(
        "/api/v1/tenants",
        headers={"X-User-Id": user_id},
        json={"name": slug.title(), "slug": slug},
    )
    assert response.status_code == 201
    return response.json()["id"]


def test_rejects_tenant_access_without_membership() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")

    response = client.get(
        "/api/v1/parameters",
        headers={"X-User-Id": USER_B, "X-Tenant-Id": tenant_a},
    )

    assert response.status_code == 403


def test_lists_only_data_for_active_tenant() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")

    created = client.post(
        "/api/v1/parameters",
        headers={"X-User-Id": USER_A, "X-Tenant-Id": tenant_a},
        json={"code": "active_content", "name": "Active Content", "unit": "% p/p"},
    )
    assert created.status_code == 201

    response_a = client.get(
        "/api/v1/parameters",
        headers={"X-User-Id": USER_A, "X-Tenant-Id": tenant_a},
    )
    response_b = client.get(
        "/api/v1/parameters",
        headers={"X-User-Id": USER_B, "X-Tenant-Id": tenant_b},
    )

    assert len(response_a.json()) == 1
    assert response_b.json() == []


def test_raw_material_from_another_tenant_is_not_found() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")

    created = client.post(
        "/api/v1/raw-materials",
        headers={"X-User-Id": USER_A, "X-Tenant-Id": tenant_a},
        json={"name": "Active A", "code": "ACT-A"},
    )
    raw_material_id = created.json()["id"]

    response = client.get(
        f"/api/v1/raw-materials/{raw_material_id}",
        headers={"X-User-Id": USER_B, "X-Tenant-Id": tenant_b},
    )

    assert response.status_code == 404


def test_raw_material_aliases_are_tenant_scoped() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    headers_a = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_a}
    headers_b = {"X-User-Id": USER_B, "X-Tenant-Id": tenant_b}
    raw_material = client.post(
        "/api/v1/raw-materials",
        headers=headers_a,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()

    created = client.post(
        f"/api/v1/raw-materials/{raw_material['id']}/aliases",
        headers=headers_a,
        json={"alias": "Active Alpha"},
    )
    listed = client.get(
        f"/api/v1/raw-materials/{raw_material['id']}/aliases",
        headers=headers_a,
    )
    forbidden = client.get(
        f"/api/v1/raw-materials/{raw_material['id']}/aliases",
        headers=headers_b,
    )

    assert created.status_code == 201
    assert created.json()["normalized_alias"] == "active alpha"
    assert listed.status_code == 200
    assert listed.json()[0]["alias"] == "Active Alpha"
    assert forbidden.status_code == 404


def test_raw_material_prices_are_tenant_scoped_and_ordered() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    headers_a = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_a}
    headers_b = {"X-User-Id": USER_B, "X-Tenant-Id": tenant_b}
    raw_material = client.post(
        "/api/v1/raw-materials",
        headers=headers_a,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()

    first = client.post(
        f"/api/v1/raw-materials/{raw_material['id']}/prices",
        headers=headers_a,
        json={"price": 1.2, "currency": "EUR", "unit": "kg", "valid_from": "2026-01-01"},
    )
    second = client.post(
        f"/api/v1/raw-materials/{raw_material['id']}/prices",
        headers=headers_a,
        json={"price": 1.8, "currency": "EUR", "unit": "kg", "valid_from": "2026-03-01"},
    )
    listed = client.get(
        f"/api/v1/raw-materials/{raw_material['id']}/prices",
        headers=headers_a,
    )
    forbidden = client.get(
        f"/api/v1/raw-materials/{raw_material['id']}/prices",
        headers=headers_b,
    )

    assert first.status_code == 201
    assert second.status_code == 201
    assert listed.status_code == 200
    assert [price["price"] for price in listed.json()] == [1.8, 1.2]
    assert [price["valid_from"] for price in listed.json()] == ["2026-03-01", "2026-01-01"]
    assert forbidden.status_code == 404


def test_persisted_formula_calculation_uses_backend_core() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}

    parameter = client.post(
        "/api/v1/parameters",
        headers=headers,
        json={"code": "active_content", "name": "Active Content", "unit": "% p/p"},
    ).json()
    rm_1 = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    rm_2 = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()

    for raw_material, price, parameter_value in [
        (rm_1, 2.0, 40),
        (rm_2, 1.0, 10),
    ]:
        price_response = client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/prices",
            headers=headers,
            json={"price": price, "currency": "EUR", "unit": "kg"},
        )
        assert price_response.status_code == 201
        value_response = client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/parameter-values",
            headers=headers,
            json={"parameter_id": parameter["id"], "value": parameter_value},
        )
        assert value_response.status_code == 201

    formula = client.post(
        "/api/v1/formulas",
        headers=headers,
        json={
            "name": "Test Formula",
            "items": [
                {"raw_material_id": rm_1["id"], "percentage": 25},
                {"raw_material_id": rm_2["id"], "percentage": 75},
            ],
        },
    ).json()

    response = client.post(
        f"/api/v1/formulas/{formula['id']}/calculate",
        headers=headers,
    )

    assert response.status_code == 200
    result = response.json()
    assert result["total_percentage"] == 100
    assert result["price_total"] == 1.25
    assert result["parameters"] == [
        {"code": "active_content", "value": 17.5, "unit": "% p/p"}
    ]
    assert result["warnings"] == []


def test_persisted_formula_calculation_uses_latest_price_history() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    raw_material = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    for price, valid_from in [(1.0, "2026-01-01"), (2.5, "2026-04-01")]:
        created = client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/prices",
            headers=headers,
            json={"price": price, "currency": "EUR", "unit": "kg", "valid_from": valid_from},
        )
        assert created.status_code == 201
    formula = client.post(
        "/api/v1/formulas",
        headers=headers,
        json={
            "name": "Latest Price Formula",
            "items": [{"raw_material_id": raw_material["id"], "percentage": 100}],
        },
    ).json()

    response = client.post(
        f"/api/v1/formulas/{formula['id']}/calculate",
        headers=headers,
    )

    assert response.status_code == 200
    assert response.json()["price_total"] == 2.5


def test_formula_items_keep_saved_order() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    active = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    carrier = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()
    formula = client.post(
        "/api/v1/formulas",
        headers=headers,
        json={
            "name": "Ordered Formula",
            "items": [
                {"raw_material_id": carrier["id"], "percentage": 75, "order_index": 0},
                {"raw_material_id": active["id"], "percentage": 25, "order_index": 1},
            ],
        },
    ).json()

    patched = client.patch(
        f"/api/v1/formulas/{formula['id']}",
        headers=headers,
        json={
            "items": [
                {"raw_material_id": active["id"], "percentage": 25, "order_index": 0},
                {"raw_material_id": carrier["id"], "percentage": 75, "order_index": 1},
            ],
        },
    )
    reopened = client.get(f"/api/v1/formulas/{formula['id']}", headers=headers)

    assert patched.status_code == 200
    assert [item["raw_material_id"] for item in patched.json()["items"]] == [
        active["id"],
        carrier["id"],
    ]
    assert reopened.status_code == 200
    assert [item["raw_material_id"] for item in reopened.json()["items"]] == [
        active["id"],
        carrier["id"],
    ]


def test_formula_persists_objective() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    material = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()

    created = client.post(
        "/api/v1/formulas",
        headers=headers,
        json={
            "name": "Optimized Formula",
            "objective": "minimize_price",
            "items": [{"raw_material_id": material["id"], "percentage": 100}],
        },
    )

    assert created.status_code == 201
    assert created.json()["objective"] == "minimize_price"

    manual = client.post(
        "/api/v1/formulas",
        headers=headers,
        json={
            "name": "Manual Formula",
            "items": [{"raw_material_id": material["id"], "percentage": 100}],
        },
    )

    assert manual.status_code == 201
    assert manual.json()["objective"] is None


def test_formula_comparison_is_tenant_scoped_and_returns_deltas() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    headers_a = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_a}
    headers_b = {"X-User-Id": USER_B, "X-Tenant-Id": tenant_b}
    parameter = client.post(
        "/api/v1/parameters",
        headers=headers_a,
        json={"code": "active_content", "name": "Active Content", "unit": "% p/p"},
    ).json()
    active = client.post(
        "/api/v1/raw-materials",
        headers=headers_a,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    carrier = client.post(
        "/api/v1/raw-materials",
        headers=headers_a,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()
    for raw_material, price, parameter_value in [
        (active, 2.0, 40),
        (carrier, 1.0, 10),
    ]:
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/prices",
            headers=headers_a,
            json={"price": price, "currency": "EUR", "unit": "kg"},
        )
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/parameter-values",
            headers=headers_a,
            json={"parameter_id": parameter["id"], "value": parameter_value},
        )
    left = client.post(
        "/api/v1/formulas",
        headers=headers_a,
        json={
            "name": "Premium Formula",
            "items": [{"raw_material_id": active["id"], "percentage": 100}],
        },
    ).json()
    right = client.post(
        "/api/v1/formulas",
        headers=headers_a,
        json={
            "name": "Balanced Formula",
            "items": [
                {"raw_material_id": active["id"], "percentage": 50},
                {"raw_material_id": carrier["id"], "percentage": 50},
            ],
        },
    ).json()

    response = client.post(
        "/api/v1/formulas/compare",
        headers=headers_a,
        json={"left_formula_id": left["id"], "right_formula_id": right["id"]},
    )
    forbidden = client.post(
        "/api/v1/formulas/compare",
        headers=headers_b,
        json={"left_formula_id": left["id"], "right_formula_id": right["id"]},
    )

    assert response.status_code == 200
    comparison = response.json()
    assert comparison["left"]["price_total"] == 2.0
    assert comparison["right"]["price_total"] == 1.5
    assert comparison["delta"]["price_total"] == -0.5
    assert comparison["delta"]["total_percentage"] == 0
    assert comparison["delta"]["parameters"] == [
        {
            "code": "active_content",
            "left_value": 40,
            "right_value": 25,
            "delta": -15,
            "unit": "% p/p",
        }
    ]
    assert forbidden.status_code == 404


def test_requirement_parser_uses_tenant_context() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}

    response = client.post(
        "/api/v1/requirements/parse",
        headers=headers,
        json={
            "text": "Minimiza coste con active content entre 20 y 40 y precio maximo 2 EUR/kg",
            "active_parameter_code": "active_content",
            "active_parameter_name": "Active content",
        },
    )

    assert response.status_code == 200
    parsed = response.json()
    assert parsed["tenant_id"] == tenant_id
    assert parsed["user_id"] == USER_A
    assert parsed["source"] == "deterministic"
    assert parsed["objectives"] == [{"type": "minimize", "target": "price"}]
    assert parsed["parameter_bounds"] == [
        {
            "code": "active_content",
            "min_value": 20.0,
            "max_value": 40.0,
            "source_text": "active content entre 20 y 40",
        }
    ]
    assert parsed["price_constraint"]["max_price"] == 2.0


def test_requirement_parser_rejects_cross_tenant_access() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")

    response = client.post(
        "/api/v1/requirements/parse",
        headers={"X-User-Id": USER_B, "X-Tenant-Id": tenant_id},
        json={"text": "Minimiza coste con active content minimo 20"},
    )

    assert response.status_code == 403


def test_requirement_parser_logs_ai_run_and_tool_call() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}

    parsed = client.post(
        "/api/v1/requirements/parse",
        headers=headers,
        json={"text": "Minimiza coste con active content entre 20 y 40"},
    )
    runs = client.get("/api/v1/ai/runs", headers=headers)

    assert parsed.status_code == 200
    assert runs.status_code == 200
    listed_runs = runs.json()
    assert len(listed_runs) == 1
    run = listed_runs[0]
    assert run["tenant_id"] == tenant_id
    assert run["user_id"] == USER_A
    assert run["run_type"] == "requirement_parser"
    assert run["provider"] == "deterministic"
    assert run["model"] == "rules:v1"
    assert run["status"] == "success"
    assert run["input_json"]["text"] == "Minimiza coste con active content entre 20 y 40"
    assert run["output_json"]["source"] == "deterministic"
    assert run["finished_at"] is not None

    detail = client.get(f"/api/v1/ai/runs/{run['id']}", headers=headers)

    assert detail.status_code == 200
    detailed_run = detail.json()
    assert len(detailed_run["tool_calls"]) == 1
    tool_call = detailed_run["tool_calls"][0]
    assert tool_call["tenant_id"] == tenant_id
    assert tool_call["ai_run_id"] == run["id"]
    assert tool_call["tool_name"] == "RequirementParserTool"
    assert tool_call["status"] == "success"
    assert tool_call["output_json"]["parameter_bounds"][0]["min_value"] == 20.0


def test_ai_runs_are_tenant_scoped() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    headers_a = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_a}
    headers_b = {"X-User-Id": USER_B, "X-Tenant-Id": tenant_b}

    client.post(
        "/api/v1/requirements/parse",
        headers=headers_a,
        json={"text": "Minimiza coste con active content minimo 20"},
    )
    runs_a = client.get("/api/v1/ai/runs", headers=headers_a).json()
    runs_b = client.get("/api/v1/ai/runs", headers=headers_b)
    forbidden_detail = client.get(f"/api/v1/ai/runs/{runs_a[0]['id']}", headers=headers_b)

    assert len(runs_a) == 1
    assert runs_b.status_code == 200
    assert runs_b.json() == []
    assert forbidden_detail.status_code == 404


def test_optimization_validation_accepts_minimize_price_contract() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    parameter = client.post(
        "/api/v1/parameters",
        headers=headers,
        json={"code": "active_content", "name": "Active Content", "unit": "% p/p"},
    ).json()
    active = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    carrier = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()

    response = client.post(
        "/api/v1/optimizations/validate",
        headers=headers,
        json={
            "objective": "minimize_price",
            "candidate_raw_material_ids": [active["id"], carrier["id"]],
            "raw_material_bounds": [
                {
                    "raw_material_id": active["id"],
                    "min_percentage": 10,
                    "max_percentage": 80,
                }
            ],
            "parameter_bounds": [
                {
                    "code": parameter["code"],
                    "min_value": 20,
                    "max_value": 40,
                }
            ],
        },
    )

    assert response.status_code == 200
    assert response.json() == {
        "status": "valid",
        "objective": "minimize_price",
        "candidate_count": 2,
        "raw_material_bound_count": 1,
        "parameter_bound_count": 1,
        "issues": [],
    }


def test_optimization_validation_does_not_accept_other_tenant_materials() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    headers_a = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_a}
    headers_b = {"X-User-Id": USER_B, "X-Tenant-Id": tenant_b}
    foreign_material = client.post(
        "/api/v1/raw-materials",
        headers=headers_b,
        json={"name": "Foreign Active", "code": "FOR-A"},
    ).json()

    response = client.post(
        "/api/v1/optimizations/validate",
        headers=headers_a,
        json={
            "objective": "minimize_price",
            "candidate_raw_material_ids": [foreign_material["id"]],
        },
    )

    assert response.status_code == 200
    assert response.json()["status"] == "invalid"
    assert response.json()["issues"] == [
        {
            "code": "candidate_not_found",
            "target": foreign_material["id"],
            "message": "Candidate raw material was not found for the active tenant",
        }
    ]


def test_optimization_validation_reports_incoherent_ranges() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    parameter = client.post(
        "/api/v1/parameters",
        headers=headers,
        json={"code": "active_content", "name": "Active Content", "unit": "% p/p"},
    ).json()
    active = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()

    response = client.post(
        "/api/v1/optimizations/validate",
        headers=headers,
        json={
            "objective": "minimize_price",
            "candidate_raw_material_ids": [active["id"]],
            "raw_material_bounds": [
                {
                    "raw_material_id": active["id"],
                    "min_percentage": 80,
                    "max_percentage": 20,
                }
            ],
            "parameter_bounds": [
                {
                    "code": parameter["code"],
                    "min_value": 40,
                    "max_value": 20,
                }
            ],
        },
    )

    assert response.status_code == 200
    assert response.json()["status"] == "invalid"
    assert [issue["code"] for issue in response.json()["issues"]] == [
        "raw_material_range_invalid",
        "parameter_range_invalid",
    ]


def test_optimization_run_returns_lowest_price_candidate_formula() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    parameter = client.post(
        "/api/v1/parameters",
        headers=headers,
        json={"code": "active_content", "name": "Active Content", "unit": "% p/p"},
    ).json()
    active = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    carrier = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()
    for raw_material, price, parameter_value in [
        (active, 4.0, 50),
        (carrier, 1.0, 0),
    ]:
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/prices",
            headers=headers,
            json={"price": price, "currency": "EUR", "unit": "kg"},
        )
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/parameter-values",
            headers=headers,
            json={"parameter_id": parameter["id"], "value": parameter_value},
        )

    response = client.post(
        "/api/v1/optimizations/run",
        headers=headers,
        json={
            "objective": "minimize_price",
            "candidate_raw_material_ids": [active["id"], carrier["id"]],
            "parameter_bounds": [{"code": "active_content", "min_value": 20}],
        },
    )

    assert response.status_code == 200
    result = response.json()
    assert result["id"]
    assert result["created_at"]
    assert result["status"] == "success"
    assert result["items"] == [
        {"raw_material_id": active["id"], "percentage": 40.0},
        {"raw_material_id": carrier["id"], "percentage": 60.0},
    ]
    assert result["calculation"]["price_total"] == 2.2
    assert result["calculation"]["parameters"] == [
        {"code": "active_content", "value": 20.0, "unit": "% p/p"}
    ]
    assert result["issues"] == []

    runs = client.get("/api/v1/optimizations/runs", headers=headers)
    assert runs.status_code == 200
    assert len(runs.json()) == 1
    run = runs.json()[0]
    assert run["id"] == result["id"]
    assert run["tenant_id"] == tenant_id
    assert run["user_id"] == USER_A
    assert run["formula_id"] is None
    assert run["status"] == "success"
    assert run["objective"] == "minimize_price"
    assert run["request_json"]["parameter_bounds"] == [
        {"code": "active_content", "min_value": 20.0, "max_value": None}
    ]
    assert run["result_json"]["calculation"]["price_total"] == 2.2


def test_optimization_run_is_tenant_scoped() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    headers_a = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_a}
    headers_b = {"X-User-Id": USER_B, "X-Tenant-Id": tenant_b}
    foreign_material = client.post(
        "/api/v1/raw-materials",
        headers=headers_b,
        json={"name": "Foreign Active", "code": "FOR-A"},
    ).json()

    response = client.post(
        "/api/v1/optimizations/run",
        headers=headers_a,
        json={
            "objective": "minimize_price",
            "candidate_raw_material_ids": [foreign_material["id"]],
        },
    )

    assert response.status_code == 200
    assert response.json()["status"] == "invalid"
    assert response.json()["items"] == []
    assert response.json()["issues"][0]["code"] == "candidate_not_found"

    runs_a = client.get("/api/v1/optimizations/runs", headers=headers_a)
    runs_b = client.get("/api/v1/optimizations/runs", headers=headers_b)

    assert runs_a.status_code == 200
    assert runs_b.status_code == 200
    assert [run["status"] for run in runs_a.json()] == ["invalid"]
    assert runs_a.json()[0]["request_json"]["candidate_raw_material_ids"] == [
        foreign_material["id"]
    ]
    assert runs_b.json() == []


def test_optimized_formula_can_link_to_source_run() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    parameter = client.post(
        "/api/v1/parameters",
        headers=headers,
        json={"code": "active_content", "name": "Active Content", "unit": "% p/p"},
    ).json()
    active = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    carrier = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()
    for raw_material, price, parameter_value in [
        (active, 4.0, 50),
        (carrier, 1.0, 0),
    ]:
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/prices",
            headers=headers,
            json={"price": price, "currency": "EUR", "unit": "kg"},
        )
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/parameter-values",
            headers=headers,
            json={"parameter_id": parameter["id"], "value": parameter_value},
        )

    optimization = client.post(
        "/api/v1/optimizations/run",
        headers=headers,
        json={
            "objective": "minimize_price",
            "candidate_raw_material_ids": [active["id"], carrier["id"]],
            "parameter_bounds": [{"code": "active_content", "min_value": 20}],
        },
    ).json()

    formula = client.post(
        "/api/v1/formulas",
        headers=headers,
        json={
            "name": "Optimized Low Cost Formula",
            "objective": "minimize_price",
            "optimization_run_id": optimization["id"],
            "items": [
                {
                    "raw_material_id": item["raw_material_id"],
                    "percentage": item["percentage"],
                    "order_index": index,
                }
                for index, item in enumerate(optimization["items"])
            ],
        },
    )

    assert formula.status_code == 201
    runs = client.get("/api/v1/optimizations/runs", headers=headers).json()
    assert runs[0]["formula_id"] == formula.json()["id"]


def test_optimization_run_reports_infeasible_problem() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    material_a = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Material A", "code": "MAT-A"},
    ).json()
    material_b = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Material B", "code": "MAT-B"},
    ).json()
    for raw_material in [material_a, material_b]:
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/prices",
            headers=headers,
            json={"price": 1.0, "currency": "EUR", "unit": "kg"},
        )

    response = client.post(
        "/api/v1/optimizations/run",
        headers=headers,
        json={
            "objective": "minimize_price",
            "candidate_raw_material_ids": [material_a["id"], material_b["id"]],
            "raw_material_bounds": [
                {"raw_material_id": material_a["id"], "max_percentage": 40},
                {"raw_material_id": material_b["id"], "max_percentage": 40},
            ],
        },
    )

    assert response.status_code == 200
    assert response.json()["status"] == "infeasible"
    assert response.json()["items"] == []
    assert response.json()["calculation"] is None
    assert (
        "Raw material maximum percentages total 80%, below 100%."
        in response.json()["messages"]
    )
    runs = client.get("/api/v1/optimizations/runs", headers=headers)
    assert runs.status_code == 200
    assert runs.json()[0]["status"] == "infeasible"
    assert (
        "Raw material maximum percentages total 80%, below 100%."
        in runs.json()[0]["result_json"]["messages"]
    )


def test_persisted_formula_requires_active_tenant_parameters() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}

    parameter = client.post(
        "/api/v1/parameters",
        headers=headers,
        json={"code": "viscosity", "name": "Viscosity", "unit": "cP"},
    )
    assert parameter.status_code == 201
    raw_material = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Base A", "code": "BASE-A"},
    ).json()
    price_response = client.post(
        f"/api/v1/raw-materials/{raw_material['id']}/prices",
        headers=headers,
        json={"price": 1.5, "currency": "EUR", "unit": "kg"},
    )
    assert price_response.status_code == 201
    formula = client.post(
        "/api/v1/formulas",
        headers=headers,
        json={
            "name": "Missing Parameter Formula",
            "items": [{"raw_material_id": raw_material["id"], "percentage": 100}],
        },
    ).json()

    response = client.post(
        f"/api/v1/formulas/{formula['id']}/calculate",
        headers=headers,
    )

    assert response.status_code == 200
    warnings = response.json()["warnings"]
    assert warnings[0]["code"] == "missing_parameter"
    assert warnings[0]["parameter_code"] == "viscosity"


def test_formula_calculation_history_is_tenant_scoped() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    headers_a = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_a}
    headers_b = {"X-User-Id": USER_B, "X-Tenant-Id": tenant_b}
    parameter = client.post(
        "/api/v1/parameters",
        headers=headers_a,
        json={"code": "active_content", "name": "Active Content", "unit": "% p/p"},
    ).json()
    rm_1 = client.post(
        "/api/v1/raw-materials",
        headers=headers_a,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    rm_2 = client.post(
        "/api/v1/raw-materials",
        headers=headers_a,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()
    for raw_material, price, parameter_value in [
        (rm_1, 2.0, 40),
        (rm_2, 1.0, 10),
    ]:
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/prices",
            headers=headers_a,
            json={"price": price, "currency": "EUR", "unit": "kg"},
        )
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/parameter-values",
            headers=headers_a,
            json={"parameter_id": parameter["id"], "value": parameter_value},
        )
    formula = client.post(
        "/api/v1/formulas",
        headers=headers_a,
        json={
            "name": "History Formula",
            "items": [
                {"raw_material_id": rm_1["id"], "percentage": 25},
                {"raw_material_id": rm_2["id"], "percentage": 75},
            ],
        },
    ).json()

    first = client.post(f"/api/v1/formulas/{formula['id']}/calculate", headers=headers_a)
    second = client.post(f"/api/v1/formulas/{formula['id']}/calculate", headers=headers_a)
    assert first.status_code == 200
    assert second.status_code == 200
    history = client.get(
        f"/api/v1/formulas/{formula['id']}/calculations",
        headers=headers_a,
    )
    forbidden = client.get(
        f"/api/v1/formulas/{formula['id']}/calculations",
        headers=headers_b,
    )

    assert history.status_code == 200
    assert len(history.json()) == 2
    assert history.json()[0]["result_json"]["price_total"] == 1.25
    assert forbidden.status_code == 404


def test_formula_excel_export_is_tenant_scoped_and_contains_workbook() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    headers_a = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_a}
    headers_b = {"X-User-Id": USER_B, "X-Tenant-Id": tenant_b}

    parameter = client.post(
        "/api/v1/parameters",
        headers=headers_a,
        json={"code": "active_content", "name": "Active Content", "unit": "% p/p"},
    ).json()
    active = client.post(
        "/api/v1/raw-materials",
        headers=headers_a,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    carrier = client.post(
        "/api/v1/raw-materials",
        headers=headers_a,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()
    for raw_material, price, parameter_value in [
        (active, 2.0, 40),
        (carrier, 1.0, 10),
    ]:
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/prices",
            headers=headers_a,
            json={"price": price, "currency": "EUR", "unit": "kg"},
        )
        client.post(
            f"/api/v1/raw-materials/{raw_material['id']}/parameter-values",
            headers=headers_a,
            json={"parameter_id": parameter["id"], "value": parameter_value},
        )
    formula = client.post(
        "/api/v1/formulas",
        headers=headers_a,
        json={
            "name": "Export Formula",
            "items": [
                {"raw_material_id": active["id"], "percentage": 25},
                {"raw_material_id": carrier["id"], "percentage": 75},
            ],
        },
    ).json()

    response = client.get(
        f"/api/v1/formulas/{formula['id']}/export/excel",
        headers=headers_a,
    )
    forbidden = client.get(
        f"/api/v1/formulas/{formula['id']}/export/excel",
        headers=headers_b,
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="Export_Formula.xlsx"' in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["Summary", "Lines", "Parameters", "Warnings"]
    assert workbook["Summary"]["B2"].value == "Export Formula"
    assert workbook["Summary"]["B6"].value == 100
    assert workbook["Summary"]["B7"].value == 1.25
    assert workbook["Lines"]["C2"].value == "Active A"
    assert workbook["Lines"]["D2"].value == 25
    assert workbook["Lines"]["G2"].value == 0.5
    assert workbook["Parameters"]["A2"].value == "active_content"
    assert workbook["Parameters"]["B2"].value == 17.5
    assert forbidden.status_code == 404


def test_unknown_raw_material_in_ad_hoc_calculation_returns_warning() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")

    response = client.post(
        "/api/v1/formulas/calculate",
        headers={"X-User-Id": USER_A, "X-Tenant-Id": tenant_id},
        json={
            "items": [
                {"raw_material_id": str(uuid.uuid4()), "percentage": 100},
            ]
        },
    )

    assert response.status_code == 200
    assert response.json()["warnings"][0]["code"] == "missing_raw_material"
