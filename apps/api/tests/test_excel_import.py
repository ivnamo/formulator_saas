from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel, create_engine

from formulia_api.excel_import import list_formula_xlsx_sheets, parse_formula_xlsx
from formulia_api.main import create_app


USER_A = "10000000-0000-0000-0000-000000000001"
USER_B = "20000000-0000-0000-0000-000000000001"


def make_client() -> TestClient:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    app = create_app(engine)
    SQLModel.metadata.create_all(engine)
    return TestClient(app)


def create_tenant(client: TestClient, user_id: str, slug: str) -> str:
    response = client.post(
        "/api/v1/tenants",
        headers={"X-User-Id": user_id},
        json={"name": slug.title(), "slug": slug},
    )
    assert response.status_code == 201
    return response.json()["id"]


def workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def workbook_with_sheets(sheets: dict[str, list[list[object]]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets.items():
        worksheet = workbook.create_sheet(title=title)
        for row in rows:
            worksheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def atlantica_workbook_bytes() -> bytes:
    workbook = Workbook()
    calculator = workbook.active
    calculator.title = "Calculadora"
    calculator.append(["Materia Prima", "%", "Precio", "Ntotal", "K2O"])
    calculator.append(["AGUA", 50, 0, 0, 0])
    calculator.append(["GLICINA", 50, 1, 18, 0])
    calculator.append([])
    calculator.append(["TOTAL", "=SUM(B2:B4)", "=SUMPRODUCT($B$2:$B$3,C2:C3)/100"])

    lab = workbook.create_sheet("Hoja Lab")
    lab["F2"] = "FECHA: 02-10-2025"
    lab["C3"] = "MICROCAT BON + AA - MUESTRA: F1"
    lab["A5"] = "Órden de adición"
    lab["D5"] = "Cantidad % peso"
    lab["E5"] = "ENSAYO 0,5 kg"
    lab["F5"] = "ENSAYO 2 kg"
    lab["A6"] = 1
    lab["C6"] = "AGUA INICIAL"
    lab["D6"] = 50
    lab["E6"] = "=D6*10/2"
    lab["F6"] = "=D6*10*2"
    lab["A7"] = 2
    lab["C7"] = "GLICINA"
    lab["D7"] = 50
    lab["E7"] = "=D7*10/2"
    lab["F7"] = "=D7*10*2"
    lab["G7"] = "transparente"
    lab["D9"] = "=SUM(D6:D7)"
    lab["E9"] = "=SUM(E6:E7)"
    lab["F9"] = "=SUM(F6:F7)"
    lab["C11"] = "EXPERIMENTAL 02/10/2025"
    lab["D11"] = "Densidad"
    lab["E11"] = 1.01592
    lab["D12"] = "pH"
    lab["E12"] = 10.48

    composition = workbook.create_sheet("Composición")
    composition.append([])
    composition.append(["Parámetro", "% p/p"])
    composition.append(["Precio", 0.5])
    composition.append(["Ntotal", 9])
    composition.append(["K2O", 0])

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def compact_lab_trial_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Aminocat 30%"
    worksheet["C2"] = "Lote"
    worksheet["E2"] = "Fecha 24/05/24"
    worksheet["C5"] = "AMINOCAT 30% -  NORLAN 6/24"
    worksheet.merge_cells("C5:E5")
    worksheet["D7"] = "Cantidad % peso"
    worksheet["E7"] = "ensayo 2 Kg"
    rows = [
        ("Agua", 2.14),
        ("EDTA(Na)4", 0.06),
        ("Hidróxido Potásico", 0.3),
        ("Glicina", 1),
        ("AA 24% Sintetico", 25),
        ("NORLAN 6/24", 39.5),
        ("AA 27%", 20),
        ("Glutamato monosodico", 10.8),
        ("Nitrato amónico", 1),
        ("Biopol KF78", 0.2),
    ]
    for offset, (name, percentage) in enumerate(rows, start=8):
        worksheet.cell(row=offset, column=3).value = name
        worksheet.cell(row=offset, column=4).value = percentage
        worksheet.cell(row=offset, column=5).value = f"=D{offset}*20"
    worksheet["D19"] = "=SUM(D8:D18)"
    worksheet["E19"] = "=SUM(E8:E17)"
    worksheet["C21"] = "Densidad Experimental"
    worksheet["D21"] = 1.216
    worksheet["C22"] = "pH"
    worksheet["D22"] = 7.11

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def compact_lab_trial_with_addition_order_workbook_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "F1"
    worksheet["D2"] = "FECHA: 06/05/2026"
    worksheet["C3"] = "ID.608 - PROTIUM FULVIC EG - F2 - CALIDAD"
    worksheet["A5"] = "Orden de adicion"
    worksheet["D5"] = "Cantidad % peso"
    worksheet["E5"] = "ENSAYO 2000g"
    rows = [
        ("AGUA", 32.69),
        ("EDTA TETRASODICO", 0.5),
        ("BIOPOL KF78", 0.2),
        ("PROTE FOAM ESS", 0.01),
        ("SULFATO DE ZINC MONOHIDRATO", 5),
        ("SULFATO DE MANGANESO", 5),
        ("DAVE 25L", 45),
        ("POLISACARIDOS", 10),
        ("MONOPROPILENGLICOL", 1.5),
        ("GOMA XANTANA", 0.1),
    ]
    for order, (name, percentage) in enumerate(rows, start=1):
        row_number = order + 5
        worksheet.cell(row=row_number, column=1).value = order
        worksheet.cell(row=row_number, column=3).value = name
        worksheet.cell(row=row_number, column=4).value = percentage
        worksheet.cell(row=row_number, column=5).value = f"=D{row_number}*20"
    worksheet["D17"] = "=SUM(D6:D15)"
    worksheet["E17"] = "=SUM(E6:E15)"
    worksheet["C19"] = "EXPERIMENTAL 06/05/2026"
    worksheet["D19"] = "Densidad"
    worksheet["E19"] = 1.20764
    worksheet["D20"] = "pH"
    worksheet["E20"] = 3.65

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_parser_detects_columns_and_percent_rows() -> None:
    content = workbook_bytes(
        [
            ["Code", "Raw Material", "Percentage"],
            ["ACT-A", "Active A", 25],
            ["CAR-B", "Carrier B", 75],
        ]
    )

    parsed = parse_formula_xlsx(content)

    assert parsed.columns.material_code == "code"
    assert parsed.columns.material_name == "raw material"
    assert parsed.columns.percentage == "percentage"
    assert parsed.total_percentage == 100
    assert [row.material_code for row in parsed.rows] == ["ACT-A", "CAR-B"]


def test_parser_detects_atlantica_id_lab_template() -> None:
    parsed = parse_formula_xlsx(atlantica_workbook_bytes())

    assert parsed.parser == "atlantica_id_lab"
    assert parsed.sheet_name == "Calculadora"
    assert parsed.formula_name == "MICROCAT BON + AA - MUESTRA: F1"
    assert parsed.parameter_headers == ["Ntotal", "K2O"]
    assert parsed.total_percentage == 100
    assert parsed.rows[0].material_name == "AGUA"
    assert parsed.rows[0].lab_material_name == "AGUA INICIAL"
    assert parsed.rows[1].price == 1
    assert parsed.rows[1].parameters == {"Ntotal": 18.0, "K2O": 0.0}
    assert parsed.rows[1].lab_observation == "transparente"


def test_parser_detects_compact_lab_trial_template() -> None:
    parsed = parse_formula_xlsx(compact_lab_trial_workbook_bytes())

    assert parsed.parser == "compact_lab_trial"
    assert parsed.sheet_name == "Aminocat 30%"
    assert parsed.formula_name == "AMINOCAT 30% -  NORLAN 6/24"
    assert parsed.columns.material_name == "material"
    assert parsed.columns.percentage == "Cantidad % peso"
    assert parsed.total_percentage == 100
    assert len(parsed.rows) == 10
    assert parsed.rows[0].material_name == "Agua"
    assert parsed.rows[0].percentage == 2.14
    assert parsed.rows[-1].material_name == "Biopol KF78"
    assert parsed.rows[-1].percentage == 0.2
    assert "Densidad Experimental" not in {
        row.material_name for row in parsed.rows
    }


def test_parser_ignores_compact_lab_addition_order_column() -> None:
    parsed = parse_formula_xlsx(compact_lab_trial_with_addition_order_workbook_bytes())

    assert parsed.parser == "compact_lab_trial"
    assert parsed.sheet_name == "F1"
    assert parsed.formula_name == "ID.608 - PROTIUM FULVIC EG - F2 - CALIDAD"
    assert parsed.columns.material_name_index == 2
    assert parsed.columns.percentage_index == 3
    assert parsed.total_percentage == 100
    assert [row.material_name for row in parsed.rows[:3]] == [
        "AGUA",
        "EDTA TETRASODICO",
        "BIOPOL KF78",
    ]
    assert parsed.rows[-1].material_name == "GOMA XANTANA"
    assert parsed.rows[-1].percentage == 0.1


def test_parser_lists_and_selects_worksheets() -> None:
    content = workbook_with_sheets(
        {
            "Notes": [["Comment"], ["Not a formula"]],
            "Formula": [["Code", "Percentage"], ["ACT-A", 100]],
        }
    )

    parsed = parse_formula_xlsx(content, sheet_name="Formula")

    assert list_formula_xlsx_sheets(content) == ["Notes", "Formula"]
    assert parsed.sheet_name == "Formula"
    assert parsed.available_sheets == ["Notes", "Formula"]
    assert parsed.rows[0].material_code == "ACT-A"


def test_preview_matches_by_code_and_normalized_name() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    active = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    carrier = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()
    content = workbook_bytes(
        [
            ["Code", "Raw Material", "Percentage"],
            ["ACT-A", "Active A", 25],
            ["", "Carrier B", 75],
        ]
    )

    response = client.post(
        "/api/v1/imports/formulas/excel/preview",
        headers=headers,
        files={"file": ("formula.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    preview = response.json()
    assert preview["total_percentage"] == 100
    assert preview["resolved_rows"] == 2
    assert preview["pending_rows"] == 0
    assert preview["rows"][0]["raw_material_id"] == active["id"]
    assert preview["rows"][0]["matched_by"] == "code"
    assert preview["rows"][0]["resolved_material_name"] == "Active A"
    assert preview["rows"][1]["raw_material_id"] == carrier["id"]
    assert preview["rows"][1]["matched_by"] == "name"
    assert preview["rows"][1]["resolved_material_name"] == "Carrier B"


def test_preview_does_not_match_obsolete_materials() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    obsolete = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Obsolete A", "code": "OBS-A"},
    ).json()
    alias = client.post(
        f"/api/v1/raw-materials/{obsolete['id']}/aliases",
        headers=headers,
        json={"alias": "Old Alpha"},
    )
    updated = client.patch(
        f"/api/v1/raw-materials/{obsolete['id']}",
        headers=headers,
        json={"is_obsolete": True},
    )
    content = workbook_bytes(
        [
            ["Code", "Raw Material", "Percentage"],
            ["OBS-A", "Obsolete A", 50],
            ["", "Old Alpha", 50],
        ]
    )

    response = client.post(
        "/api/v1/imports/formulas/excel/preview",
        headers=headers,
        files={"file": ("formula.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert alias.status_code == 201
    assert updated.status_code == 200
    assert response.status_code == 200
    preview = response.json()
    assert preview["resolved_rows"] == 0
    assert preview["pending_rows"] == 2
    assert [row["status"] for row in preview["rows"]] == ["needs_review", "needs_review"]
    assert [row["raw_material_id"] for row in preview["rows"]] == [None, None]
    assert [row["suggested_raw_material_id"] for row in preview["rows"]] == [None, None]


def test_preview_exposes_atlantica_template_metadata() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    water = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "AGUA", "code": "WATER"},
    ).json()
    glycine = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "GLICINA", "code": "GLY"},
    ).json()

    response = client.post(
        "/api/v1/imports/formulas/excel/preview",
        headers=headers,
        files={
            "file": (
                "formula.xlsx",
                atlantica_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    preview = response.json()
    assert preview["parser"] == "atlantica_id_lab"
    assert preview["formula_name"] == "MICROCAT BON + AA - MUESTRA: F1"
    assert preview["parameter_headers"] == ["Ntotal", "K2O"]
    assert preview["rows"][0]["raw_material_id"] == water["id"]
    assert preview["rows"][1]["raw_material_id"] == glycine["id"]
    assert preview["rows"][1]["imported_price"] == 1
    assert preview["rows"][1]["imported_parameters"] == {"Ntotal": 18.0, "K2O": 0.0}
    assert preview["rows"][1]["lab_observation"] == "transparente"


def test_preview_matches_by_alias() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    material = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()
    alias = client.post(
        f"/api/v1/raw-materials/{material['id']}/aliases",
        headers=headers,
        json={"alias": "Carrier beta"},
    )
    assert alias.status_code == 201
    content = workbook_bytes(
        [
            ["Raw Material", "Percentage"],
            ["Carrier beta", 100],
        ]
    )

    response = client.post(
        "/api/v1/imports/formulas/excel/preview",
        headers=headers,
        files={"file": ("formula.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    row = response.json()["rows"][0]
    assert row["raw_material_id"] == material["id"]
    assert row["matched_by"] == "alias"
    assert row["material_name"] == "Carrier beta"
    assert row["resolved_material_code"] == "CAR-B"
    assert row["resolved_material_name"] == "Carrier B"


def test_preview_matches_material_code_by_alias() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    material = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Carrier B", "code": "CAR-B"},
    ).json()
    alias = client.post(
        f"/api/v1/raw-materials/{material['id']}/aliases",
        headers=headers,
        json={"alias": "ALT-CAR-B"},
    )
    assert alias.status_code == 201
    content = workbook_bytes(
        [
            ["Code", "Percentage"],
            ["ALT-CAR-B", 100],
        ]
    )

    response = client.post(
        "/api/v1/imports/formulas/excel/preview",
        headers=headers,
        files={"file": ("formula.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    row = response.json()["rows"][0]
    assert row["raw_material_id"] == material["id"]
    assert row["matched_by"] == "alias"
    assert row["material_code"] == "ALT-CAR-B"
    assert row["resolved_material_code"] == "CAR-B"
    assert row["resolved_material_name"] == "Carrier B"


def test_preview_suggests_fuzzy_match_without_resolving() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    material = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Carrier Meta 009", "code": "CAR-009"},
    ).json()
    content = workbook_bytes(
        [
            ["Raw Material", "Percentage"],
            ["Carier Meta 009", 100],
        ]
    )

    response = client.post(
        "/api/v1/imports/formulas/excel/preview",
        headers=headers,
        files={"file": ("formula.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    preview = response.json()
    row = preview["rows"][0]
    assert preview["resolved_rows"] == 0
    assert preview["pending_rows"] == 1
    assert row["status"] == "needs_review"
    assert row["raw_material_id"] is None
    assert row["matched_by"] is None
    assert row["suggested_raw_material_id"] == material["id"]
    assert row["suggested_material_name"] == "Carrier Meta 009"
    assert row["suggested_match_score"] >= 0.82


def test_preview_selected_sheet_in_multi_sheet_workbook() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    material = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()
    content = workbook_with_sheets(
        {
            "Notes": [["Comment"], ["Not a formula"]],
            "Formula": [["Code", "Percentage"], ["ACT-A", 100]],
        }
    )

    sheets = client.post(
        "/api/v1/imports/formulas/excel/sheets",
        headers=headers,
        files={"file": ("formula.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    response = client.post(
        "/api/v1/imports/formulas/excel/preview",
        headers=headers,
        files={"file": ("formula.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_name": "Formula"},
    )

    assert sheets.status_code == 200
    assert sheets.json() == {"sheets": ["Notes", "Formula"], "default_sheet": "Notes"}
    assert response.status_code == 200
    preview = response.json()
    assert preview["sheet_name"] == "Formula"
    assert preview["available_sheets"] == ["Notes", "Formula"]
    assert preview["rows"][0]["raw_material_id"] == material["id"]


def test_preview_rejects_missing_sheet_name() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    content = workbook_with_sheets({"Formula": [["Code", "Percentage"], ["ACT-A", 100]]})

    response = client.post(
        "/api/v1/imports/formulas/excel/preview",
        headers=headers,
        files={"file": ("formula.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_name": "Missing"},
    )

    assert response.status_code == 400
    assert "Worksheet 'Missing' was not found" in response.json()["detail"]


def test_preview_flags_unmatched_and_invalid_rows() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    content = workbook_bytes(
        [
            ["Code", "Raw Material", "Percentage"],
            ["MISSING", "Missing Material", 50],
            ["", "Invalid Material", "abc"],
        ]
    )

    response = client.post(
        "/api/v1/imports/formulas/excel/preview",
        headers=headers,
        files={"file": ("formula.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    rows = response.json()["rows"]
    assert rows[0]["status"] == "needs_review"
    assert rows[1]["status"] == "invalid_percentage"


def test_save_imported_rows_as_formula_and_calculate() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    parameter = client.post(
        "/api/v1/parameters",
        headers=headers,
        json={"code": "active_content", "name": "Active Content", "unit": "% p/p"},
    ).json()
    active = _create_material(client, headers, "Active A", "ACT-A", parameter["id"], 2, 40)
    carrier = _create_material(client, headers, "Carrier B", "CAR-B", parameter["id"], 1, 10)

    response = client.post(
        "/api/v1/imports/formulas/excel/save",
        headers=headers,
        json={
            "name": "Imported Formula",
            "objective": "Imported formula used to verify save and calculation.",
            "rows": [
                {"raw_material_id": active["id"], "percentage": 25},
                {"raw_material_id": carrier["id"], "percentage": 75},
            ],
        },
    )

    assert response.status_code == 201
    formula = response.json()
    assert formula["name"] == "Imported Formula"
    calculation = client.post(
        f"/api/v1/formulas/{formula['id']}/calculate",
        headers=headers,
    )
    assert calculation.status_code == 200
    assert calculation.json()["price_total"] == 1.25
    assert calculation.json()["parameters"][0]["value"] == 17.5


def test_save_import_requires_description() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}

    response = client.post(
        "/api/v1/imports/formulas/excel/save",
        headers=headers,
        json={"name": "No Description Import", "objective": "", "rows": []},
    )

    assert response.status_code == 422


def test_save_import_requires_name() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}

    response = client.post(
        "/api/v1/imports/formulas/excel/save",
        headers=headers,
        json={"name": " ", "objective": "Imported formula description.", "rows": []},
    )

    assert response.status_code == 422


def test_save_import_rejects_negative_percentage() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    headers = {"X-User-Id": USER_A, "X-Tenant-Id": tenant_id}
    raw_material = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": "Active A", "code": "ACT-A"},
    ).json()

    response = client.post(
        "/api/v1/imports/formulas/excel/save",
        headers=headers,
        json={
            "name": "Invalid Import",
            "objective": "Import used to verify negative percentage rejection.",
            "rows": [{"raw_material_id": raw_material["id"], "percentage": -1}],
        },
    )

    assert response.status_code == 422


def test_import_preview_requires_tenant_membership() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    content = workbook_bytes([["Code", "Percentage"], ["ACT-A", 100]])

    response = client.post(
        "/api/v1/imports/formulas/excel/preview",
        headers={"X-User-Id": USER_B, "X-Tenant-Id": tenant_a},
        files={"file": ("formula.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 403


def _create_material(
    client: TestClient,
    headers: dict[str, str],
    name: str,
    code: str,
    parameter_id: str,
    price: float,
    parameter_value: float,
) -> dict[str, object]:
    material = client.post(
        "/api/v1/raw-materials",
        headers=headers,
        json={"name": name, "code": code},
    ).json()
    assert client.post(
        f"/api/v1/raw-materials/{material['id']}/prices",
        headers=headers,
        json={"price": price, "currency": "EUR", "unit": "kg"},
    ).status_code == 201
    assert client.post(
        f"/api/v1/raw-materials/{material['id']}/parameter-values",
        headers=headers,
        json={"parameter_id": parameter_id, "value": parameter_value},
    ).status_code == 201
    return material
