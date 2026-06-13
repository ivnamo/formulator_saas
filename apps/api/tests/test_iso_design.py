import uuid
from io import BytesIO
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel, Session, create_engine, select

from formulia_api.main import create_app
from formulia_api.models import (
    FormulaReviewRequest,
    IsoDesignProject,
    IsoDesignTrial,
    IsoProductValidation,
    TenantMember,
)


USER_A = "10000000-0000-0000-0000-000000000001"
USER_B = "20000000-0000-0000-0000-000000000001"


def make_client() -> TestClient:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    app = create_app(engine)
    SQLModel.metadata.create_all(engine)
    return TestClient(app)


def create_tenant(client: TestClient, user_id: str, slug: str) -> str:
    response = client.post(
        "/api/v1/tenants",
        headers={"X-User-Id": user_id},
        json={"name": slug.title(), "slug": slug},
    )
    assert response.status_code == 201
    return response.json()["id"]


def headers(user_id: str, tenant_id: str) -> dict[str, str]:
    return {"X-User-Id": user_id, "X-Tenant-Id": tenant_id}


def add_member(client: TestClient, tenant_id: str, user_id: str, role: str) -> None:
    client.get("/api/v1/me", headers={"X-User-Id": user_id})
    with Session(client.app.state.engine) as session:
        session.add(
            TenantMember(
                tenant_id=uuid.UUID(tenant_id),
                user_id=uuid.UUID(user_id),
                role=role,
            )
        )
        session.commit()


def enable_iso(client: TestClient, tenant_id: str, user_id: str = USER_A) -> dict:
    response = client.patch(
        "/api/v1/iso/settings",
        headers=headers(user_id, tenant_id),
        json={"enabled": True},
    )
    assert response.status_code == 200
    return response.json()


def create_iso_project(client: TestClient, tenant_id: str) -> dict:
    response = client.post(
        "/api/v1/iso/design-projects",
        headers=headers(USER_A, tenant_id),
        json={
            "iso_request_number": "1/2026",
            "project_code": "FLOWER",
            "product_name": "Flower Power",
            "requester": "Comercial",
            "accepted_status": "accepted",
            "lifecycle_status": "design",
        },
    )
    assert response.status_code == 201
    return response.json()


def create_iso_trial(
    client: TestClient,
    tenant_id: str,
    project_id: str,
    technical_result: str,
    trial_name: str = "F1",
) -> dict:
    response = client.post(
        f"/api/v1/iso/design-projects/{project_id}/trials",
        headers=headers(USER_A, tenant_id),
        json={
            "technical_result": technical_result,
            "trial_name": trial_name,
            "formula_version": 1,
        },
    )
    assert response.status_code == 201
    return response.json()


def workbook_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def upload_legacy_xlsx(
    client: TestClient,
    tenant_id: str,
    endpoint: str,
    content: bytes,
    *,
    user_id: str = USER_A,
    sheet_name: str | None = None,
):
    data = {"sheet_name": sheet_name} if sheet_name else None
    return client.post(
        endpoint,
        headers=headers(user_id, tenant_id),
        files={
            "file": (
                "legacy.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data=data,
    )


def f10_01_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2026"
    sheet.append(
        [
            "No Solicitud",
            "ID",
            "SOLICITANTE",
            "NOMBRE",
            "TIPO DE PRODUCTO",
            "NECESIDAD",
            "ACEPTADO",
            "FINALIZADO",
        ]
    )
    sheet.append(
        [
            "10/2026",
            "AA-10",
            "I+D",
            "Legacy Calcium",
            "Fertilizante",
            "Cliente requiere calcio",
            "SI",
            "NO",
        ]
    )
    sheet.append(["11/2026", "AA-11", "I+D", None, "Fertilizante", "Sin producto", "SI", "NO"])
    return workbook_bytes(workbook)


def f10_02_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "12-2026 Legacy Amino"
    sheet["A1"] = "F10-02 - Diseno de producto"
    sheet["A3"] = "No Solicitud:"
    sheet["B3"] = "12/2026"
    sheet["A4"] = "Responsable:"
    sheet["B4"] = "Ivan"
    sheet["A5"] = "Producto:"
    sheet["B5"] = "Legacy Amino"
    sheet["A7"] = "1. DATOS DE PARTIDA DEL DISENO"
    sheet["A8"] = "Cliente solicita aminoacidos"
    sheet["A10"] = "Ensayo 1"
    sheet.append(["ID", "Nombre", "Fecha", "Resultado"])
    sheet.append(["ID-1", "Legacy Amino F1", "2026-05-01", "NOK"])
    sheet.append(["Materia prima", "% peso"])
    sheet.append(["Materia A", "60"])
    sheet.append(["Materia B", "40"])
    sheet.append(["Motivo / comentario", "No cumple estabilidad"])
    sheet.append([])
    sheet.append(["Ensayo 2"])
    sheet.append(["ID", "Nombre", "Fecha", "Resultado"])
    sheet.append(["ID-2", "Legacy Amino F2", "2026-05-02", "LIBERADO"])
    sheet.append(["Materia prima", "% peso"])
    sheet.append(["Materia A", "55"])
    sheet.append(["Materia C", "45"])
    sheet.append(["Motivo / comentario", "Liberado por calidad"])
    return workbook_bytes(workbook)


def f10_03_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "13-2026 Legacy Zinc"
    sheet["A1"] = "F10-03 - Validacion de producto"
    sheet["A3"] = "No Solicitud:"
    sheet["B3"] = "13/2026"
    sheet["A4"] = "Responsable:"
    sheet["B4"] = "Calidad"
    sheet["A5"] = "Producto:"
    sheet["B5"] = "Legacy Zinc"
    sheet["A6"] = "Formula OK:"
    sheet["B6"] = "ID-777 || Legacy Zinc F3"
    sheet["A8"] = "1. ESPECIFICACION"
    sheet["A9"] = "Descripcion"
    sheet["B9"] = "Producto validado"
    sheet["A10"] = "Aspecto"
    sheet["B10"] = "Liquido"
    sheet["A11"] = "Color"
    sheet["B11"] = "Verde"
    sheet["A13"] = "Parametro"
    sheet["B13"] = "Valor"
    sheet["A14"] = "pH"
    sheet["B14"] = "6.5"
    sheet["A16"] = "2. VALIDACION"
    sheet["A17"] = "Fecha validacion:"
    sheet["B17"] = "2026-06-01"
    sheet["A19"] = "Area"
    sheet["B19"] = "Aspecto a validar"
    sheet["C19"] = "OK/NOK"
    sheet["D19"] = "Comentarios"
    sheet.append(["I+D", "Formula liberada", "OK", "Correcto"])
    sheet.append(["Calidad", "Analitica", "OK", "Correcto"])
    return workbook_bytes(workbook)


def test_iso_settings_are_disabled_by_default_and_gate_endpoints() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")

    settings = client.get("/api/v1/iso/settings", headers=headers(USER_A, tenant_id))
    projects = client.get("/api/v1/iso/design-projects", headers=headers(USER_A, tenant_id))

    assert settings.status_code == 200
    assert settings.json()["enabled"] is False
    assert settings.json()["config"]["formats"]["f10_01"]["enabled"] is True
    assert projects.status_code == 403
    assert projects.json()["detail"] == "ISO 9001 module is not enabled."


def test_only_admin_can_enable_iso_module() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    add_member(client, tenant_id, USER_B, "formulator")

    forbidden = client.patch(
        "/api/v1/iso/settings",
        headers=headers(USER_B, tenant_id),
        json={"enabled": True},
    )
    enabled = client.patch(
        "/api/v1/iso/settings",
        headers=headers(USER_A, tenant_id),
        json={
            "enabled": True,
            "config_patch": {"jira": {"allow_poc_without_project": True}},
        },
    )

    assert forbidden.status_code == 403
    assert enabled.status_code == 200
    assert enabled.json()["enabled"] is True
    assert enabled.json()["config"]["jira"]["allow_poc_without_project"] is True


def test_iso_design_project_crud_and_rejection_rule() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    enable_iso(client, tenant_id)

    missing_reason = client.post(
        "/api/v1/iso/design-projects",
        headers=headers(USER_A, tenant_id),
        json={
            "iso_request_number": "2/2026",
            "product_name": "Rejected Product",
            "accepted_status": "rejected",
        },
    )
    project = create_iso_project(client, tenant_id)
    listed = client.get("/api/v1/iso/design-projects", headers=headers(USER_A, tenant_id))
    duplicate = client.post(
        "/api/v1/iso/design-projects",
        headers=headers(USER_A, tenant_id),
        json={
            "iso_request_number": "1/2026",
            "project_code": "FLOWER",
            "product_name": "Flower Power Duplicate",
        },
    )
    updated = client.patch(
        f"/api/v1/iso/design-projects/{project['id']}",
        headers=headers(USER_A, tenant_id),
        json={"comments": "Planificado para Q3", "rd_hours": 4.5},
    )

    assert missing_reason.status_code == 400
    assert "Rejection reason" in missing_reason.json()["detail"]
    assert project["year"] == 2026
    assert project["trial_count"] == 0
    assert listed.status_code == 200
    assert [item["id"] for item in listed.json()] == [project["id"]]
    assert duplicate.status_code == 409
    assert updated.status_code == 200
    assert updated.json()["comments"] == "Planificado para Q3"
    assert updated.json()["rd_hours"] == 4.5


def test_iso_design_trial_can_be_created_from_jira_review_snapshot() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    enable_iso(client, tenant_id)
    project = create_iso_project(client, tenant_id)
    review_id = uuid.uuid4()
    formula_id = uuid.uuid4()
    with Session(client.app.state.engine) as session:
        session.add(
            FormulaReviewRequest(
                id=review_id,
                tenant_id=uuid.UUID(tenant_id),
                formula_id=formula_id,
                formula_version=3,
                jira_connection_id=uuid.uuid4(),
                review_status="closed",
                jira_issue_key="ID-665",
                jira_issue_url="https://example.atlassian.net/browse/ID-665",
                jira_status="FINALIZADO",
                sent_by_user_id=uuid.UUID(USER_A),
                snapshot_json={
                    "formula": {"name": "Flower Power", "jira_issue_type": "Calidad"},
                    "jira": {
                        "issue_summary": "CALIDAD - FLOWER - Flower Power",
                        "technical_result_raw": "NOK técnico",
                    },
                },
            )
        )
        session.commit()

    created = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/trials/from-jira-review",
        headers=headers(USER_A, tenant_id),
        json={"review_id": str(review_id), "trial_number": 1},
    )
    repeated = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/trials/from-jira-review",
        headers=headers(USER_A, tenant_id),
        json={"review_id": str(review_id), "trial_number": 1},
    )
    trials = client.get(
        f"/api/v1/iso/design-projects/{project['id']}/trials",
        headers=headers(USER_A, tenant_id),
    )

    assert created.status_code == 201
    trial = created.json()
    assert trial["technical_result"] == "NOK"
    assert trial["raw_result_label"] == "NOK técnico"
    assert trial["raw_status_label"] == "FINALIZADO"
    assert trial["trial_code"] == "ID-665"
    assert trial["formula_version"] == 3
    assert trial["snapshot_checksum"]
    assert repeated.status_code == 201
    assert repeated.json()["id"] == trial["id"]
    assert trials.status_code == 200
    assert [item["id"] for item in trials.json()] == [trial["id"]]


def test_iso_technical_results_are_configurable_per_tenant() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    enable_iso(client, tenant_id)
    client.patch(
        "/api/v1/iso/settings",
        headers=headers(USER_A, tenant_id),
        json={"config_patch": {"technical_results": ["CUSTOM_OK", "pending_result"]}},
    )
    project = create_iso_project(client, tenant_id)

    rejected = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/trials",
        headers=headers(USER_A, tenant_id),
        json={"technical_result": "LIBERADO", "trial_name": "F1"},
    )
    accepted = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/trials",
        headers=headers(USER_A, tenant_id),
        json={"technical_result": "CUSTOM_OK", "trial_name": "F2"},
    )

    assert rejected.status_code == 400
    assert "not allowed" in rejected.json()["detail"]
    assert accepted.status_code == 201
    assert accepted.json()["technical_result"] == "CUSTOM_OK"


def test_f10_03_validation_requires_released_trial_and_publishes_project() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    enable_iso(client, tenant_id)
    project = create_iso_project(client, tenant_id)
    draft_trial = create_iso_trial(
        client,
        tenant_id,
        project["id"],
        "OK_NO_LIBERADO",
        trial_name="Draft Formula",
    )
    released_trial = create_iso_trial(
        client,
        tenant_id,
        project["id"],
        "LIBERADO",
        trial_name="Released Formula",
    )

    rejected = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/validation",
        headers=headers(USER_A, tenant_id),
        json={"released_trial_id": draft_trial["id"]},
    )
    created = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/validation",
        headers=headers(USER_A, tenant_id),
        json={"released_trial_id": released_trial["id"]},
    )

    assert rejected.status_code == 400
    assert "LIBERADO" in rejected.json()["detail"]
    assert created.status_code == 201
    validation = created.json()
    assert validation["status"] == "draft"
    assert validation["formula_ok"] == "Released Formula"
    assert validation["validation_checks"]
    assert {check["result"] for check in validation["validation_checks"]} == {"pending"}

    blocked = client.post(
        f"/api/v1/iso/product-validations/{validation['id']}/publish",
        headers=headers(USER_A, tenant_id),
    )
    checks = [
        {**check, "result": "ok", "comments": "Validado"}
        for check in validation["validation_checks"]
    ]
    updated = client.put(
        f"/api/v1/iso/product-validations/{validation['id']}/checks",
        headers=headers(USER_A, tenant_id),
        json={"checks": checks},
    )
    published = client.post(
        f"/api/v1/iso/product-validations/{validation['id']}/publish",
        headers=headers(USER_A, tenant_id),
    )
    refreshed_project = client.get(
        f"/api/v1/iso/design-projects/{project['id']}",
        headers=headers(USER_A, tenant_id),
    )

    assert blocked.status_code == 400
    assert "Required validation checks" in blocked.json()["detail"]
    assert updated.status_code == 200
    assert {check["result"] for check in updated.json()["validation_checks"]} == {"ok"}
    assert published.status_code == 200
    assert published.json()["status"] == "published"
    assert published.json()["validation_at"] is not None
    assert published.json()["published_at"] is not None
    assert refreshed_project.status_code == 200
    assert refreshed_project.json()["lifecycle_status"] == "validated"


def test_f10_03_validation_matrix_is_configurable_per_tenant() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    enable_iso(client, tenant_id)
    client.patch(
        "/api/v1/iso/settings",
        headers=headers(USER_A, tenant_id),
        json={
            "config_patch": {
                "f10_03": {
                    "validation_matrix": [
                        {"area": "Registros", "aspect": "Etiqueta", "required": True},
                        {"area": "Ventas", "aspect": "Tarifa", "required": False},
                    ]
                }
            }
        },
    )
    project = create_iso_project(client, tenant_id)
    released_trial = create_iso_trial(client, tenant_id, project["id"], "LIBERADO")

    created = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/validation",
        headers=headers(USER_A, tenant_id),
        json={"released_trial_id": released_trial["id"]},
    )

    assert created.status_code == 201
    checks = created.json()["validation_checks"]
    assert [check["area"] for check in checks] == ["Registros", "Ventas"]
    assert [check["aspect"] for check in checks] == ["Etiqueta", "Tarifa"]
    assert [check["required"] for check in checks] == [True, False]


def test_iso_legacy_f10_01_import_previews_and_applies_ready_projects() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    enable_iso(client, tenant_id)
    content = f10_01_workbook()

    preview = upload_legacy_xlsx(
        client,
        tenant_id,
        "/api/v1/iso/imports/f10-01/preview",
        content,
    )
    applied = upload_legacy_xlsx(
        client,
        tenant_id,
        "/api/v1/iso/imports/f10-01/apply",
        content,
    )
    reapplied = upload_legacy_xlsx(
        client,
        tenant_id,
        "/api/v1/iso/imports/f10-01/apply",
        content,
    )
    projects = client.get("/api/v1/iso/design-projects", headers=headers(USER_A, tenant_id))

    assert preview.status_code == 200
    assert preview.json()["total_rows"] == 2
    assert preview.json()["ready_rows"] == 1
    assert preview.json()["ambiguous_rows"] == 1
    assert applied.status_code == 200
    assert applied.json()["created_projects"] == 1
    assert applied.json()["skipped_rows"] == 1
    assert applied.json()["rows"][0]["status"] == "applied"
    assert applied.json()["rows"][1]["status"] == "ambiguous"
    assert reapplied.status_code == 200
    assert reapplied.json()["created_projects"] == 0
    assert reapplied.json()["updated_projects"] == 1
    assert projects.status_code == 200
    assert len(projects.json()) == 1
    assert projects.json()[0]["iso_request_number"] == "10/2026"
    assert projects.json()[0]["project_code"] == "AA-10"
    assert projects.json()[0]["product_name"] == "Legacy Calcium"


def test_iso_legacy_f10_02_import_creates_project_trials_and_snapshots() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    enable_iso(client, tenant_id)

    applied = upload_legacy_xlsx(
        client,
        tenant_id,
        "/api/v1/iso/imports/f10-02/apply",
        f10_02_workbook(),
    )
    projects = client.get("/api/v1/iso/design-projects", headers=headers(USER_A, tenant_id))
    project = projects.json()[0]
    trials = client.get(
        f"/api/v1/iso/design-projects/{project['id']}/trials",
        headers=headers(USER_A, tenant_id),
    )
    reapplied = upload_legacy_xlsx(
        client,
        tenant_id,
        "/api/v1/iso/imports/f10-02/apply",
        f10_02_workbook(),
    )

    assert applied.status_code == 200
    assert applied.json()["created_projects"] == 1
    assert applied.json()["created_trials"] == 2
    assert projects.status_code == 200
    assert project["iso_request_number"] == "12/2026"
    assert project["project_code"] == "12-2026 Legacy Amino"
    assert project["trial_count"] == 2
    assert trials.status_code == 200
    assert [trial["technical_result"] for trial in trials.json()] == ["NOK", "LIBERADO"]
    assert trials.json()[0]["snapshot"]["composition"] == [
        {"material_name": "Materia A", "percentage": "60"},
        {"material_name": "Materia B", "percentage": "40"},
    ]
    assert trials.json()[1]["reason_comment"] == "Liberado por calidad"
    assert reapplied.status_code == 200
    assert reapplied.json()["created_trials"] == 0
    assert reapplied.json()["updated_trials"] == 2


def test_iso_legacy_f10_03_import_creates_published_validation() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    enable_iso(client, tenant_id)

    applied = upload_legacy_xlsx(
        client,
        tenant_id,
        "/api/v1/iso/imports/f10-03/apply",
        f10_03_workbook(),
    )

    assert applied.status_code == 200
    assert applied.json()["created_projects"] == 1
    assert applied.json()["created_trials"] == 1
    assert applied.json()["created_validations"] == 1
    assert applied.json()["rows"][0]["status"] == "applied"
    with Session(client.app.state.engine) as session:
        project = session.exec(
            select(IsoDesignProject).where(
                IsoDesignProject.tenant_id == uuid.UUID(tenant_id),
                IsoDesignProject.iso_request_number == "13/2026",
            )
        ).first()
        assert project is not None
        validation = session.exec(
            select(IsoProductValidation).where(
                IsoProductValidation.tenant_id == uuid.UUID(tenant_id),
                IsoProductValidation.design_project_id == project.id,
            )
        ).first()
        assert validation is not None
        released_trial = session.get(IsoDesignTrial, validation.released_trial_id)

    assert project.lifecycle_status == "validated"
    assert validation.status == "published"
    assert validation.validation_at is not None
    assert validation.validation_at.date().isoformat() == "2026-06-01"
    assert validation.formula_ok == "ID-777 || Legacy Zinc F3"
    assert validation.specification_json["descripcion"] == "Producto validado"
    assert {check["result"] for check in validation.validation_checks_json} == {"ok"}
    assert released_trial is not None
    assert released_trial.technical_result == "LIBERADO"
    assert released_trial.trial_code == "ID-777"


def test_iso_exports_are_downloadable_and_tenant_scoped() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    enable_iso(client, tenant_id)
    enable_iso(client, tenant_b, USER_B)
    project = create_iso_project(client, tenant_id)
    create_iso_trial(
        client,
        tenant_id,
        project["id"],
        "NOK",
        trial_name="Rejected Formula",
    )
    released_trial = create_iso_trial(
        client,
        tenant_id,
        project["id"],
        "LIBERADO",
        trial_name="Released Formula",
    )
    validation = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/validation",
        headers=headers(USER_A, tenant_id),
        json={"released_trial_id": released_trial["id"]},
    ).json()
    checks = [{**check, "result": "ok"} for check in validation["validation_checks"]]
    client.put(
        f"/api/v1/iso/product-validations/{validation['id']}/checks",
        headers=headers(USER_A, tenant_id),
        json={"checks": checks},
    )

    f10_01 = client.post(
        "/api/v1/iso/exports/f10-01?year=2026",
        headers=headers(USER_A, tenant_id),
    )
    f10_02 = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/exports/f10-02",
        headers=headers(USER_A, tenant_id),
    )
    f10_03 = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/exports/f10-03",
        headers=headers(USER_A, tenant_id),
    )
    dossier = client.post(
        f"/api/v1/iso/design-projects/{project['id']}/dossier",
        headers=headers(USER_A, tenant_id),
    )

    assert f10_01.status_code == 200
    assert f10_01.json()["artifact_type"] == "iso_f10_01_xlsx"
    assert f10_02.status_code == 200
    assert f10_02.json()["artifact_type"] == "iso_f10_02_xlsx"
    assert f10_03.status_code == 200
    assert f10_03.json()["artifact_type"] == "iso_f10_03_xlsx"
    assert dossier.status_code == 200
    assert dossier.json()["artifact_type"] == "iso_dossier_zip"

    downloaded_f10_01 = client.get(
        f"/api/v1/iso/artifacts/{f10_01.json()['id']}/download",
        headers=headers(USER_A, tenant_id),
    )
    workbook_f10_01 = load_workbook(BytesIO(downloaded_f10_01.content), data_only=True)
    values_f10_01 = [
        cell
        for row in workbook_f10_01["F10-01"].iter_rows(values_only=True)
        for cell in row
    ]
    assert downloaded_f10_01.status_code == 200
    assert project["iso_request_number"] in values_f10_01
    assert "Flower Power" in values_f10_01

    downloaded_f10_02 = client.get(
        f"/api/v1/iso/artifacts/{f10_02.json()['id']}/download",
        headers=headers(USER_A, tenant_id),
    )
    workbook_f10_02 = load_workbook(BytesIO(downloaded_f10_02.content), data_only=True)
    values_f10_02 = [
        cell
        for row in workbook_f10_02["F10-02"].iter_rows(values_only=True)
        for cell in row
    ]
    assert "Released Formula" in values_f10_02
    assert "LIBERADO" in values_f10_02

    downloaded_f10_03 = client.get(
        f"/api/v1/iso/artifacts/{f10_03.json()['id']}/download",
        headers=headers(USER_A, tenant_id),
    )
    workbook_f10_03 = load_workbook(BytesIO(downloaded_f10_03.content), data_only=True)
    assert {"F10-03", "Validaciones", "Especificacion"}.issubset(workbook_f10_03.sheetnames)

    downloaded_dossier = client.get(
        f"/api/v1/iso/artifacts/{dossier.json()['id']}/download",
        headers=headers(USER_A, tenant_id),
    )
    with ZipFile(BytesIO(downloaded_dossier.content)) as archive:
        names = set(archive.namelist())
    assert "metadata.json" in names
    assert any(name.startswith("formulia_iso_f10_01") for name in names)
    assert any(name.startswith("formulia_iso_f10_02") for name in names)
    assert any(name.startswith("formulia_iso_f10_03") for name in names)

    cross_tenant_download = client.get(
        f"/api/v1/iso/artifacts/{f10_01.json()['id']}/download",
        headers=headers(USER_B, tenant_b),
    )
    assert cross_tenant_download.status_code == 404
