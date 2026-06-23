from datetime import date
from io import BytesIO
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import SQLModel, Session, create_engine, select

from formulia_api import jira_integration
from formulia_api.jira_client import (
    JiraAttachmentResult,
    JiraClientError,
    JiraConnectionCheckResult,
    JiraIssueResult,
)
from formulia_api.jira_oauth import JiraOAuthCallbackResult
from formulia_api.main import create_app
from formulia_api.models import (
    FormulaReviewArtifact,
    IntegrationEvent,
    IsoDesignTrial,
    TenantMember,
)


USER_A = "10000000-0000-0000-0000-000000000001"
USER_B = "20000000-0000-0000-0000-000000000001"


class FakeJiraClient:
    def __init__(
        self,
        *,
        fail_attachment: bool = False,
        issue_status: str = "Cambios solicitados",
        technical_result: str | None = None,
    ) -> None:
        self.fail_attachment = fail_attachment
        self.issue_status = issue_status
        self.technical_result = technical_result
        self.created_payloads: list[dict] = []
        self.attachments: list[dict] = []

    def list_projects(self) -> dict:
        return {
            "values": [
                {
                    "id": "10000",
                    "key": "LAB",
                    "name": "Formula Lab",
                    "projectTypeKey": "software",
                    "simplified": True,
                }
            ]
        }

    def get_project(self, project_key: str) -> dict:
        assert project_key == "LAB"
        return {
            "key": "LAB",
            "name": "Formula Lab",
            "issueTypes": [
                {
                    "id": "10001",
                    "name": "Review",
                    "description": "Formula review",
                    "subtask": False,
                },
                {
                    "id": "10002",
                    "name": "Prototype",
                    "description": "Prototype review",
                    "subtask": False,
                },
                {
                    "id": "10003",
                    "name": "Calidad",
                    "description": "Quality review",
                    "subtask": False,
                },
                {
                    "id": "10004",
                    "name": "PoC",
                    "description": "Proof of concept",
                    "subtask": False,
                },
            ],
        }

    def get_create_issue_fields(self, project_key: str, issue_type_id: str) -> dict:
        assert project_key == "LAB"
        assert issue_type_id in {"10001", "10003", "10004"}
        if issue_type_id == "10004":
            return {
                "fields": [
                    {
                        "fieldId": "project",
                        "name": "Project",
                        "required": True,
                        "schema": {"type": "project"},
                        "allowedValues": [{"id": "10000", "key": "LAB", "name": "Formula Lab"}],
                    },
                    {
                        "fieldId": "issuetype",
                        "name": "Issue type",
                        "required": True,
                        "schema": {"type": "issuetype"},
                        "allowedValues": [{"id": "10004", "name": "PoC"}],
                    },
                    {
                        "fieldId": "summary",
                        "name": "Summary",
                        "required": True,
                        "schema": {"type": "string"},
                    },
                    {
                        "fieldId": "description",
                        "name": "Description",
                        "required": False,
                        "schema": {"type": "string"},
                    },
                    {
                        "fieldId": "labels",
                        "name": "Labels",
                        "required": False,
                        "schema": {"type": "array"},
                    },
                ]
            }
        return {
            "fields": [
                {
                    "fieldId": "project",
                    "name": "Project",
                    "required": True,
                    "schema": {"type": "project"},
                    "allowedValues": [{"id": "10000", "key": "LAB", "name": "Formula Lab"}],
                },
                {
                    "fieldId": "issuetype",
                    "name": "Issue type",
                    "required": True,
                    "schema": {"type": "issuetype"},
                    "allowedValues": [{"id": issue_type_id, "name": "Review" if issue_type_id == "10001" else "Calidad"}],
                },
                {
                    "fieldId": "summary",
                    "name": "Summary",
                    "required": True,
                    "schema": {"type": "string"},
                },
                {
                    "fieldId": "description",
                    "name": "Description",
                    "required": False,
                    "schema": {"type": "string"},
                },
                {
                    "fieldId": "labels",
                    "name": "Labels",
                    "required": False,
                    "schema": {"type": "array"},
                },
                {
                    "fieldId": "customfield_10010",
                    "name": "Formula name",
                    "required": False,
                    "schema": {"type": "string", "custom": "text"},
                },
                {
                    "fieldId": "customfield_20010",
                    "name": "Functional project",
                    "required": True,
                    "schema": {"type": "string", "custom": "text"},
                },
                {
                    "fieldId": "customfield_20011",
                    "name": "Product type",
                    "required": True,
                    "schema": {"type": "option", "custom": "select"},
                    "allowedValues": [
                        {"id": "1", "value": "Nuevo"},
                        {"id": "2", "value": "Mod A"},
                    ],
                },
            ]
        }

    def create_issue(self, payload: dict) -> JiraIssueResult:
        self.created_payloads.append(payload)
        return JiraIssueResult(
            key="LAB-321",
            url="https://example.atlassian.net/browse/LAB-321",
        )

    def get_issue(self, issue_key: str, fields: str | None = None) -> dict:
        assert issue_key == "LAB-321"
        issue_fields = {
            "summary": "Review Formula",
            "status": {"name": self.issue_status},
        }
        if fields and "customfield_11024" in fields and self.technical_result:
            issue_fields["customfield_11024"] = {"value": self.technical_result}
        return {
            "key": issue_key,
            "fields": issue_fields,
        }

    def get_issue_transitions(self, issue_key: str) -> dict:
        assert issue_key == "LAB-321"
        return {
            "transitions": [
                {"id": "31", "name": "OK"},
                {"id": "41", "name": "NOK"},
            ]
        }

    def add_attachment(self, issue_key: str, artifact: object) -> JiraAttachmentResult:
        if self.fail_attachment:
            raise JiraClientError("Attachment upload failed.")
        self.attachments.append(
            {
                "issue_key": issue_key,
                "file_name": getattr(artifact, "file_name"),
                "content_type": getattr(artifact, "content_type"),
                "size_bytes": getattr(artifact, "size_bytes"),
            }
        )
        return JiraAttachmentResult(
            attachment_id="10001",
            file_name=getattr(artifact, "file_name"),
        )


def use_fake_jira_client(monkeypatch, fake: FakeJiraClient) -> None:
    monkeypatch.setattr(jira_integration, "make_jira_client", lambda connection: fake)


def make_client() -> TestClient:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    app = create_app(engine)
    SQLModel.metadata.create_all(engine)
    return TestClient(app)


def create_tenant(client: TestClient, user_id: str, slug: str) -> str:
    response = client.post(
        "/api/v1/tenants",
        headers={"X-User-Id": user_id},
        json={"name": slug.title(), "slug": slug},
    )
    assert response.status_code == 201
    return response.json()["id"]


def headers(user_id: str, tenant_id: str) -> dict[str, str]:
    return {"X-User-Id": user_id, "X-Tenant-Id": tenant_id}


def add_member(client: TestClient, tenant_id: str, user_id: str, role: str) -> None:
    client.get("/api/v1/me", headers={"X-User-Id": user_id})
    with Session(client.app.state.engine) as session:
        session.add(
            TenantMember(
                tenant_id=uuid.UUID(tenant_id),
                user_id=uuid.UUID(user_id),
                role=role,
            )
        )
        session.commit()


def create_jira_connection(client: TestClient, tenant_id: str, user_id: str = USER_A) -> dict:
    response = client.post(
        "/api/v1/integrations/jira",
        headers=headers(user_id, tenant_id),
        json={
            "base_url": "https://example.atlassian.net",
            "auth_email": "lab@example.com",
            "api_token": "secret-token",
            "default_project_key": "LAB",
            "default_issue_type": "Revision de formula",
        },
    )
    assert response.status_code == 201
    return response.json()


def create_oauth_jira_connection(
    client: TestClient,
    tenant_id: str,
    user_id: str = USER_A,
) -> dict:
    response = client.post(
        "/api/v1/integrations/jira",
        headers=headers(user_id, tenant_id),
        json={
            "base_url": "https://example.atlassian.net",
            "auth_type": "oauth",
            "default_project_key": "LAB",
            "default_issue_type": "Calidad",
            "field_mapping": {
                "formula_name": "customfield_10010",
                "jira_project_id": "customfield_20010",
                "jira_product_type_option": "customfield_20011",
                "technical_result": "customfield_11024",
            },
        },
    )
    assert response.status_code == 201
    return response.json()


def create_formula(
    client: TestClient,
    tenant_id: str,
    user_id: str = USER_A,
    *,
    jira_issue_type: str = "Calidad",
    jira_product_type: str = "Nuevo",
) -> dict:
    request_headers = headers(user_id, tenant_id)
    material = client.post(
        "/api/v1/raw-materials",
        headers=request_headers,
        json={"name": "Material A", "code": "MAT-A"},
    ).json()
    formula_response = client.post(
        "/api/v1/formulas",
        headers=request_headers,
        json={
            "name": "Review Formula",
            "objective": "Review formula prepared from FormulIA.",
            "jira_project_id": "FLOWER",
            "jira_issue_type": jira_issue_type,
            "jira_product_type": jira_product_type,
            "items": [{"raw_material_id": material["id"], "percentage": 100}],
        },
    )
    assert formula_response.status_code == 201
    return formula_response.json()


def enable_iso(client: TestClient, tenant_id: str, user_id: str = USER_A) -> None:
    response = client.patch(
        "/api/v1/iso/settings",
        headers=headers(user_id, tenant_id),
        json={"enabled": True},
    )
    assert response.status_code == 200


def create_iso_project(client: TestClient, tenant_id: str, user_id: str = USER_A) -> dict:
    response = client.post(
        "/api/v1/iso/design-projects",
        headers=headers(user_id, tenant_id),
        json={
            "iso_request_number": "1/2026",
            "project_code": "FLOWER",
            "product_name": "Flower Power",
            "accepted_status": "accepted",
            "lifecycle_status": "design",
        },
    )
    assert response.status_code == 201
    return response.json()


def test_owner_configures_and_tests_jira_connection(monkeypatch) -> None:
    monkeypatch.setattr(
        jira_integration,
        "check_jira_connection",
        lambda connection: JiraConnectionCheckResult(
            status="ready_for_client",
            message="Connected to Jira as Lab User; project LAB is ready.",
        ),
    )
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")

    created = client.post(
        "/api/v1/integrations/jira",
        headers=headers(USER_A, tenant_id),
        json={
            "base_url": " https://example.atlassian.net/ ",
            "auth_email": " lab@example.com ",
            "api_token": "secret-token",
            "default_project_key": " lab ",
            "default_issue_type": "Revision de formula",
            "default_assignee": "lab-team",
            "field_mapping": {" formula_code ": " customfield_10001 "},
        },
    )
    listed = client.get("/api/v1/integrations/jira", headers=headers(USER_A, tenant_id))

    assert created.status_code == 201
    connection = created.json()
    assert "api_token" not in connection
    assert "credential_json" not in connection
    assert connection["base_url"] == "https://example.atlassian.net"
    assert connection["auth_email"] == "lab@example.com"
    assert connection["credential_status"] == "configured"
    assert connection["default_project_key"] == "LAB"
    assert connection["field_mapping"] == {"formula_code": "customfield_10001"}
    assert listed.status_code == 200
    assert [item["id"] for item in listed.json()] == [connection["id"]]

    tested = client.post(
        f"/api/v1/integrations/jira/{connection['id']}/test",
        headers=headers(USER_A, tenant_id),
    )

    assert tested.status_code == 200
    assert tested.json()["status"] == "ready_for_client"
    assert "Connected to Jira" in tested.json()["message"]


def test_owner_loads_jira_connector_metadata(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient()
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    connection = create_jira_connection(client, tenant_id)
    request_headers = headers(USER_A, tenant_id)

    projects = client.get(
        f"/api/v1/integrations/jira/{connection['id']}/projects",
        headers=request_headers,
    )
    issue_types = client.get(
        f"/api/v1/integrations/jira/{connection['id']}/issue-types",
        headers=request_headers,
        params={"project_key": "lab"},
    )
    fields = client.get(
        f"/api/v1/integrations/jira/{connection['id']}/fields",
        headers=request_headers,
        params={"project_key": "LAB", "issue_type": "Review"},
    )

    assert projects.status_code == 200
    assert projects.json() == [
        {
            "id": "10000",
            "key": "LAB",
            "name": "Formula Lab",
            "project_type_key": "software",
            "simplified": True,
        }
    ]
    assert issue_types.status_code == 200
    assert issue_types.json()[0] == {
        "id": "10001",
        "name": "Review",
        "description": "Formula review",
        "subtask": False,
    }
    assert fields.status_code == 200
    fields_by_id = {field["field_id"]: field for field in fields.json()}
    assert fields_by_id["customfield_20010"] == {
        "field_id": "customfield_20010",
        "name": "Functional project",
        "required": True,
        "schema_type": "string",
        "custom": "text",
        "allowed_values": [],
    }
    assert fields_by_id["customfield_20011"]["allowed_values"] == [
        {"id": "1", "key": None, "name": None, "value": "Nuevo"},
        {"id": "2", "key": None, "name": None, "value": "Mod A"},
    ]


def test_jira_connector_metadata_requires_admin(monkeypatch) -> None:
    client = make_client()
    use_fake_jira_client(monkeypatch, FakeJiraClient())
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    add_member(client, tenant_id, USER_B, "formulator")
    connection = create_jira_connection(client, tenant_id)

    response = client.get(
        f"/api/v1/integrations/jira/{connection['id']}/projects",
        headers=headers(USER_B, tenant_id),
    )

    assert response.status_code == 403


def test_jira_oauth_authorize_url_requires_admin_and_returns_redirect(monkeypatch) -> None:
    monkeypatch.setenv("FORMULIA_JIRA_OAUTH_CLIENT_ID", "client-id")
    monkeypatch.setenv("FORMULIA_JIRA_OAUTH_CLIENT_SECRET", "client-secret")
    monkeypatch.setenv("FORMULIA_JIRA_OAUTH_REDIRECT_URI", "http://localhost:3000/callback")
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    add_member(client, tenant_id, USER_B, "formulator")

    authorized = client.get(
        "/api/v1/integrations/jira/oauth/authorize-url",
        headers=headers(USER_A, tenant_id),
    )
    forbidden = client.get(
        "/api/v1/integrations/jira/oauth/authorize-url",
        headers=headers(USER_B, tenant_id),
    )

    assert authorized.status_code == 200
    assert authorized.json()["authorization_url"].startswith(
        "https://auth.atlassian.com/authorize?"
    )
    assert authorized.json()["state"]
    assert forbidden.status_code == 403


def test_jira_oauth_callback_route_completes_exchange(monkeypatch) -> None:
    client = make_client()

    def fake_complete(code: str, state: str) -> JiraOAuthCallbackResult:
        assert code == "authorization-code"
        assert state == "signed-state"
        return JiraOAuthCallbackResult(
            cloud_id="cloud-123",
            site_url="https://example.atlassian.net",
            expires_at=1234567890,
            scope="read:issue:jira write:issue:jira offline_access",
        )

    monkeypatch.setattr(jira_integration.jira_oauth, "complete_jira_oauth_callback", fake_complete)

    response = client.post(
        "/api/v1/integrations/jira/oauth/callback",
        json={"code": "authorization-code", "state": "signed-state"},
    )

    assert response.status_code == 200
    assert response.json() == {
        "status": "connected",
        "cloud_id": "cloud-123",
        "site_url": "https://example.atlassian.net",
        "expires_at": 1234567890,
        "scope": "read:issue:jira write:issue:jira offline_access",
    }


def test_jira_connections_are_tenant_scoped() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")

    connection = client.post(
        "/api/v1/integrations/jira",
        headers=headers(USER_A, tenant_a),
        json={
            "base_url": "https://example.atlassian.net",
            "auth_email": "lab@example.com",
            "api_token": "secret-token",
            "default_project_key": "LAB",
            "default_issue_type": "Revision de formula",
        },
    ).json()

    listed_b = client.get("/api/v1/integrations/jira", headers=headers(USER_B, tenant_b))
    cross_patch = client.patch(
        f"/api/v1/integrations/jira/{connection['id']}",
        headers=headers(USER_B, tenant_b),
        json={"default_project_key": "OTHER"},
    )

    assert listed_b.status_code == 200
    assert listed_b.json() == []
    assert cross_patch.status_code == 404


def test_only_owner_or_admin_can_mutate_jira_connection() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    add_member(client, tenant_id, USER_B, "formulator")

    created = client.post(
        "/api/v1/integrations/jira",
        headers=headers(USER_A, tenant_id),
        json={
            "base_url": "https://example.atlassian.net",
            "auth_email": "lab@example.com",
            "api_token": "secret-token",
            "default_project_key": "LAB",
            "default_issue_type": "Revision de formula",
        },
    ).json()

    listed = client.get("/api/v1/integrations/jira", headers=headers(USER_B, tenant_id))
    forbidden_patch = client.patch(
        f"/api/v1/integrations/jira/{created['id']}",
        headers=headers(USER_B, tenant_id),
        json={"default_project_key": "LAB2"},
    )
    forbidden_test = client.post(
        f"/api/v1/integrations/jira/{created['id']}/test",
        headers=headers(USER_B, tenant_id),
    )

    assert listed.status_code == 200
    assert listed.json()[0]["id"] == created["id"]
    assert forbidden_patch.status_code == 403
    assert forbidden_test.status_code == 403


def test_formula_jira_review_request_captures_snapshot() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    connection = create_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)

    created = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_id),
        json={"notes": "Priorizar revision de estabilidad."},
    )
    listed = client.get(
        f"/api/v1/formulas/{formula['id']}/reviews",
        headers=headers(USER_A, tenant_id),
    )

    assert created.status_code == 201
    review = created.json()
    assert review["tenant_id"] == tenant_id
    assert review["formula_id"] == formula["id"]
    assert review["formula_version"] == 1
    assert review["jira_connection_id"] == connection["id"]
    assert review["review_status"] == "ready_for_jira"
    assert review["jira_issue_key"] is None
    assert review["snapshot"]["formula"]["name"] == "Review Formula"
    assert review["snapshot"]["items"] == [
        {
            "raw_material_id": formula["items"][0]["raw_material_id"],
            "code": "MAT-A",
            "name": "Material A",
            "percentage": 100.0,
            "quantity": None,
            "unit": None,
            "order_index": 0,
            "price": None,
            "currency": None,
            "parameters": [],
        }
    ]
    assert review["snapshot"]["jira"]["project_key"] == "LAB"
    assert review["snapshot"]["formula"]["jira_project_id"] == "FLOWER"
    assert review["snapshot"]["formula"]["jira_issue_type"] == "Calidad"
    assert review["snapshot"]["formula"]["jira_product_type"] == "Nuevo"
    assert review["snapshot"]["jira"]["issue_summary"] == "Review Formula"
    assert review["snapshot"]["jira"]["issue_description"] is None
    assert review["snapshot"]["notes"] == "Priorizar revision de estabilidad."
    assert listed.status_code == 200
    assert [item["id"] for item in listed.json()] == [review["id"]]


def test_formula_jira_review_can_link_iso_trial_and_sync_result(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient(issue_status="FINALIZADO", technical_result="Liberado")
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    enable_iso(client, tenant_id)
    project = create_iso_project(client, tenant_id)
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    request_headers = headers(USER_A, tenant_id)

    review_response = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=request_headers,
        json={
            "design_project_id": project["id"],
            "iso_trial_number": 2,
            "iso_reason_comment": "Ensayo desde revision Jira",
        },
    )
    assert review_response.status_code == 201
    review = review_response.json()
    initial_trials = client.get(
        f"/api/v1/iso/design-projects/{project['id']}/trials",
        headers=request_headers,
    )

    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=request_headers,
    )
    synced = client.post(
        f"/api/v1/formula-reviews/{review['id']}/sync",
        headers=request_headers,
    )
    final_trials = client.get(
        f"/api/v1/iso/design-projects/{project['id']}/trials",
        headers=request_headers,
    )

    assert review["snapshot"]["iso"] == {
        "design_project_id": project["id"],
        "trial_number": 2,
        "reason_comment": "Ensayo desde revision Jira",
        "trial_intent": "f10_02_trial",
    }
    assert initial_trials.status_code == 200
    assert len(initial_trials.json()) == 1
    assert initial_trials.json()[0]["review_request_id"] == review["id"]
    assert initial_trials.json()[0]["technical_result"] == "pending_result"
    assert initial_trials.json()[0]["trial_number"] == 2
    assert sent.status_code == 200
    assert synced.status_code == 200
    final_trial = final_trials.json()[0]
    assert final_trial["id"] == initial_trials.json()[0]["id"]
    assert final_trial["jira_issue_key"] == "LAB-321"
    assert final_trial["technical_result"] == "LIBERADO"
    assert final_trial["raw_result_label"] == "Liberado"
    assert final_trial["raw_status_label"] == "FINALIZADO"
    with Session(client.app.state.engine) as session:
        trials = session.exec(
            select(IsoDesignTrial).where(IsoDesignTrial.tenant_id == uuid.UUID(tenant_id))
        ).all()
    assert len(trials) == 1


def test_formula_jira_review_excel_artifact_is_generated_and_downloadable() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    request_headers = headers(USER_A, tenant_id)

    calculation = client.post(
        f"/api/v1/formulas/{formula['id']}/calculate",
        headers=request_headers,
    )
    created_review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=request_headers,
        json={"notes": "Preparar paquete Excel."},
    )
    review = created_review.json()

    generated = client.post(
        f"/api/v1/formula-reviews/{review['id']}/artifacts/excel",
        headers=request_headers,
    )
    repeated = client.post(
        f"/api/v1/formula-reviews/{review['id']}/artifacts/excel",
        headers=request_headers,
    )
    listed = client.get(
        f"/api/v1/formula-reviews/{review['id']}/artifacts",
        headers=request_headers,
    )

    assert calculation.status_code == 200
    assert created_review.status_code == 201
    assert generated.status_code == 201
    assert repeated.status_code == 200
    artifact = generated.json()
    assert artifact["id"] == repeated.json()["id"]
    assert artifact["artifact_type"] == "jira_review_xlsx"
    assert artifact["file_name"] == f"DT_Review_Formula_{date.today().isoformat()}.xlsx"
    assert artifact["size_bytes"] > 0
    assert listed.status_code == 200
    assert [item["id"] for item in listed.json()] == [artifact["id"]]

    downloaded = client.get(
        f"/api/v1/formula-review-artifacts/{artifact['id']}/download",
        headers=request_headers,
    )

    assert downloaded.status_code == 200
    assert artifact["file_name"] in downloaded.headers["content-disposition"]

    workbook = load_workbook(BytesIO(downloaded.content), data_only=False)
    assert workbook.sheetnames == ["Calculadora", "Hoja Lab", "Composición"]

    calculator = workbook["Calculadora"]
    assert calculator["A1"].value == "Materia Prima"
    assert calculator["A2"].value == "Material A"
    assert calculator["B2"].value == 100
    assert calculator["A4"].value == "TOTAL"
    assert calculator["B4"].value == "=SUM(B2:B3)"
    assert calculator["C4"].value == "=SUMPRODUCT($B$2:$B$2,C2:C2)/100"

    lab = workbook["Hoja Lab"]
    assert lab["C3"].value == "Review Formula"
    assert lab["C6"].value == "Material A"
    assert lab["D6"].value == 100
    assert lab["E6"].value == "=D6*10/2"
    assert lab["F6"].value == "=D6*10*2"
    assert lab["D8"].value == "=SUM(D6:D6)"

    composition = workbook["Composición"]
    assert composition["A2"].value == "Parámetro"
    assert composition["A3"].value == "Precio"
    assert composition["B3"].value == "='Calculadora'!C4"


def test_formula_jira_review_excel_artifact_updates_legacy_filename() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    request_headers = headers(USER_A, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=request_headers,
        json={},
    ).json()

    with Session(client.app.state.engine) as session:
        artifact = FormulaReviewArtifact(
            tenant_id=uuid.UUID(tenant_id),
            review_request_id=uuid.UUID(review["id"]),
            artifact_type="jira_review_xlsx",
            file_name="formulia_id_lab_Review_Formula_v1_legacy.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            checksum_sha256="legacy",
            size_bytes=3,
            content=b"old",
        )
        session.add(artifact)
        session.commit()
        session.refresh(artifact)
        artifact_id = str(artifact.id)

    generated = client.post(
        f"/api/v1/formula-reviews/{review['id']}/artifacts/excel",
        headers=request_headers,
    )

    assert generated.status_code == 200
    updated = generated.json()
    assert updated["id"] == artifact_id
    assert updated["file_name"] == f"DT_Review_Formula_{date.today().isoformat()}.xlsx"
    assert updated["checksum_sha256"] != "legacy"
    assert updated["size_bytes"] > 3


def test_formula_jira_review_can_be_sent_to_jira_with_excel_attachment(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient()
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    connection = create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    request_headers = headers(USER_A, tenant_id)

    client.post(
        f"/api/v1/formulas/{formula['id']}/calculate",
        headers=request_headers,
    )
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=request_headers,
        json={"description": "Descripcion redactada por I+D."},
    ).json()

    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=request_headers,
    )
    listed_artifacts = client.get(
        f"/api/v1/formula-reviews/{review['id']}/artifacts",
        headers=request_headers,
    )

    assert connection["auth_type"] == "oauth"
    assert connection["credential_status"] == "external"
    assert sent.status_code == 200
    sent_review = sent.json()
    assert sent_review["review_status"] == "sent_to_jira"
    assert sent_review["jira_issue_key"] == "LAB-321"
    assert sent_review["jira_issue_url"] == "https://example.atlassian.net/browse/LAB-321"
    assert sent_review["jira_status"] == "created"
    assert sent_review["sent_by_user_id"] == USER_A
    assert sent_review["sent_at"] is not None

    assert len(fake_jira.created_payloads) == 1
    fields = fake_jira.created_payloads[0]["fields"]
    assert fields["project"] == {"key": "LAB"}
    assert fields["issuetype"] == {"name": "Calidad"}
    assert fields["summary"] == "Review Formula"
    assert "reporter" not in fields
    assert fields["description"]["type"] == "doc"
    assert fields["description"]["content"][0]["content"][0]["text"] == "Descripcion redactada por I+D."
    assert "labels" not in fields
    assert fields["customfield_10010"] == "Review Formula"
    assert fields["customfield_20010"] == "FLOWER"
    assert fields["customfield_20011"] == {"value": "Nuevo"}

    assert len(fake_jira.attachments) == 1
    assert fake_jira.attachments[0]["issue_key"] == "LAB-321"
    assert fake_jira.attachments[0]["file_name"].endswith(".xlsx")
    assert fake_jira.attachments[0]["size_bytes"] > 0
    assert listed_artifacts.status_code == 200
    assert len(listed_artifacts.json()) == 1


def test_formula_jira_review_filters_fields_not_available_for_issue_type(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient()
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id, jira_issue_type="PoC")
    request_headers = headers(USER_A, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=request_headers,
        json={},
    ).json()

    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=request_headers,
    )

    assert sent.status_code == 200
    fields = fake_jira.created_payloads[0]["fields"]
    assert fields["issuetype"] == {"name": "PoC"}
    assert fields["summary"] == "Review Formula"
    assert "description" not in fields
    assert "labels" not in fields
    assert "customfield_10010" not in fields
    assert "customfield_20010" not in fields
    assert "customfield_20011" not in fields


def test_formula_jira_review_send_requires_issue_type_fields(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient()
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id, jira_issue_type="Calidad", jira_product_type="")
    request_headers = headers(USER_A, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=request_headers,
        json={},
    ).json()

    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=request_headers,
    )

    assert sent.status_code == 400
    assert "Product type" in sent.json()["detail"]
    assert fake_jira.created_payloads == []


def test_formula_jira_review_send_rejects_invalid_issue_type_option(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient()
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id, jira_issue_type="Calidad", jira_product_type="Mod C")
    request_headers = headers(USER_A, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=request_headers,
        json={},
    ).json()

    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=request_headers,
    )

    assert sent.status_code == 400
    assert "Product type" in sent.json()["detail"]
    assert "Mod C" in sent.json()["detail"]
    assert fake_jira.created_payloads == []


def test_formula_jira_review_send_avoids_duplicate_issue(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient()
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_id),
        json={},
    ).json()

    first_send = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=headers(USER_A, tenant_id),
    )
    duplicate_send = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=headers(USER_A, tenant_id),
    )

    assert first_send.status_code == 200
    assert duplicate_send.status_code == 409
    assert duplicate_send.json()["detail"] == "Formula review is already sent to Jira."
    assert len(fake_jira.created_payloads) == 1


def test_formula_jira_review_send_marks_partial_failure_when_attachment_fails(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient(fail_attachment=True)
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_id),
        json={},
    ).json()

    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=headers(USER_A, tenant_id),
    )

    assert sent.status_code == 200
    sent_review = sent.json()
    assert sent_review["review_status"] == "partial_failure"
    assert sent_review["jira_issue_key"] == "LAB-321"
    assert sent_review["jira_status"] == "attachment_failed"
    assert len(fake_jira.created_payloads) == 1
    assert fake_jira.attachments == []


def test_formula_jira_review_retry_attachment_after_partial_failure(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient(fail_attachment=True)
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_id),
        json={},
    ).json()
    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=headers(USER_A, tenant_id),
    )
    assert sent.status_code == 200
    assert sent.json()["review_status"] == "partial_failure"

    fake_jira.fail_attachment = False
    retried = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/retry-attachment",
        headers=headers(USER_A, tenant_id),
    )

    assert retried.status_code == 200
    retried_review = retried.json()
    assert retried_review["review_status"] == "sent_to_jira"
    assert retried_review["jira_status"] == "created"
    assert len(fake_jira.created_payloads) == 1
    assert len(fake_jira.attachments) == 1
    assert fake_jira.attachments[0]["issue_key"] == "LAB-321"


def test_formula_jira_review_sync_updates_status_from_jira(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient(issue_status="OK")
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_id),
        json={},
    ).json()
    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=headers(USER_A, tenant_id),
    )
    assert sent.status_code == 200

    synced = client.post(
        f"/api/v1/formula-reviews/{review['id']}/sync",
        headers=headers(USER_A, tenant_id),
    )

    assert synced.status_code == 200
    synced_review = synced.json()
    assert synced_review["review_status"] == "approved"
    assert synced_review["jira_status"] == "OK"
    assert synced_review["last_sync_at"] is not None
    with Session(client.app.state.engine) as session:
        events = session.exec(
            select(IntegrationEvent).where(IntegrationEvent.tenant_id == uuid.UUID(tenant_id))
        ).all()
    sync_event = [event for event in events if event.event_type == "jira_status_sync"][0]
    assert sync_event.status == "success"
    assert sync_event.payload_summary["available_transitions"] == ["OK", "NOK"]


def test_formula_jira_review_sync_records_technical_result_from_jira(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient(issue_status="FINALIZADO", technical_result="Liberado")
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_id),
        json={},
    ).json()
    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=headers(USER_A, tenant_id),
    )
    assert sent.status_code == 200

    synced = client.post(
        f"/api/v1/formula-reviews/{review['id']}/sync",
        headers=headers(USER_A, tenant_id),
    )

    assert synced.status_code == 200
    synced_review = synced.json()
    assert synced_review["jira_status"] == "FINALIZADO"
    assert synced_review["snapshot"]["jira"]["technical_result_raw"] == "Liberado"
    assert synced_review["snapshot"]["jira"]["technical_result"] == "LIBERADO"
    with Session(client.app.state.engine) as session:
        events = session.exec(
            select(IntegrationEvent).where(IntegrationEvent.tenant_id == uuid.UUID(tenant_id))
        ).all()
    sync_event = [event for event in events if event.event_type == "jira_status_sync"][0]
    assert sync_event.payload_summary["jira_technical_result"] == "Liberado"
    assert sync_event.payload_summary["technical_result"] == "LIBERADO"


def test_formula_jira_review_sync_keeps_internal_status_when_unmapped(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient(issue_status="Cliente QA")
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_id),
        json={},
    ).json()
    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=headers(USER_A, tenant_id),
    )
    assert sent.status_code == 200

    synced = client.post(
        f"/api/v1/formula-reviews/{review['id']}/sync",
        headers=headers(USER_A, tenant_id),
    )

    assert synced.status_code == 200
    synced_review = synced.json()
    assert synced_review["review_status"] == "sent_to_jira"
    assert synced_review["jira_status"] == "Cliente QA"
    with Session(client.app.state.engine) as session:
        events = session.exec(
            select(IntegrationEvent).where(IntegrationEvent.tenant_id == uuid.UUID(tenant_id))
        ).all()
    sync_event = [event for event in events if event.event_type == "jira_status_sync"][0]
    assert sync_event.payload_summary["mapped"] is False


def test_formula_jira_review_sync_requires_sent_issue() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_id),
        json={},
    ).json()

    synced = client.post(
        f"/api/v1/formula-reviews/{review['id']}/sync",
        headers=headers(USER_A, tenant_id),
    )

    assert synced.status_code == 409
    assert synced.json()["detail"] == "Formula review must be sent to Jira before sync."


def test_jira_send_records_integration_events(monkeypatch) -> None:
    client = make_client()
    fake_jira = FakeJiraClient()
    use_fake_jira_client(monkeypatch, fake_jira)
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    create_oauth_jira_connection(client, tenant_id)
    formula = create_formula(client, tenant_id)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_id),
        json={},
    ).json()

    sent = client.post(
        f"/api/v1/formula-reviews/{review['id']}/jira/send",
        headers=headers(USER_A, tenant_id),
    )

    assert sent.status_code == 200
    with Session(client.app.state.engine) as session:
        events = session.exec(
            select(IntegrationEvent).where(IntegrationEvent.tenant_id == uuid.UUID(tenant_id))
        ).all()
    assert [(event.event_type, event.status) for event in events] == [
        ("jira_review_send", "success")
    ]


def test_formula_jira_review_request_requires_active_connection() -> None:
    client = make_client()
    tenant_id = create_tenant(client, USER_A, "tenant-a")
    formula = create_formula(client, tenant_id)

    response = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_id),
        json={},
    )

    assert response.status_code == 409
    assert response.json()["detail"] == "Active Jira connection is required."


def test_formula_jira_review_requests_are_tenant_scoped() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    create_jira_connection(client, tenant_a, USER_A)
    create_jira_connection(client, tenant_b, USER_B)
    formula = create_formula(client, tenant_a, USER_A)

    cross_create = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_B, tenant_b),
        json={},
    )
    cross_list = client.get(
        f"/api/v1/formulas/{formula['id']}/reviews",
        headers=headers(USER_B, tenant_b),
    )

    assert cross_create.status_code == 404
    assert cross_list.status_code == 404


def test_formula_jira_review_artifacts_are_tenant_scoped() -> None:
    client = make_client()
    tenant_a = create_tenant(client, USER_A, "tenant-a")
    tenant_b = create_tenant(client, USER_B, "tenant-b")
    create_jira_connection(client, tenant_a, USER_A)
    create_jira_connection(client, tenant_b, USER_B)
    formula = create_formula(client, tenant_a, USER_A)
    review = client.post(
        f"/api/v1/formulas/{formula['id']}/reviews/jira",
        headers=headers(USER_A, tenant_a),
        json={},
    ).json()
    artifact = client.post(
        f"/api/v1/formula-reviews/{review['id']}/artifacts/excel",
        headers=headers(USER_A, tenant_a),
    ).json()

    cross_list = client.get(
        f"/api/v1/formula-reviews/{review['id']}/artifacts",
        headers=headers(USER_B, tenant_b),
    )
    cross_download = client.get(
        f"/api/v1/formula-review-artifacts/{artifact['id']}/download",
        headers=headers(USER_B, tenant_b),
    )

    assert cross_list.status_code == 404
    assert cross_download.status_code == 404
