from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from .formula_excel_contract import (
    CALCULATOR_FIRST_ITEM_ROW,
    CALCULATOR_FIRST_PARAMETER_COLUMN,
    CALCULATOR_MATERIAL_COLUMN,
    CALCULATOR_PERCENTAGE_COLUMN,
    CALCULATOR_PRICE_COLUMN,
    CALCULATOR_SHEET_NAME,
    COMPOSITION_FIRST_PARAMETER_ROW,
    COMPOSITION_SHEET_NAME,
    FORMULA_ID_LAB_TEMPLATE_KEY,
    LAB_FIRST_ITEM_ROW,
    LAB_MATERIAL_COLUMN,
    LAB_OBSERVATION_COLUMN,
    LAB_PERCENTAGE_COLUMN,
    LAB_SHEET_NAME,
)


class ExcelImportError(ValueError):
    pass


@dataclass(frozen=True)
class DetectedColumns:
    material_name: str | None
    material_code: str | None
    percentage: str
    material_name_index: int | None
    material_code_index: int | None
    percentage_index: int


@dataclass(frozen=True)
class ParsedFormulaRow:
    row_number: int
    material_name: str | None
    material_code: str | None
    percentage: float | None
    status: str
    message: str | None = None
    price: float | None = None
    parameters: dict[str, float] = field(default_factory=dict)
    lab_material_name: str | None = None
    lab_observation: str | None = None


@dataclass(frozen=True)
class ParsedFormulaImport:
    sheet_name: str
    available_sheets: list[str]
    columns: DetectedColumns
    rows: list[ParsedFormulaRow]
    parser: str = "generic_table"
    formula_name: str | None = None
    parameter_headers: list[str] = field(default_factory=list)
    warnings: list[dict[str, Any]] = field(default_factory=list)

    @property
    def total_percentage(self) -> float:
        return sum(row.percentage or 0 for row in self.rows)


MATERIAL_NAME_HEADERS = {
    "materia prima",
    "ingrediente",
    "raw material",
    "material",
    "mp",
    "componente",
    "nombre",
    "name",
}
MATERIAL_CODE_HEADERS = {
    "codigo",
    "código",
    "code",
    "cod",
    "sku",
    "referencia",
    "reference",
}
PERCENTAGE_HEADERS = {
    "%",
    "porcentaje",
    "percentage",
    "cantidad %",
    "cantidad % peso",
    "cantidad %peso",
    "dosificacion",
    "dosificación",
    "share",
}
COMPACT_LAB_PERCENTAGE_HEADERS = {
    "cantidad %",
    "cantidad % peso",
    "cantidad %peso",
}

COMPACT_LAB_TRIAL_PARSER = "compact_lab_trial"
COMPACT_LAB_HEADER_SEARCH_ROWS = 20
COMPACT_LAB_SCAN_ROWS = 80
COMPACT_LAB_MIN_MATERIAL_ROWS = 2
COMPACT_LAB_ORDER_HEADERS = {
    "orden",
    "orden de adicion",
    "orden de adición",
    "órden de adicion",
    "órden de adición",
    "order",
    "addition order",
}


def _is_atlantica_id_lab_workbook(workbook) -> bool:
    if CALCULATOR_SHEET_NAME not in workbook.sheetnames:
        return False
    worksheet = workbook[CALCULATOR_SHEET_NAME]
    return (
        _normalize_header(worksheet["A1"].value) == "materia prima"
        and _normalize_header(worksheet["B1"].value) == "%"
        and _normalize_header(worksheet["C1"].value) == "precio"
        and bool(_normalize_header(worksheet["D1"].value))
        and _find_total_row(worksheet) is not None
    )


def _parse_atlantica_id_lab_workbook(workbook) -> ParsedFormulaImport:
    worksheet = workbook[CALCULATOR_SHEET_NAME]
    total_row = _find_total_row(worksheet)
    if total_row is None:
        raise ExcelImportError("Could not detect TOTAL row in Calculadora.")

    parameter_headers = _atlantica_parameter_headers(worksheet)
    lab_rows = _atlantica_lab_rows(workbook)
    parsed_rows: list[ParsedFormulaRow] = []
    for row_number in range(CALCULATOR_FIRST_ITEM_ROW, total_row):
        material_name = _cell_text(
            worksheet.cell(row=row_number, column=CALCULATOR_MATERIAL_COLUMN).value
        )
        if material_name is None:
            continue
        percentage_cell = worksheet.cell(row=row_number, column=CALCULATOR_PERCENTAGE_COLUMN)
        percentage = _parse_percentage(percentage_cell.value, percentage_cell.number_format)
        price = _parse_number(worksheet.cell(row=row_number, column=CALCULATOR_PRICE_COLUMN).value)
        parameters = {
            header: _parse_number(worksheet.cell(row=row_number, column=column).value) or 0.0
            for column, header in parameter_headers
        }
        lab_row = lab_rows.get(len(parsed_rows))
        if percentage is None:
            parsed_rows.append(
                ParsedFormulaRow(
                    row_number=row_number,
                    material_name=material_name,
                    material_code=None,
                    percentage=None,
                    status="invalid_percentage",
                    message="Percentage is missing or invalid.",
                    price=price,
                    parameters=parameters,
                    lab_material_name=lab_row["name"] if lab_row else None,
                    lab_observation=lab_row["observation"] if lab_row else None,
                )
            )
            continue
        parsed_rows.append(
            ParsedFormulaRow(
                row_number=row_number,
                material_name=material_name,
                material_code=None,
                percentage=percentage,
                status="parsed",
                price=price,
                parameters=parameters,
                lab_material_name=lab_row["name"] if lab_row else None,
                lab_observation=lab_row["observation"] if lab_row else None,
            )
        )

    if not parsed_rows:
        raise ExcelImportError("No formula rows were found in Calculadora.")

    warnings = _atlantica_warnings(workbook, parsed_rows)
    return ParsedFormulaImport(
        sheet_name=CALCULATOR_SHEET_NAME,
        available_sheets=list(workbook.sheetnames),
        columns=DetectedColumns(
            material_name="materia prima",
            material_code=None,
            percentage="%",
            material_name_index=0,
            material_code_index=None,
            percentage_index=1,
        ),
        rows=parsed_rows,
        parser=FORMULA_ID_LAB_TEMPLATE_KEY,
        formula_name=_atlantica_formula_name(workbook),
        parameter_headers=[header for _, header in parameter_headers],
        warnings=warnings,
    )


def _find_total_row(worksheet) -> int | None:
    for row_number in range(1, worksheet.max_row + 1):
        if _normalize_header(worksheet.cell(row=row_number, column=1).value) == "total":
            return row_number
    return None


def _atlantica_parameter_headers(worksheet) -> list[tuple[int, str]]:
    headers: list[tuple[int, str]] = []
    for column in range(CALCULATOR_FIRST_PARAMETER_COLUMN, worksheet.max_column + 1):
        header = _cell_text(worksheet.cell(row=1, column=column).value)
        if header:
            headers.append((column, header))
    return headers


def _atlantica_lab_rows(workbook) -> dict[int, dict[str, str | None]]:
    if LAB_SHEET_NAME not in workbook.sheetnames:
        return {}
    worksheet = workbook[LAB_SHEET_NAME]
    rows: dict[int, dict[str, str | None]] = {}
    index = 0
    for row_number in range(LAB_FIRST_ITEM_ROW, worksheet.max_row + 1):
        name = _cell_text(worksheet.cell(row=row_number, column=LAB_MATERIAL_COLUMN).value)
        percentage = _parse_number(worksheet.cell(row=row_number, column=LAB_PERCENTAGE_COLUMN).value)
        if name is None or percentage is None:
            continue
        rows[index] = {
            "name": name,
            "observation": _cell_text(
                worksheet.cell(row=row_number, column=LAB_OBSERVATION_COLUMN).value
            ),
        }
        index += 1
    return rows


def _atlantica_formula_name(workbook) -> str | None:
    if LAB_SHEET_NAME not in workbook.sheetnames:
        return None
    return _cell_text(workbook[LAB_SHEET_NAME]["C3"].value)


def _atlantica_warnings(
    workbook,
    rows: list[ParsedFormulaRow],
) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    if COMPOSITION_SHEET_NAME in workbook.sheetnames:
        composition = workbook[COMPOSITION_SHEET_NAME]
        calc_parameters = {
            key.casefold(): value
            for row in rows
            for key, value in row.parameters.items()
        }
        for row_number in range(COMPOSITION_FIRST_PARAMETER_ROW, composition.max_row + 1):
            label = _cell_text(composition.cell(row=row_number, column=1).value)
            if not label:
                continue
            if label.casefold() not in calc_parameters:
                warnings.append(
                    {
                        "code": "composition_parameter_not_in_calculator",
                        "message": f"Composición parameter '{label}' was not found in Calculadora.",
                    }
                )
                break
    return warnings


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, int | float):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", ".")
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def list_formula_xlsx_sheets(content: bytes) -> list[str]:
    workbook = _load_formula_workbook(content)
    return list(workbook.sheetnames)


def parse_formula_xlsx(content: bytes, sheet_name: str | None = None) -> ParsedFormulaImport:
    workbook = _load_formula_workbook(content)
    if _is_atlantica_id_lab_workbook(workbook):
        return _parse_atlantica_id_lab_workbook(workbook)
    if compact_lab_import := _parse_compact_lab_trial_workbook(workbook, sheet_name):
        return compact_lab_import
    worksheet = _select_worksheet(workbook, sheet_name)
    rows = list(worksheet.iter_rows())
    if not rows:
        raise ExcelImportError("XLSX file is empty.")

    header_index, headers = _find_header_row(rows)
    columns = _detect_columns(headers)
    parsed_rows = [
        parsed
        for row in rows[header_index + 1 :]
        if (parsed := _parse_row(row, columns)) is not None
    ]
    if not parsed_rows:
        raise ExcelImportError("No formula rows were found.")

    return ParsedFormulaImport(
        sheet_name=worksheet.title,
        available_sheets=list(workbook.sheetnames),
        columns=columns,
        rows=parsed_rows,
    )


def _parse_compact_lab_trial_workbook(
    workbook,
    sheet_name: str | None,
) -> ParsedFormulaImport | None:
    worksheet = _select_worksheet(workbook, sheet_name)
    table = _find_compact_lab_trial_table(worksheet)
    if table is None:
        return None

    header_row, material_column, percentage_column = table
    parsed_rows: list[ParsedFormulaRow] = []
    for row_number in range(
        header_row + 1,
        min(worksheet.max_row + 1, header_row + COMPACT_LAB_SCAN_ROWS),
    ):
        material_name = _cell_text(worksheet.cell(row=row_number, column=material_column).value)
        percentage_cell = worksheet.cell(row=row_number, column=percentage_column)
        percentage = _parse_percentage(percentage_cell.value, percentage_cell.number_format)

        if material_name is None and percentage is None:
            if parsed_rows:
                break
            continue
        if parsed_rows and _is_compact_lab_metadata_label(material_name):
            break
        if material_name is None:
            continue

        if percentage is None:
            parsed_rows.append(
                ParsedFormulaRow(
                    row_number=row_number,
                    material_name=material_name,
                    material_code=None,
                    percentage=None,
                    status="invalid_percentage",
                    message="Percentage is missing or invalid.",
                )
            )
            continue
        if percentage < 0:
            parsed_rows.append(
                ParsedFormulaRow(
                    row_number=row_number,
                    material_name=material_name,
                    material_code=None,
                    percentage=percentage,
                    status="invalid_percentage",
                    message="Percentage cannot be negative.",
                )
            )
            continue

        parsed_rows.append(
            ParsedFormulaRow(
                row_number=row_number,
                material_name=material_name,
                material_code=None,
                percentage=percentage,
                status="parsed",
            )
        )

    valid_rows = [row for row in parsed_rows if row.percentage is not None]
    if len(valid_rows) < COMPACT_LAB_MIN_MATERIAL_ROWS:
        return None

    percentage_header = _cell_text(
        worksheet.cell(row=header_row, column=percentage_column).value
    )
    return ParsedFormulaImport(
        sheet_name=worksheet.title,
        available_sheets=list(workbook.sheetnames),
        columns=DetectedColumns(
            material_name="material",
            material_code=None,
            percentage=percentage_header or "%",
            material_name_index=material_column - 1,
            material_code_index=None,
            percentage_index=percentage_column - 1,
        ),
        rows=parsed_rows,
        parser=COMPACT_LAB_TRIAL_PARSER,
        formula_name=_compact_lab_formula_name(worksheet, header_row, material_column),
    )


def _find_compact_lab_trial_table(worksheet) -> tuple[int, int, int] | None:
    for row_number in range(
        1,
        min(worksheet.max_row, COMPACT_LAB_HEADER_SEARCH_ROWS) + 1,
    ):
        for column_number in range(1, worksheet.max_column + 1):
            header = _normalize_header(worksheet.cell(row=row_number, column=column_number).value)
            if header not in COMPACT_LAB_PERCENTAGE_HEADERS:
                continue
            material_column = _infer_compact_lab_material_column(
                worksheet,
                row_number,
                column_number,
            )
            if material_column is not None:
                return row_number, material_column, column_number
    return None


def _infer_compact_lab_material_column(
    worksheet,
    header_row: int,
    percentage_column: int,
) -> int | None:
    best_column: int | None = None
    best_ranking = (0, 0, 0)
    for candidate_column in range(1, percentage_column):
        if _normalize_header(
            worksheet.cell(row=header_row, column=candidate_column).value
        ) in COMPACT_LAB_ORDER_HEADERS:
            continue
        score = 0
        material_name_score = 0
        for row_number in range(
            header_row + 1,
            min(worksheet.max_row + 1, header_row + COMPACT_LAB_SCAN_ROWS),
        ):
            material_value = worksheet.cell(row=row_number, column=candidate_column).value
            material_name = _cell_text(material_value)
            percentage_cell = worksheet.cell(row=row_number, column=percentage_column)
            percentage = _parse_percentage(percentage_cell.value, percentage_cell.number_format)
            if material_name is None and percentage is None and score:
                break
            if material_name is not None and percentage is not None:
                score += 1
                if _looks_like_compact_lab_material_name(material_value):
                    material_name_score += 1
        ranking = (material_name_score, score, candidate_column)
        if ranking > best_ranking:
            best_ranking = ranking
            best_column = candidate_column
    if best_ranking[0] < COMPACT_LAB_MIN_MATERIAL_ROWS:
        return None
    return best_column


def _looks_like_compact_lab_material_name(value: Any) -> bool:
    text = _cell_text(value)
    if text is None:
        return False
    return _parse_number(text) is None


def _compact_lab_formula_name(
    worksheet,
    header_row: int,
    material_column: int,
) -> str | None:
    ignored = {"lote", "fecha"}
    for row_number in range(header_row - 1, 0, -1):
        candidates: list[str] = []
        for column_number in range(material_column, worksheet.max_column + 1):
            value = _cell_text(worksheet.cell(row=row_number, column=column_number).value)
            if value and _normalize_header(value) not in ignored:
                candidates.append(value)
        if candidates:
            return max(candidates, key=len)
    return None


def _is_compact_lab_metadata_label(value: str | None) -> bool:
    normalized = _normalize_header(value)
    return normalized in {
        "densidad",
        "densidad experimental",
        "ph",
        "p.h.",
    }


def _load_formula_workbook(content: bytes):
    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl error types vary
        raise ExcelImportError("Could not read XLSX file.") from exc


def _select_worksheet(workbook, sheet_name: str | None):
    if sheet_name is None:
        return workbook.active
    selected = sheet_name.strip()
    if not selected:
        return workbook.active
    if selected not in workbook.sheetnames:
        raise ExcelImportError(f"Worksheet '{selected}' was not found.")
    return workbook[selected]


def _find_header_row(rows: list[tuple[Cell, ...]]) -> tuple[int, dict[int, str]]:
    for index, row in enumerate(rows[:10]):
        headers = {
            position: _normalize_header(cell.value)
            for position, cell in enumerate(row)
            if _normalize_header(cell.value)
        }
        if headers and _detect_columns(headers, require_complete=False).percentage:
            return index, headers
    raise ExcelImportError("Could not detect header row.")


def _detect_columns(
    headers: dict[int, str],
    *,
    require_complete: bool = True,
) -> DetectedColumns:
    material_name_index = _find_column(headers, MATERIAL_NAME_HEADERS)
    material_code_index = _find_column(headers, MATERIAL_CODE_HEADERS)
    percentage_index = _find_column(headers, PERCENTAGE_HEADERS)

    if require_complete:
        if material_name_index is None and material_code_index is None:
            raise ExcelImportError("Could not detect material name or code column.")
        if percentage_index is None:
            raise ExcelImportError("Could not detect percentage column.")

    return DetectedColumns(
        material_name=_column_name(headers, material_name_index),
        material_code=_column_name(headers, material_code_index),
        percentage=_column_name(headers, percentage_index) or "",
        material_name_index=material_name_index,
        material_code_index=material_code_index,
        percentage_index=percentage_index or 0,
    )


def _parse_row(row: tuple[Cell, ...], columns: DetectedColumns) -> ParsedFormulaRow | None:
    material_name = _text_at(row, columns.material_name_index)
    material_code = _text_at(row, columns.material_code_index)
    if material_name is None and material_code is None:
        return None

    percentage_cell = _cell_at(row, columns.percentage_index)
    percentage = _parse_percentage(
        percentage_cell.value if percentage_cell else None,
        percentage_cell.number_format if percentage_cell else None,
    )
    if percentage is None:
        return ParsedFormulaRow(
            row_number=row[0].row,
            material_name=material_name,
            material_code=material_code,
            percentage=None,
            status="invalid_percentage",
            message="Percentage is missing or invalid.",
        )
    if percentage < 0:
        return ParsedFormulaRow(
            row_number=row[0].row,
            material_name=material_name,
            material_code=material_code,
            percentage=percentage,
            status="invalid_percentage",
            message="Percentage cannot be negative.",
        )

    return ParsedFormulaRow(
        row_number=row[0].row,
        material_name=material_name,
        material_code=material_code,
        percentage=percentage,
        status="parsed",
    )


def _find_column(headers: dict[int, str], aliases: set[str]) -> int | None:
    for index, header in headers.items():
        if header in aliases:
            return index
    return None


def _column_name(headers: dict[int, str], index: int | None) -> str | None:
    if index is None:
        return None
    return headers[index]


def _cell_at(row: tuple[Cell, ...], index: int | None) -> Cell | None:
    if index is None or index >= len(row):
        return None
    return row[index]


def _text_at(row: tuple[Cell, ...], index: int | None) -> str | None:
    cell = _cell_at(row, index)
    if cell is None or cell.value is None:
        return None
    value = str(cell.value).strip()
    return value or None


def _parse_percentage(value: Any, number_format: str | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip().replace(",", ".")
        if not stripped:
            return None
        has_percent = stripped.endswith("%")
        stripped = stripped.rstrip("%").strip()
        try:
            parsed = float(stripped)
        except ValueError:
            return None
        return parsed if has_percent else parsed
    if isinstance(value, int | float):
        parsed = float(value)
        if number_format and "%" in number_format and abs(parsed) <= 1:
            return parsed * 100
        return parsed
    return None


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())
