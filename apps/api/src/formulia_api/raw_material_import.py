from __future__ import annotations

import csv
import hashlib
import io
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import date
from difflib import SequenceMatcher
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlmodel import Session, select

from .models import (
    Parameter,
    RawMaterial,
    RawMaterialAlias,
    RawMaterialImport,
    RawMaterialImportRow,
    RawMaterialParameterValue,
    RawMaterialPrice,
)
from .raw_material_master import (
    clean_raw_material_payload,
    ensure_raw_material_identity_available,
    ensure_valid_raw_material_price,
    normalize_name,
)


MATCH_THRESHOLD = 0.82


COLUMN_ALIASES = {
    "name": {
        "materia prima",
        "nombre",
        "descripcion",
        "description",
        "itemname sap",
        "item name sap",
        "material",
        "raw material",
    },
    "code": {
        "codigo",
        "codigo interno",
        "code",
        "id",
    },
    "external_code": {
        "codigo sap",
        "cod sap",
        "sap",
        "sap code",
        "codigo erp",
        "erp code",
        "itemcode sap",
        "item code sap",
    },
    "price": {
        "precio",
        "precio kg",
        "precio eur kg",
        "precio sap",
        "precio sap kg",
        "precio sap eur kg",
        "price",
        "price kg",
        "price eur kg",
    },
    "currency": {
        "moneda",
        "currency",
    },
    "unit": {
        "unidad",
        "unit",
    },
    "supplier": {
        "proveedor",
        "supplier",
        "vendor",
    },
    "family": {
        "familia",
        "family",
    },
    "sap_status": {
        "estado sap",
        "sap status",
        "estado",
        "status",
    },
}


@dataclass(frozen=True)
class ParsedSapRow:
    row_number: int
    raw: dict[str, Any]
    name: str | None
    code: str | None
    external_code: str | None
    price: float | None
    currency: str
    unit: str
    supplier: str | None
    family: str | None
    sap_status: str | None


@dataclass(frozen=True)
class MatchResult:
    material: RawMaterial | None
    match_type: str
    match_score: float | None
    message: str | None = None


def create_sap_import_preview(
    session: Session,
    tenant_id: uuid.UUID,
    *,
    file_name: str,
    content: bytes,
    sheet_name: str | None,
    source: str | None,
    valid_from: date,
) -> RawMaterialImport:
    rows, selected_sheet = parse_sap_file(file_name, content, sheet_name=sheet_name)
    source_name = source or _source_name(file_name)
    import_record = RawMaterialImport(
        tenant_id=tenant_id,
        file_name=file_name,
        source=source_name,
        source_hash=hashlib.sha256(content).hexdigest(),
        status="preview",
        summary_json={
            "import_type": "sap_upload",
            "sheet_name": selected_sheet,
            "valid_from": valid_from.isoformat(),
            "rows": len(rows),
        },
    )
    session.add(import_record)
    session.flush()

    parsed_rows = [parse_sap_row(index, row) for index, row in enumerate(rows, start=2)]
    summary: dict[str, int] = {
        "new_material": 0,
        "price_update": 0,
        "metadata_update": 0,
        "unchanged": 0,
        "needs_review": 0,
        "error": 0,
    }
    for parsed in parsed_rows:
        action, status, material, message, details = classify_sap_row(
            session,
            tenant_id,
            parsed,
        )
        summary[action] = summary.get(action, 0) + 1
        session.add(
            RawMaterialImportRow(
                tenant_id=tenant_id,
                import_id=import_record.id,
                row_number=parsed.row_number,
                raw_material_id=material.id if material else None,
                raw_name=parsed.name,
                action=action,
                status=status,
                raw_row_json={
                    "raw": parsed.raw,
                    "parsed": parsed_row_json(parsed),
                    **details,
                },
                message=message,
            )
        )

    import_record.summary_json = {
        **import_record.summary_json,
        **summary,
    }
    session.add(import_record)
    session.commit()
    session.refresh(import_record)
    return import_record


def parse_sap_file(
    file_name: str,
    content: bytes,
    *,
    sheet_name: str | None,
) -> tuple[list[dict[str, Any]], str | None]:
    lowered = file_name.lower()
    if lowered.endswith(".csv"):
        return _parse_csv(content), None
    if lowered.endswith(".xlsx") or lowered.endswith(".xlsm"):
        return _parse_xlsx(content, sheet_name=sheet_name)
    raise HTTPException(status_code=400, detail="Only .xlsx, .xlsm, and .csv files are supported.")


def parse_sap_row(row_number: int, row: dict[str, Any]) -> ParsedSapRow:
    mapped = map_sap_columns(row)
    price = parse_decimal(mapped.get("price"))
    currency = clean_text(mapped.get("currency")) or "EUR"
    unit = clean_text(mapped.get("unit")) or "kg"
    return ParsedSapRow(
        row_number=row_number,
        raw={str(key): _json_value(value) for key, value in row.items()},
        name=clean_text(mapped.get("name")),
        code=clean_text(mapped.get("code")),
        external_code=clean_text(mapped.get("external_code")),
        price=price,
        currency=currency,
        unit=unit,
        supplier=clean_text(mapped.get("supplier")),
        family=clean_text(mapped.get("family")),
        sap_status=clean_text(mapped.get("sap_status")),
    )


def map_sap_columns(row: dict[str, Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {}
    for key, value in row.items():
        canonical = canonical_column_name(str(key))
        if canonical and canonical not in mapped:
            mapped[canonical] = value
    return mapped


def canonical_column_name(header: str) -> str | None:
    normalized = normalize_header(header)
    for name, aliases in COLUMN_ALIASES.items():
        if normalized in aliases:
            return name
    return None


def classify_sap_row(
    session: Session,
    tenant_id: uuid.UUID,
    parsed: ParsedSapRow,
) -> tuple[str, str, RawMaterial | None, str | None, dict[str, Any]]:
    if not parsed.name and not parsed.external_code and not parsed.code:
        return "error", "error", None, "Row needs a material name, code, or SAP code.", {}
    if parsed.price is not None and parsed.price < 0:
        return "error", "error", None, "Raw material price cannot be negative.", {}

    match = match_sap_row(session, tenant_id, parsed)
    details = {
        "match": {
            "type": match.match_type,
            "score": match.match_score,
            "message": match.message,
            "raw_material_id": str(match.material.id) if match.material else None,
        },
        "diff": {},
    }
    if match.match_type == "fuzzy":
        return "needs_review", "needs_review", match.material, match.message, details
    if match.material is None:
        if not parsed.name:
            return "error", "error", None, "New material rows need a name.", details
        return "new_material", "ready", None, "New material detected.", details

    diff = material_diff(session, tenant_id, match.material, parsed)
    details["diff"] = diff
    if diff.get("price"):
        return "price_update", "ready", match.material, "Price update detected.", details
    metadata_changes = {key: value for key, value in diff.items() if key != "price"}
    if metadata_changes:
        return "metadata_update", "ready", match.material, "Metadata update detected.", details
    return "unchanged", "skipped", match.material, "No changes detected.", details


def match_sap_row(session: Session, tenant_id: uuid.UUID, parsed: ParsedSapRow) -> MatchResult:
    materials = session.exec(
        select(RawMaterial).where(RawMaterial.tenant_id == tenant_id)
    ).all()
    aliases = session.exec(
        select(RawMaterialAlias).where(RawMaterialAlias.tenant_id == tenant_id)
    ).all()
    by_alias = {normalize_name(alias.alias): alias.raw_material_id for alias in aliases}
    by_id = {material.id: material for material in materials}

    external_code = identity_key(parsed.external_code)
    if external_code:
        for material in materials:
            if identity_key(material.external_code) == external_code:
                return MatchResult(material, "external_code", 1.0)

    code = identity_key(parsed.code)
    if code:
        for material in materials:
            if identity_key(material.code) == code:
                return MatchResult(material, "code", 1.0)

    if parsed.name:
        normalized = normalize_name(parsed.name)
        for material in materials:
            if normalize_name(material.name) == normalized:
                return MatchResult(material, "name", 1.0)
        alias_material_id = by_alias.get(normalized)
        if alias_material_id and alias_material_id in by_id:
            return MatchResult(by_id[alias_material_id], "alias", 1.0)
        fuzzy = best_fuzzy_match(materials, normalized)
        if fuzzy:
            material, score = fuzzy
            return MatchResult(
                material,
                "fuzzy",
                score,
                f"Possible match: {material.name} ({score:.2f}).",
            )

    return MatchResult(None, "none", None)


def best_fuzzy_match(
    materials: list[RawMaterial],
    normalized_name: str,
) -> tuple[RawMaterial, float] | None:
    best: tuple[RawMaterial, float] | None = None
    for material in materials:
        score = SequenceMatcher(None, normalized_name, normalize_name(material.name)).ratio()
        if score >= MATCH_THRESHOLD and (best is None or score > best[1]):
            best = (material, score)
    return best


def material_diff(
    session: Session,
    tenant_id: uuid.UUID,
    material: RawMaterial,
    parsed: ParsedSapRow,
) -> dict[str, Any]:
    diff: dict[str, Any] = {}
    if parsed.price is not None:
        current_price = session.exec(
            select(RawMaterialPrice)
            .where(
                RawMaterialPrice.tenant_id == tenant_id,
                RawMaterialPrice.raw_material_id == material.id,
            )
            .order_by(RawMaterialPrice.valid_from.desc(), RawMaterialPrice.created_at.desc())
        ).first()
        if current_price is None or abs(current_price.price - parsed.price) > 0.000001:
            diff["price"] = {
                "before": current_price.price if current_price else None,
                "after": parsed.price,
            }
    if parsed.external_code and not material.external_code:
        diff["external_code"] = {"before": None, "after": parsed.external_code}
    if parsed.family and parsed.family != material.family:
        diff["family"] = {"before": material.family, "after": parsed.family}
    if parsed.name and normalize_name(parsed.name) != normalize_name(material.name):
        diff["name"] = {"before": material.name, "after": parsed.name}
    return diff


def apply_sap_import(
    session: Session,
    tenant_id: uuid.UUID,
    import_record: RawMaterialImport,
) -> RawMaterialImport:
    if import_record.status == "applied":
        return import_record
    rows = session.exec(
        select(RawMaterialImportRow)
        .where(
            RawMaterialImportRow.tenant_id == tenant_id,
            RawMaterialImportRow.import_id == import_record.id,
        )
        .order_by(RawMaterialImportRow.row_number)
    ).all()
    summary = dict(import_record.summary_json or {})
    applied = 0
    errored = 0
    skipped = 0
    for row in rows:
        if row.status != "ready":
            skipped += 1
            continue
        try:
            apply_sap_import_row(session, tenant_id, import_record, row)
            row.status = "applied"
            row.message = "Applied."
            applied += 1
        except HTTPException as exc:
            row.status = "error"
            row.message = str(exc.detail)
            errored += 1
        session.add(row)
    import_record.status = "applied" if errored == 0 else "applied_with_errors"
    import_record.summary_json = {
        **summary,
        "applied_rows": applied,
        "error_rows": errored,
        "skipped_rows": skipped,
    }
    session.add(import_record)
    session.commit()
    session.refresh(import_record)
    return import_record


def apply_sap_import_row(
    session: Session,
    tenant_id: uuid.UUID,
    import_record: RawMaterialImport,
    row: RawMaterialImportRow,
) -> None:
    parsed = parsed_row_from_json(row.raw_row_json.get("parsed") or {}, row.row_number)
    if row.action == "new_material":
        material = create_material_from_sap_row(session, tenant_id, parsed)
        row.raw_material_id = material.id
        upsert_price_from_sap_row(session, tenant_id, material.id, parsed, import_record)
        return
    if row.raw_material_id is None:
        raise HTTPException(status_code=400, detail="Import row has no matched raw material.")
    material = session.get(RawMaterial, row.raw_material_id)
    if material is None or material.tenant_id != tenant_id:
        raise HTTPException(status_code=404, detail="Raw material not found.")
    if row.action in {"price_update", "metadata_update"}:
        update_material_metadata_from_sap_row(session, tenant_id, material, parsed)
    if row.action == "price_update":
        upsert_price_from_sap_row(session, tenant_id, material.id, parsed, import_record)


def create_material_from_sap_row(
    session: Session,
    tenant_id: uuid.UUID,
    parsed: ParsedSapRow,
) -> RawMaterial:
    if not parsed.name:
        raise HTTPException(status_code=400, detail="New material rows need a name.")
    values = clean_raw_material_payload(
        {
            "code": parsed.code or parsed.external_code,
            "external_code": parsed.external_code,
            "name": parsed.name,
            "family": parsed.family,
            "notes": None,
        }
    )
    normalized_name = normalize_name(values["name"])
    ensure_raw_material_identity_available(
        session,
        tenant_id,
        code=values.get("code"),
        external_code=values.get("external_code"),
        normalized_name=normalized_name,
    )
    material = RawMaterial(
        tenant_id=tenant_id,
        normalized_name=normalized_name,
        **values,
    )
    session.add(material)
    session.flush()
    ensure_zero_values_for_raw_material(session, tenant_id, material)
    return material


def update_material_metadata_from_sap_row(
    session: Session,
    tenant_id: uuid.UUID,
    material: RawMaterial,
    parsed: ParsedSapRow,
) -> None:
    updates: dict[str, Any] = {}
    if parsed.external_code and not material.external_code:
        updates["external_code"] = parsed.external_code
    if parsed.family and parsed.family != material.family:
        updates["family"] = parsed.family
    if not updates:
        return
    ensure_raw_material_identity_available(
        session,
        tenant_id,
        code=material.code,
        external_code=updates.get("external_code", material.external_code),
        normalized_name=material.normalized_name,
        exclude_raw_material_id=material.id,
    )
    for key, value in updates.items():
        setattr(material, key, value)
    session.add(material)


def upsert_price_from_sap_row(
    session: Session,
    tenant_id: uuid.UUID,
    raw_material_id: uuid.UUID,
    parsed: ParsedSapRow,
    import_record: RawMaterialImport,
) -> None:
    if parsed.price is None:
        return
    ensure_valid_raw_material_price(parsed.price)
    valid_from = date.fromisoformat(str(import_record.summary_json["valid_from"]))
    existing = session.exec(
        select(RawMaterialPrice).where(
            RawMaterialPrice.tenant_id == tenant_id,
            RawMaterialPrice.raw_material_id == raw_material_id,
            RawMaterialPrice.source == import_record.source,
            RawMaterialPrice.valid_from == valid_from,
        )
    ).first()
    if existing is None:
        session.add(
            RawMaterialPrice(
                tenant_id=tenant_id,
                raw_material_id=raw_material_id,
                price=parsed.price,
                currency=parsed.currency,
                unit=parsed.unit,
                supplier=parsed.supplier,
                source=import_record.source,
                valid_from=valid_from,
            )
        )
        return
    existing.price = parsed.price
    existing.currency = parsed.currency
    existing.unit = parsed.unit
    existing.supplier = parsed.supplier
    session.add(existing)


def ensure_zero_values_for_raw_material(
    session: Session,
    tenant_id: uuid.UUID,
    material: RawMaterial,
) -> None:
    parameters = session.exec(
        select(Parameter).where(
            Parameter.tenant_id == tenant_id,
            Parameter.is_active.is_(True),
        )
    ).all()
    for parameter in parameters:
        exists = session.exec(
            select(RawMaterialParameterValue).where(
                RawMaterialParameterValue.tenant_id == tenant_id,
                RawMaterialParameterValue.raw_material_id == material.id,
                RawMaterialParameterValue.parameter_id == parameter.id,
            )
        ).first()
        if exists is not None:
            continue
        session.add(
            RawMaterialParameterValue(
                tenant_id=tenant_id,
                raw_material_id=material.id,
                parameter_id=parameter.id,
                value=0,
                source="default_zero",
            )
        )


def load_import_rows(
    session: Session,
    tenant_id: uuid.UUID,
    import_id: uuid.UUID,
) -> list[RawMaterialImportRow]:
    return session.exec(
        select(RawMaterialImportRow)
        .where(
            RawMaterialImportRow.tenant_id == tenant_id,
            RawMaterialImportRow.import_id == import_id,
        )
        .order_by(RawMaterialImportRow.row_number)
    ).all()


def import_read(import_record: RawMaterialImport, rows: list[RawMaterialImportRow]) -> dict[str, Any]:
    return {
        "id": import_record.id,
        "tenant_id": import_record.tenant_id,
        "file_name": import_record.file_name,
        "source": import_record.source,
        "source_hash": import_record.source_hash,
        "status": import_record.status,
        "summary_json": import_record.summary_json,
        "created_at": import_record.created_at,
        "rows": [
            {
                "id": row.id,
                "tenant_id": row.tenant_id,
                "import_id": row.import_id,
                "row_number": row.row_number,
                "raw_material_id": row.raw_material_id,
                "raw_name": row.raw_name,
                "action": row.action,
                "status": row.status,
                "raw_row_json": row.raw_row_json,
                "message": row.message,
            }
            for row in rows
        ],
    }


def parsed_row_json(parsed: ParsedSapRow) -> dict[str, Any]:
    return {
        "row_number": parsed.row_number,
        "name": parsed.name,
        "code": parsed.code,
        "external_code": parsed.external_code,
        "price": parsed.price,
        "currency": parsed.currency,
        "unit": parsed.unit,
        "supplier": parsed.supplier,
        "family": parsed.family,
        "sap_status": parsed.sap_status,
    }


def parsed_row_from_json(payload: dict[str, Any], row_number: int) -> ParsedSapRow:
    return ParsedSapRow(
        row_number=row_number,
        raw={},
        name=payload.get("name"),
        code=payload.get("code"),
        external_code=payload.get("external_code"),
        price=payload.get("price"),
        currency=payload.get("currency") or "EUR",
        unit=payload.get("unit") or "kg",
        supplier=payload.get("supplier"),
        family=payload.get("family"),
        sap_status=payload.get("sap_status"),
    )


def normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    chars = [character.lower() if character.isalnum() else " " for character in ascii_value]
    return " ".join("".join(chars).split())


def identity_key(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = " ".join(value.strip().casefold().split())
    return cleaned or None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = " ".join(str(value).replace("\xa0", " ").strip().split())
    return cleaned or None


def parse_decimal(value: Any) -> float | None:
    cleaned = clean_text(value)
    if cleaned is None:
        return None
    text = (
        cleaned.replace("EUR", "")
        .replace("eur", "")
        .replace("€", "")
        .replace(" ", "")
    )
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    sample = text[:2048]
    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes, *, sheet_name: str | None) -> tuple[list[dict[str, Any]], str | None]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], sheet.title
    headers = [clean_text(value) or f"Column {index + 1}" for index, value in enumerate(rows[0])]
    parsed_rows = []
    for values in rows[1:]:
        if values is None or all(value in (None, "") for value in values):
            continue
        parsed_rows.append(
            {
                headers[index]: value
                for index, value in enumerate(values)
                if index < len(headers)
            }
        )
    return parsed_rows, sheet.title


def _source_name(file_name: str) -> str:
    return file_name.rsplit(".", 1)[0]


def _json_value(value: Any) -> Any:
    if isinstance(value, date):
        return value.isoformat()
    return value
