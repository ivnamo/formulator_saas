from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


class IsoLegacyImportError(ValueError):
    pass


@dataclass(frozen=True)
class IsoLegacyImportRow:
    format_key: str
    sheet_name: str
    row_number: int
    record_key: str | None
    action: str
    status: str
    message: str | None
    payload: dict[str, Any]


@dataclass(frozen=True)
class IsoLegacyImport:
    format_key: str
    available_sheets: list[str]
    rows: list[IsoLegacyImportRow]


F10_01_ALIASES = {
    "iso_request_number": {"nº solicitud", "no solicitud", "n solicitud", "solicitud"},
    "legacy_id": {"id"},
    "requester": {"solicitante"},
    "product_name": {"nombre", "producto"},
    "commercial_name": {"nom_comercial", "nombre comercial"},
    "need": {"necesidad"},
    "product_type": {"tipo de producto", "tipo producto"},
    "packaging": {"envase"},
    "destination_country": {"pais destino", "país destino"},
    "accepted": {"aceptado"},
    "finished": {"finalizado"},
    "rejection_reason": {"motivo denegado", "motivo rechazo"},
    "approved_at": {"fecha de aprobacion sol", "fecha de aprobación sol"},
    "planned_finish_at": {"fecha finalizacion estimada", "fecha finalización estimada"},
    "finished_at": {"fecha de finalizacion real", "fecha de finalización real"},
    "estimated_days": {"tiempo estimado"},
}


def parse_iso_legacy_xlsx(
    format_key: str,
    content: bytes,
    *,
    sheet_name: str | None = None,
) -> IsoLegacyImport:
    workbook = _load_workbook(content)
    selected_sheets = _selected_sheets(workbook, sheet_name)
    if format_key == "f10_01":
        rows = [row for sheet in selected_sheets for row in _parse_f10_01_sheet(sheet)]
    elif format_key == "f10_02":
        rows = [row for sheet in selected_sheets for row in _parse_f10_02_sheet(sheet)]
    elif format_key == "f10_03":
        rows = [row for sheet in selected_sheets for row in _parse_f10_03_sheet(sheet)]
    else:
        raise IsoLegacyImportError(f"Unsupported ISO legacy format {format_key!r}.")
    return IsoLegacyImport(
        format_key=format_key,
        available_sheets=list(workbook.sheetnames),
        rows=rows,
    )


def _load_workbook(content: bytes):
    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl error types vary
        raise IsoLegacyImportError("Could not read ISO legacy XLSX file.") from exc


def _selected_sheets(workbook, sheet_name: str | None) -> list[Worksheet]:
    selected = (sheet_name or "").strip()
    if selected:
        if selected not in workbook.sheetnames:
            raise IsoLegacyImportError(f"Worksheet '{selected}' was not found.")
        return [workbook[selected]]
    return list(workbook.worksheets)


def _parse_f10_01_sheet(sheet: Worksheet) -> list[IsoLegacyImportRow]:
    rows = list(sheet.iter_rows())
    header_index, headers = _find_header_row(rows, F10_01_ALIASES)
    year = _year_from_sheet(sheet.title)
    parsed: list[IsoLegacyImportRow] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if _row_is_empty(row):
            continue
        values = {field: _cell_text(row, index) for field, index in headers.items()}
        product_name = values.get("product_name")
        request_number = values.get("iso_request_number")
        legacy_id = values.get("legacy_id")
        record_key = _clean_text(request_number) or (
            f"ID-{legacy_id}" if _clean_text(legacy_id) else None
        )
        parsed_year = _year_from_request(record_key) or year
        payload = {
            "project": {
                "iso_request_number": record_key,
                "year": parsed_year,
                "project_code": _clean_text(legacy_id),
                "requester": _clean_text(values.get("requester")),
                "product_name": _clean_text(product_name),
                "commercial_name": _clean_text(values.get("commercial_name")),
                "need": _clean_text(values.get("need")),
                "product_type": _clean_text(values.get("product_type")),
                "destination_country": _clean_text(values.get("destination_country")),
                "packaging": _clean_text(values.get("packaging")),
                "accepted_status": _accepted_status(values.get("accepted")),
                "lifecycle_status": _lifecycle_status(values.get("finished")),
                "rejection_reason": _clean_text(values.get("rejection_reason")),
                "planned_finish_at": _date_text(values.get("planned_finish_at")),
                "finished_at": _date_text(values.get("finished_at")),
                "estimated_days": _clean_text(values.get("estimated_days")),
                "source_type": "legacy_f10_01",
                "source_ref": _source_ref(sheet.title, row_number),
            },
            "legacy": values,
        }
        missing = []
        if not record_key:
            missing.append("No Solicitud/ID")
        if not _clean_text(product_name):
            missing.append("Producto")
        status = "ready" if not missing else "ambiguous"
        parsed.append(
            IsoLegacyImportRow(
                format_key="f10_01",
                sheet_name=sheet.title,
                row_number=row_number,
                record_key=record_key,
                action="upsert_project" if status == "ready" else "skip",
                status=status,
                message=f"Missing {', '.join(missing)}." if missing else None,
                payload=payload,
            )
        )
    return parsed


def _parse_f10_02_sheet(sheet: Worksheet) -> list[IsoLegacyImportRow]:
    rows = list(sheet.iter_rows())
    header = _project_header_from_sheet(sheet)
    trials: list[IsoLegacyImportRow] = []
    for index, row in enumerate(rows):
        label = _cell_text(row, 0)
        match = re.fullmatch(r"ensayo\s+(\d+)", _normalize(label))
        if not match:
            continue
        trial_number = int(match.group(1))
        data_row_index = index + 2
        if data_row_index >= len(rows):
            continue
        data_row = rows[data_row_index]
        trial_code = _cell_text(data_row, 0)
        trial_name = _cell_text(data_row, 1)
        trial_date = _cell_text(data_row, 2)
        raw_result = _cell_text(data_row, 3)
        composition, reason = _trial_composition(rows[data_row_index + 1 :])
        payload = {
            "project": header,
            "trial": {
                "trial_number": trial_number,
                "trial_code": _clean_text(trial_code),
                "trial_name": _clean_text(trial_name),
                "trial_at_raw": _clean_text(trial_date),
                "technical_result": _technical_result(raw_result),
                "raw_result_label": _clean_text(raw_result),
                "result_source": "legacy_import",
                "reason_comment": _clean_text(reason),
                "snapshot": {
                    "source": "legacy_f10_02",
                    "source_ref": _source_ref(sheet.title, data_row[0].row),
                    "composition": composition,
                    "raw_trial_date": _clean_text(trial_date),
                },
            },
        }
        missing = []
        if not header.get("iso_request_number"):
            missing.append("No Solicitud")
        if not header.get("product_name"):
            missing.append("Producto")
        if not _clean_text(trial_code) and not _clean_text(trial_name):
            missing.append("Ensayo")
        status = "ready" if not missing else "ambiguous"
        trials.append(
            IsoLegacyImportRow(
                format_key="f10_02",
                sheet_name=sheet.title,
                row_number=data_row[0].row,
                record_key=_clean_text(trial_code) or _clean_text(trial_name),
                action="upsert_trial" if status == "ready" else "skip",
                status=status,
                message=f"Missing {', '.join(missing)}." if missing else None,
                payload=payload,
            )
        )
    if not trials:
        trials.append(
            IsoLegacyImportRow(
                format_key="f10_02",
                sheet_name=sheet.title,
                row_number=1,
                record_key=header.get("iso_request_number"),
                action="skip",
                status="ambiguous",
                message="No trial blocks were found.",
                payload={"project": header},
            )
        )
    return trials


def _parse_f10_03_sheet(sheet: Worksheet) -> list[IsoLegacyImportRow]:
    rows = list(sheet.iter_rows())
    header = _project_header_from_sheet(sheet)
    formula_ok = _value_next_to_label(rows, "formula ok")
    specification = _f10_03_specification(rows)
    validation_at = _value_next_to_label(rows, "fecha validacion")
    checks = _f10_03_checks(rows)
    payload = {
        "project": header,
        "validation": {
            "formula_ok": _clean_text(formula_ok),
            "specification": specification,
            "validation_at_raw": _clean_text(validation_at),
            "validation_checks": checks,
            "comments": None,
            "source_ref": _source_ref(sheet.title, 1),
        },
    }
    missing = []
    if not header.get("iso_request_number"):
        missing.append("No Solicitud")
    if not header.get("product_name"):
        missing.append("Producto")
    if not _clean_text(formula_ok):
        missing.append("Formula OK")
    status = "ready" if not missing else "ambiguous"
    return [
        IsoLegacyImportRow(
            format_key="f10_03",
            sheet_name=sheet.title,
            row_number=1,
            record_key=header.get("iso_request_number"),
            action="upsert_validation" if status == "ready" else "skip",
            status=status,
            message=f"Missing {', '.join(missing)}." if missing else None,
            payload=payload,
        )
    ]


def _find_header_row(
    rows: list[tuple[Cell, ...]],
    aliases: dict[str, set[str]],
) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows[:12]):
        normalized = {
            position: _normalize(cell.value)
            for position, cell in enumerate(row)
            if _normalize(cell.value)
        }
        detected: dict[str, int] = {}
        for field, names in aliases.items():
            for position, header in normalized.items():
                if header in names:
                    detected[field] = position
                    break
        if "product_name" in detected or "need" in detected:
            return index, detected
    raise IsoLegacyImportError("Could not detect ISO legacy header row.")


def _project_header_from_sheet(sheet: Worksheet) -> dict[str, Any]:
    rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12)))
    request_number = (
        _value_next_to_label(rows, "n solicitud")
        or _value_next_to_label(rows, "no solicitud")
        or _value_next_to_label(rows, "numero solicitud")
    )
    product = _value_next_to_label(rows, "producto")
    responsible = _value_next_to_label(rows, "responsable")
    need = None
    for index, row in enumerate(rows):
        if _normalize(_cell_text(row, 0)).startswith("1 datos de partida"):
            next_row = rows[index + 1] if index + 1 < len(rows) else ()
            need = _cell_text(next_row, 0) if next_row else None
            break
    year = _year_from_request(request_number) or _year_from_sheet(sheet.title)
    return {
        "iso_request_number": _clean_text(request_number),
        "year": year,
        "project_code": _clean_text(_sheet_project_code(sheet.title)),
        "requester": _clean_text(responsible),
        "product_name": _clean_text(product),
        "need": _clean_text(need),
        "accepted_status": "accepted",
        "lifecycle_status": "in_progress",
        "source_type": "legacy_import",
        "source_ref": _source_ref(sheet.title, 1),
    }


def _trial_composition(rows: list[tuple[Cell, ...]]) -> tuple[list[dict[str, Any]], str | None]:
    composition: list[dict[str, Any]] = []
    reason = None
    collecting = False
    for row in rows:
        first = _cell_text(row, 0)
        normalized = _normalize(first)
        if re.fullmatch(r"ensayo\s+\d+", normalized) or normalized.startswith("3 verificacion"):
            break
        if normalized == "materia prima":
            collecting = True
            continue
        if normalized.startswith("motivo comentario"):
            reason = _cell_text(row, 1)
            break
        if collecting and first:
            composition.append(
                {
                    "material_name": _clean_text(first),
                    "percentage": _clean_text(_cell_text(row, 1)),
                }
            )
    return composition, reason


def _f10_03_specification(rows: list[tuple[Cell, ...]]) -> dict[str, Any]:
    specification: dict[str, Any] = {}
    in_parameters = False
    for row in rows:
        first = _cell_text(row, 0)
        second = _cell_text(row, 1)
        normalized = _normalize(first)
        if normalized.startswith("2 validacion"):
            break
        if normalized in {"descripcion", "aspecto", "color"}:
            specification[normalized] = _clean_text(second)
        if normalized == "parametro":
            in_parameters = True
            continue
        if in_parameters and first:
            specification[_clean_text(first) or first] = _clean_text(second)
    return specification


def _f10_03_checks(rows: list[tuple[Cell, ...]]) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    collecting = False
    for row in rows:
        first = _cell_text(row, 0)
        if _normalize(first) == "area":
            collecting = True
            continue
        if collecting:
            area = _clean_text(first)
            aspect = _clean_text(_cell_text(row, 1))
            if not area and not aspect:
                continue
            checks.append(
                {
                    "area": area or "-",
                    "aspect": aspect or "-",
                    "required": True,
                    "result": _validation_result(_cell_text(row, 2)),
                    "comments": _clean_text(_cell_text(row, 3)),
                }
            )
    return checks


def _value_next_to_label(rows: list[tuple[Cell, ...]], label: str) -> str | None:
    normalized_label = _normalize(label)
    for row in rows:
        first = _normalize(_cell_text(row, 0))
        if first.rstrip(":") == normalized_label:
            return _cell_text(row, 1)
    return None


def _accepted_status(value: Any) -> str:
    normalized = _normalize(value)
    if normalized in {"si", "sí", "yes"}:
        return "accepted"
    if normalized in {"no", "no."}:
        return "rejected"
    return "pending"


def _lifecycle_status(value: Any) -> str:
    normalized = _normalize(value)
    if normalized in {"si", "sí", "yes"}:
        return "finished"
    if normalized in {"no", "no."}:
        return "in_progress"
    return "intake"


def _technical_result(value: Any) -> str:
    normalized = _normalize(value)
    if normalized == "liberado":
        return "LIBERADO"
    if normalized in {"ok no liberado", "ok_no_liberado"}:
        return "OK_NO_LIBERADO"
    if normalized == "nok":
        return "NOK"
    if normalized == "iterado":
        return "ITERADO"
    if normalized == "abandonado":
        return "ABANDONADO"
    if normalized.startswith("cancel"):
        return "CANCELADO"
    return "pending_result"


def _validation_result(value: Any) -> str:
    normalized = _normalize(value)
    if normalized == "ok":
        return "ok"
    if normalized == "nok":
        return "nok"
    if normalized in {"na", "n/a", "no aplica", "not applicable"}:
        return "not_applicable"
    return "pending"


def _year_from_request(value: Any) -> int | None:
    text = str(value or "")
    match = re.search(r"20\d{2}", text)
    return int(match.group(0)) if match else None


def _year_from_sheet(value: str) -> int | None:
    match = re.search(r"20\d{2}", value)
    return int(match.group(0)) if match else None


def _sheet_project_code(value: str) -> str | None:
    return _clean_text(value)


def _date_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return _clean_text(value)


def _source_ref(sheet_name: str, row_number: int) -> str:
    return f"{sheet_name}!{row_number}"


def _row_is_empty(row: tuple[Cell, ...]) -> bool:
    return all(_clean_text(cell.value) is None for cell in row)


def _cell_text(row: tuple[Cell, ...], index: int) -> str | None:
    if index >= len(row):
        return None
    return _clean_text(row[index].value)


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = " ".join(str(value).replace("\xa0", " ").strip().split())
    return cleaned or None


def _normalize(value: Any) -> str:
    text = _clean_text(value) or ""
    text = text.casefold().strip().rstrip(":")
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ü": "u",
        "º": "",
        "ª": "",
        "—": " ",
        "-": " ",
        "/": " ",
    }
    for source, replacement in replacements.items():
        text = text.replace(source, replacement)
    return " ".join(re.sub(r"[^a-z0-9%]+", " ", text).split())
