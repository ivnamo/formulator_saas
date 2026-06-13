from __future__ import annotations

import hashlib
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .models import IsoDesignProject, IsoDesignTrial, IsoProductValidation


ISO_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ISO_ZIP_CONTENT_TYPE = "application/zip"


@dataclass(frozen=True)
class IsoGeneratedArtifact:
    artifact_type: str
    file_name: str
    content_type: str
    checksum_sha256: str
    size_bytes: int
    content: bytes


def build_f10_01_excel(
    projects: list[IsoDesignProject],
    *,
    year: int | None = None,
) -> IsoGeneratedArtifact:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "F10-01"
    _append_title(sheet, "F10-01 Viabilidad y planificacion de disenos")
    _append_key_values(sheet, [("Year", year or "Todos"), ("Project count", len(projects))])
    _append_header(
        sheet,
        [
            "No Solicitud",
            "Ano",
            "ProyectoID",
            "Solicitante",
            "Producto",
            "Nombre comercial",
            "Necesidad",
            "Tipo producto",
            "Pais destino",
            "Envase",
            "Aceptado",
            "Estado ciclo",
            "Fin planificado",
            "Finalizado",
            "Horas I+D",
            "Horas Calidad",
            "Problemas",
            "Comentarios",
        ],
    )
    for project in projects:
        sheet.append(
            [
                project.iso_request_number,
                project.year,
                project.project_code,
                project.requester,
                project.product_name,
                project.commercial_name,
                project.need,
                project.product_type,
                project.destination_country,
                project.packaging,
                project.accepted_status,
                project.lifecycle_status,
                _cell_value(project.planned_finish_at),
                _cell_value(project.finished_at),
                project.rd_hours,
                project.quality_hours,
                project.problems,
                project.comments,
            ]
        )
    if not projects:
        sheet.append(["Sin proyectos ISO", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    _finish_sheet(sheet)
    suffix = str(year) if year else "all"
    return _artifact_from_workbook(
        workbook,
        artifact_type="iso_f10_01_xlsx",
        file_name=f"formulia_iso_f10_01_{suffix}.xlsx",
    )


def build_f10_02_excel(
    project: IsoDesignProject,
    trials: list[IsoDesignTrial],
) -> IsoGeneratedArtifact:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Proyecto"
    _append_title(summary, "F10-02 Diseno de producto")
    _append_project_key_values(summary, project)
    _finish_sheet(summary)

    sheet = workbook.create_sheet("F10-02")
    _append_header(
        sheet,
        [
            "Ensayo",
            "Codigo ensayo",
            "Formula",
            "Version",
            "Jira",
            "Estado Jira",
            "Resultado I+D",
            "Resultado tecnico",
            "Origen",
            "Fecha ensayo",
            "Comentario",
            "Snapshot checksum",
        ],
    )
    for trial in trials:
        sheet.append(
            [
                trial.trial_number,
                trial.trial_code,
                trial.trial_name,
                trial.formula_version,
                trial.jira_issue_key,
                trial.raw_status_label,
                trial.raw_result_label,
                trial.technical_result,
                trial.result_source,
                _cell_value(trial.trial_at),
                trial.reason_comment,
                trial.snapshot_checksum,
            ]
        )
    if not trials:
        sheet.append(["Sin ensayos", "", "", "", "", "", "", "", "", "", "", ""])
    _finish_sheet(sheet)
    return _artifact_from_workbook(
        workbook,
        artifact_type="iso_f10_02_xlsx",
        file_name=f"formulia_iso_f10_02_{_project_file_part(project)}.xlsx",
    )


def build_f10_03_excel(
    project: IsoDesignProject,
    validation: IsoProductValidation,
    released_trial: IsoDesignTrial | None,
) -> IsoGeneratedArtifact:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "F10-03"
    _append_title(summary, "F10-03 Validacion final de producto")
    _append_key_values(
        summary,
        [
            ("No Solicitud", project.iso_request_number),
            ("ProyectoID", project.project_code),
            ("Producto", validation.product_name),
            ("Formula OK", validation.formula_ok),
            ("Estado validacion", validation.status),
            ("Fecha validacion", _cell_value(validation.validation_at)),
            ("Fecha publicacion", _cell_value(validation.published_at)),
            ("Ensayo liberado", released_trial.trial_code if released_trial else validation.released_trial_id),
            ("Resultado tecnico", released_trial.technical_result if released_trial else None),
            ("Comentarios", validation.comments),
        ],
    )
    _finish_sheet(summary)

    checks = workbook.create_sheet("Validaciones")
    _append_header(checks, ["Area", "Aspecto", "Requerido", "Resultado", "Comentarios"])
    for check in _sequence(validation.validation_checks_json):
        check_data = _mapping(check)
        checks.append(
            [
                check_data.get("area"),
                check_data.get("aspect"),
                bool(check_data.get("required", True)),
                check_data.get("result"),
                check_data.get("comments"),
            ]
        )
    if checks.max_row == 1:
        checks.append(["Sin checks", "", "", "", ""])
    _finish_sheet(checks)

    spec = workbook.create_sheet("Especificacion")
    _append_header(spec, ["Campo", "Valor"])
    for key, value in _mapping(validation.specification_json).items():
        spec.append([key, _cell_value(value)])
    if spec.max_row == 1:
        spec.append(["Sin especificacion final", ""])
    _finish_sheet(spec)

    return _artifact_from_workbook(
        workbook,
        artifact_type="iso_f10_03_xlsx",
        file_name=f"formulia_iso_f10_03_{_project_file_part(project)}.xlsx",
    )


def build_iso_metadata_json(
    project: IsoDesignProject,
    artifacts: list[IsoGeneratedArtifact],
) -> bytes:
    import json

    payload = {
        "project_id": str(project.id),
        "iso_request_number": project.iso_request_number,
        "project_code": project.project_code,
        "product_name": project.product_name,
        "artifacts": [
            {
                "artifact_type": artifact.artifact_type,
                "file_name": artifact.file_name,
                "checksum_sha256": artifact.checksum_sha256,
                "size_bytes": artifact.size_bytes,
            }
            for artifact in artifacts
        ],
    }
    return json.dumps(payload, indent=2, sort_keys=True, default=str).encode("utf-8")


def generated_binary_artifact(
    *,
    artifact_type: str,
    file_name: str,
    content_type: str,
    content: bytes,
) -> IsoGeneratedArtifact:
    return IsoGeneratedArtifact(
        artifact_type=artifact_type,
        file_name=file_name,
        content_type=content_type,
        checksum_sha256=hashlib.sha256(content).hexdigest(),
        size_bytes=len(content),
        content=content,
    )


def _append_project_key_values(sheet: Worksheet, project: IsoDesignProject) -> None:
    _append_key_values(
        sheet,
        [
            ("No Solicitud", project.iso_request_number),
            ("Ano", project.year),
            ("ProyectoID", project.project_code),
            ("Producto", project.product_name),
            ("Nombre comercial", project.commercial_name),
            ("Solicitante", project.requester),
            ("Necesidad", project.need),
            ("Tipo producto", project.product_type),
            ("Pais destino", project.destination_country),
            ("Envase", project.packaging),
            ("Aceptado", project.accepted_status),
            ("Estado ciclo", project.lifecycle_status),
            ("Fin planificado", _cell_value(project.planned_finish_at)),
            ("Finalizado", _cell_value(project.finished_at)),
            ("Comentarios", project.comments),
        ],
    )


def _artifact_from_workbook(
    workbook: Workbook,
    *,
    artifact_type: str,
    file_name: str,
) -> IsoGeneratedArtifact:
    stream = BytesIO()
    workbook.save(stream)
    content = stream.getvalue()
    return generated_binary_artifact(
        artifact_type=artifact_type,
        file_name=file_name,
        content_type=ISO_XLSX_CONTENT_TYPE,
        content=content,
    )


def _append_title(sheet: Worksheet, title: str) -> None:
    sheet.append([title])
    cell = sheet["A1"]
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = PatternFill("solid", fgColor="0F766E")
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24


def _append_key_values(sheet: Worksheet, rows: list[tuple[str, Any]]) -> None:
    if sheet.max_row > 1:
        sheet.append([])
    _append_header(sheet, ["Campo", "Valor"])
    for label, value in rows:
        sheet.append([label, _cell_value(value)])


def _append_header(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _finish_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in sheet.columns:
        column = column_cells[0].column_letter
        max_width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column].width = min(max(max_width + 2, 12), 46)


def _project_file_part(project: IsoDesignProject) -> str:
    return _safe_file_part(project.project_code or project.iso_request_number or str(project.id))


def _safe_file_part(value: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return normalized.strip("._-")[:48] or "iso"


def _cell_value(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, str | int | float | bool):
        return value
    return str(value)


def _mapping(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _sequence(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []
