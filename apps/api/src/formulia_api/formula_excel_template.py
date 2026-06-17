from __future__ import annotations

import hashlib
import re
import unicodedata
import uuid
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .formula_excel_contract import (
    CALCULATOR_FIRST_ITEM_ROW,
    CALCULATOR_FIRST_PARAMETER_COLUMN,
    CALCULATOR_MATERIAL_COLUMN,
    CALCULATOR_PERCENTAGE_COLUMN,
    CALCULATOR_PRICE_COLUMN,
    CALCULATOR_SHEET_NAME,
    COMPOSITION_FIRST_PARAMETER_ROW,
    COMPOSITION_PRICE_ROW,
    COMPOSITION_SHEET_NAME,
    FORMULA_ID_LAB_EXCEL_CONTENT_TYPE,
    FORMULA_ID_LAB_EXCEL_TYPE,
    FORMULA_ID_LAB_TEMPLATE_KEY,
    LAB_FIRST_ITEM_ROW,
    LAB_HALF_KG_COLUMN,
    LAB_MATERIAL_COLUMN,
    LAB_OBSERVATION_COLUMN,
    LAB_ORDER_COLUMN,
    LAB_PERCENTAGE_COLUMN,
    LAB_SHEET_NAME,
    LAB_TWO_KG_COLUMN,
)
from .parameter_order import parameter_sort_key


@dataclass(frozen=True)
class FormulaExcelParameter:
    code: str
    label: str
    unit: str | None = None
    decimals: int = 3


@dataclass(frozen=True)
class FormulaExcelItem:
    name: str
    percentage: float
    order_index: int = 0
    code: str | None = None
    lab_name: str | None = None
    lab_observation: str | None = None
    price: float | None = None
    parameters: dict[str, float] = field(default_factory=dict)


@dataclass(frozen=True)
class FormulaExcelMetadata:
    sample_code: str | None = None
    lab_date: date | None = None
    experiment_date: date | None = None
    density: float | None = None
    ph: float | None = None
    notes: str | None = None


@dataclass(frozen=True)
class FormulaExcelContext:
    name: str
    items: list[FormulaExcelItem]
    parameters: list[FormulaExcelParameter]
    version: int = 1
    metadata: FormulaExcelMetadata = field(default_factory=FormulaExcelMetadata)


@dataclass(frozen=True)
class FormulaExcelArtifact:
    file_name: str
    content_type: str
    checksum_sha256: str
    size_bytes: int
    content: bytes


def build_formula_id_lab_excel(
    context: FormulaExcelContext,
    *,
    suffix: str | None = None,
) -> FormulaExcelArtifact:
    workbook = Workbook()
    calculator = workbook.active
    calculator.title = CALCULATOR_SHEET_NAME
    lab = workbook.create_sheet(LAB_SHEET_NAME)
    composition = workbook.create_sheet(COMPOSITION_SHEET_NAME)

    parameters = sorted_export_parameters(context.parameters)
    items = sorted(context.items, key=lambda item: item.order_index)

    calculator_total_row = _write_calculator_sheet(calculator, items, parameters)
    _write_lab_sheet(lab, context, items)
    _write_composition_sheet(composition, parameters, calculator_total_row)
    _force_recalculation(workbook)

    stream = BytesIO()
    workbook.save(stream)
    content = stream.getvalue()
    file_name = _file_name(context.name, context.version, suffix=suffix)
    return FormulaExcelArtifact(
        file_name=file_name,
        content_type=FORMULA_ID_LAB_EXCEL_CONTENT_TYPE,
        checksum_sha256=hashlib.sha256(content).hexdigest(),
        size_bytes=len(content),
        content=content,
    )


def build_formula_id_lab_excel_from_snapshot(
    snapshot: dict[str, Any],
    review_id: uuid.UUID,
) -> FormulaExcelArtifact:
    formula = _mapping(snapshot.get("formula"))
    calculation = _mapping(snapshot.get("latest_calculation"))
    items = [_item_from_snapshot(item) for item in _sequence(snapshot.get("items"))]
    parameters = _parameters_from_snapshot(snapshot, calculation)
    context = FormulaExcelContext(
        name=str(formula.get("name") or "formula"),
        version=_int_or_default(formula.get("version"), 1),
        items=items,
        parameters=parameters,
        metadata=FormulaExcelMetadata(notes=_optional_str(snapshot.get("notes"))),
    )
    return build_formula_id_lab_excel(context, suffix=str(review_id)[:8])


def formula_excel_download_file_name(
    formula_name: str,
    *,
    download_date: date | None = None,
) -> str:
    safe_name = _safe_file_part(formula_name, max_length=96)
    date_part = (download_date or date.today()).isoformat()
    return f"DT_{safe_name}_{date_part}.xlsx"


def sorted_export_parameters(
    parameters: list[FormulaExcelParameter],
) -> list[FormulaExcelParameter]:
    return sorted(
        parameters,
        key=lambda parameter: (
            parameter_sort_key(parameter.label),
            parameter_sort_key(parameter.code),
        ),
    )


def _write_calculator_sheet(
    sheet: Worksheet,
    items: list[FormulaExcelItem],
    parameters: list[FormulaExcelParameter],
) -> int:
    headers = [
        "Materia Prima",
        "%",
        "Precio",
        *[parameter.label for parameter in parameters],
    ]
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_index, value=header)

    for row_offset, item in enumerate(items, start=CALCULATOR_FIRST_ITEM_ROW):
        sheet.cell(row=row_offset, column=CALCULATOR_MATERIAL_COLUMN, value=item.name)
        sheet.cell(row=row_offset, column=CALCULATOR_PERCENTAGE_COLUMN, value=item.percentage)
        sheet.cell(
            row=row_offset,
            column=CALCULATOR_PRICE_COLUMN,
            value=item.price if item.price is not None else None,
        )
        for parameter_index, parameter in enumerate(
            parameters,
            start=CALCULATOR_FIRST_PARAMETER_COLUMN,
        ):
            value = _parameter_value(item.parameters, parameter)
            sheet.cell(row=row_offset, column=parameter_index, value=value)

    last_item_row = CALCULATOR_FIRST_ITEM_ROW + len(items) - 1
    total_row = CALCULATOR_FIRST_ITEM_ROW + len(items) + 1
    sheet.cell(row=total_row, column=CALCULATOR_MATERIAL_COLUMN, value="TOTAL")
    sheet.cell(
        row=total_row,
        column=CALCULATOR_PERCENTAGE_COLUMN,
        value=f"=SUM(B{CALCULATOR_FIRST_ITEM_ROW}:B{total_row - 1})",
    )

    last_formula_row = max(last_item_row, CALCULATOR_FIRST_ITEM_ROW)
    for column in range(
        CALCULATOR_PRICE_COLUMN,
        CALCULATOR_FIRST_PARAMETER_COLUMN + len(parameters),
    ):
        sheet.cell(
            row=total_row,
            column=column,
            value=_calculator_weighted_average_formula(column, last_formula_row),
        )

    _format_calculator_sheet(sheet, len(headers))
    return total_row


def _write_lab_sheet(
    sheet: Worksheet,
    context: FormulaExcelContext,
    items: list[FormulaExcelItem],
) -> None:
    _setup_lab_page(sheet)
    _add_logo(sheet)
    metadata = context.metadata
    lab_date = metadata.lab_date or date.today()
    experiment_date = metadata.experiment_date or lab_date

    sheet["F2"] = f"FECHA: {lab_date.strftime('%d-%m-%Y')}"
    sheet["F2"].font = Font(name="Calibri", size=14, bold=True)
    sheet["F2"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("F2:G2")

    sheet["C3"] = _lab_title(context.name, metadata.sample_code)
    sheet["C3"].font = Font(name="Calibri", size=16, bold=True)
    sheet["C3"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("C3:F3")

    sheet["F4"] = lab_date
    sheet["F4"].number_format = "m/d/yy"
    sheet["F4"].fill = PatternFill("solid", fgColor="FFFF00")
    sheet["F4"].font = Font(name="Calibri", size=11, bold=True)
    sheet["F4"].border = _medium_border()

    headers = {
        "A5": "Órden de adición",
        "D5": "Cantidad % peso",
        "E5": "ENSAYO 0,5 kg",
        "F5": "ENSAYO 2 kg",
    }
    for cell, value in headers.items():
        sheet[cell] = value
    _format_lab_header(sheet)

    first_row = LAB_FIRST_ITEM_ROW
    for row_offset, item in enumerate(items, start=first_row):
        sheet.cell(
            row=row_offset,
            column=LAB_ORDER_COLUMN,
            value=row_offset - first_row + 1,
        )
        sheet.cell(
            row=row_offset,
            column=LAB_MATERIAL_COLUMN,
            value=item.lab_name or item.name,
        )
        sheet.cell(row=row_offset, column=LAB_PERCENTAGE_COLUMN, value=item.percentage)
        sheet.cell(
            row=row_offset,
            column=LAB_HALF_KG_COLUMN,
            value=_lab_batch_formula(row_offset, multiplier="/2"),
        )
        sheet.cell(
            row=row_offset,
            column=LAB_TWO_KG_COLUMN,
            value=_lab_batch_formula(row_offset, multiplier="*2"),
        )
        if item.lab_observation:
            sheet.cell(row=row_offset, column=LAB_OBSERVATION_COLUMN, value=item.lab_observation)
        _format_lab_item_row(sheet, row_offset)

    last_item_row = first_row + len(items) - 1
    total_row = first_row + len(items) + 1
    sum_last = max(last_item_row, first_row)
    sheet.cell(
        row=total_row,
        column=LAB_PERCENTAGE_COLUMN,
        value=_sum_formula(LAB_PERCENTAGE_COLUMN, first_row, sum_last),
    )
    sheet.cell(
        row=total_row,
        column=LAB_HALF_KG_COLUMN,
        value=_sum_formula(LAB_HALF_KG_COLUMN, first_row, sum_last),
    )
    sheet.cell(
        row=total_row,
        column=LAB_TWO_KG_COLUMN,
        value=_sum_formula(LAB_TWO_KG_COLUMN, first_row, sum_last),
    )
    for column in (LAB_PERCENTAGE_COLUMN, LAB_HALF_KG_COLUMN, LAB_TWO_KG_COLUMN):
        sheet.cell(row=total_row, column=column).number_format = "0"

    date_row = total_row + 1
    experiment_row = total_row + 2
    ph_row = experiment_row + 1
    sheet.cell(row=date_row, column=5, value=lab_date)
    sheet.cell(row=date_row, column=5).number_format = "m/d/yy"
    sheet.cell(row=date_row, column=5).fill = PatternFill("solid", fgColor="FFFF00")
    sheet.cell(row=date_row, column=5).font = Font(name="Calibri", size=11, bold=True)
    sheet.cell(row=date_row, column=5).border = _medium_border()

    sheet.merge_cells(start_row=experiment_row, start_column=3, end_row=ph_row, end_column=3)
    sheet.cell(
        row=experiment_row,
        column=3,
        value=f"EXPERIMENTAL {experiment_date.strftime('%d/%m/%Y')}",
    )
    sheet.cell(row=experiment_row, column=4, value="Densidad")
    sheet.cell(row=experiment_row, column=5, value=metadata.density)
    sheet.cell(row=ph_row, column=4, value="pH")
    sheet.cell(row=ph_row, column=5, value=metadata.ph)
    _format_lab_experimental_block(sheet, experiment_row, ph_row)


def _write_composition_sheet(
    sheet: Worksheet,
    parameters: list[FormulaExcelParameter],
    calculator_total_row: int,
) -> None:
    sheet["A2"] = "Parámetro"
    sheet["B2"] = "% p/p"
    sheet.cell(row=COMPOSITION_PRICE_ROW, column=1, value="Precio")
    sheet.cell(
        row=COMPOSITION_PRICE_ROW,
        column=2,
        value=_calculator_reference_formula(CALCULATOR_PRICE_COLUMN, calculator_total_row),
    )
    for row_offset, parameter in enumerate(parameters, start=COMPOSITION_FIRST_PARAMETER_ROW):
        sheet.cell(row=row_offset, column=1, value=parameter.label)
        sheet.cell(
            row=row_offset,
            column=2,
            value=_calculator_reference_formula(
                _calculator_parameter_column(row_offset),
                calculator_total_row,
            ),
        )
    sheet.column_dimensions["A"].width = 19
    sheet.column_dimensions["B"].width = 10
    for row in range(COMPOSITION_PRICE_ROW, COMPOSITION_FIRST_PARAMETER_ROW + len(parameters)):
        sheet.cell(row=row, column=2).number_format = "0.000"


def _calculator_parameter_column(composition_row: int) -> int:
    return (
        CALCULATOR_FIRST_PARAMETER_COLUMN
        + composition_row
        - COMPOSITION_FIRST_PARAMETER_ROW
    )


def _calculator_reference_formula(column: int, row: int) -> str:
    return f"='{CALCULATOR_SHEET_NAME}'!{get_column_letter(column)}{row}"


def _calculator_weighted_average_formula(column: int, last_formula_row: int) -> str:
    letter = get_column_letter(column)
    return (
        f"=SUMPRODUCT($B${CALCULATOR_FIRST_ITEM_ROW}:$B${last_formula_row},"
        f"{letter}{CALCULATOR_FIRST_ITEM_ROW}:{letter}{last_formula_row})/100"
    )


def _lab_batch_formula(row: int, *, multiplier: str) -> str:
    source = get_column_letter(LAB_PERCENTAGE_COLUMN)
    return f"={source}{row}*10{multiplier}"


def _sum_formula(column: int, first_row: int, last_row: int) -> str:
    letter = get_column_letter(column)
    return f"=SUM({letter}{first_row}:{letter}{last_row})"


def _format_calculator_sheet(sheet: Worksheet, column_count: int) -> None:
    sheet.column_dimensions[get_column_letter(CALCULATOR_MATERIAL_COLUMN)].width = 20
    sheet.column_dimensions[get_column_letter(CALCULATOR_PERCENTAGE_COLUMN)].width = 10
    sheet.column_dimensions[get_column_letter(CALCULATOR_PRICE_COLUMN)].width = 10
    for column in range(CALCULATOR_FIRST_PARAMETER_COLUMN, column_count + 1):
        header = str(sheet.cell(row=1, column=column).value or "")
        sheet.column_dimensions[get_column_letter(column)].width = min(
            max(len(header) + 1, 4),
            19,
        )


def _setup_lab_page(sheet: Worksheet) -> None:
    sheet.merge_cells("A1:XFD1")
    sheet.row_dimensions[1].height = 80.1
    for row in range(2, 20):
        sheet.row_dimensions[row].height = 18
    widths = {
        "A": 15.33,
        "B": 4.44,
        "C": 42,
        "D": 16.78,
        "E": 14.55,
        "F": 12.89,
        "G": 26.55,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _add_logo(sheet: Worksheet) -> None:
    logo_path = Path(__file__).parent / "assets" / "atlantica_logo.png"
    if not logo_path.exists():
        return
    try:
        logo = Image(str(logo_path))
    except Exception:
        return
    logo.width = 160
    logo.height = 72
    logo.anchor = "A1"
    sheet.add_image(logo)


def _format_lab_header(sheet: Worksheet) -> None:
    sheet["A5"].font = Font(name="Calibri", size=11, bold=True)
    sheet["A5"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A5"].border = Border(
        left=_medium_side(),
        right=_medium_side(),
        top=_medium_side(),
        bottom=_thin_side(),
    )
    for cell in ("D5", "E5", "F5"):
        sheet[cell].font = Font(name="Calibri", size=12, bold=True)
        sheet[cell].alignment = Alignment(horizontal="center")
    for cell in ("E5", "F5"):
        sheet[cell].border = Border(left=_medium_side(), top=_medium_side(), bottom=_medium_side())


def _format_lab_item_row(sheet: Worksheet, row: int) -> None:
    sheet.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True)
    sheet.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=1).border = Border(
        left=_medium_side(),
        right=_medium_side(),
        top=_thin_side(),
        bottom=_thin_side(),
    )
    sheet.cell(row=row, column=3).font = Font(name="Calibri", size=12, bold=True)
    sheet.cell(row=row, column=3).border = Border(
        left=_medium_side(),
        top=_medium_side(),
        bottom=_medium_side(),
    )
    sheet.cell(row=row, column=4).font = Font(name="Calibri", size=12)
    sheet.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    sheet.cell(row=row, column=4).border = Border(top=_medium_side(), bottom=_medium_side())
    for column in (5, 6):
        cell = sheet.cell(row=row, column=column)
        cell.font = Font(name="Calibri", size=12)
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = "0.0"
        cell.border = Border(left=_medium_side(), top=_medium_side(), bottom=_medium_side())


def _format_lab_experimental_block(sheet: Worksheet, first_row: int, last_row: int) -> None:
    fill = PatternFill("solid", fgColor="E2F0D9")
    for row in range(first_row, last_row + 1):
        for column in range(3, 6):
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name="Calibri", size=14, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _medium_border()
            if column == 3:
                cell.fill = fill
    sheet.cell(row=first_row, column=5).number_format = "0.00000"
    sheet.cell(row=last_row, column=5).number_format = "0.00"


def _force_recalculation(workbook: Workbook) -> None:
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"


def _parameter_value(
    values: dict[str, float],
    parameter: FormulaExcelParameter,
) -> float:
    return values.get(parameter.code, values.get(parameter.label, 0.0))


def _lab_title(name: str, sample_code: str | None) -> str:
    cleaned_name = name.strip() or "Formula"
    if not sample_code or "MUESTRA:" in cleaned_name.upper():
        return cleaned_name
    return f"{cleaned_name} - MUESTRA: {sample_code}"


def _medium_side() -> Side:
    return Side(style="medium", color="000000")


def _thin_side() -> Side:
    return Side(style="thin", color="000000")


def _medium_border() -> Border:
    side = _medium_side()
    return Border(left=side, right=side, top=side, bottom=side)


def _file_name(name: str, version: int, *, suffix: str | None = None) -> str:
    safe_name = _safe_file_part(name, max_length=48)
    extra = f"_{suffix}" if suffix else ""
    return f"formulia_id_lab_{safe_name}_v{version}{extra}.xlsx"


def _safe_file_part(value: str, *, max_length: int) -> str:
    ascii_value = unicodedata.normalize("NFKD", value.strip()).encode("ascii", "ignore").decode()
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", ascii_value)
    return normalized.strip("._-")[:max_length] or "formula"


def _item_from_snapshot(value: Any) -> FormulaExcelItem:
    item = _mapping(value)
    return FormulaExcelItem(
        name=str(item.get("name") or "Unknown material"),
        code=_optional_str(item.get("code")),
        lab_name=_optional_str(item.get("lab_name")),
        lab_observation=_optional_str(item.get("lab_observation")),
        percentage=_float_or_default(item.get("percentage"), 0.0),
        order_index=_int_or_default(item.get("order_index"), 0),
        price=_optional_float(item.get("price")),
        parameters=_parameters_dict(item.get("parameters")),
    )


def _parameters_from_snapshot(
    snapshot: dict[str, Any],
    calculation: dict[str, Any],
) -> list[FormulaExcelParameter]:
    rows: dict[str, FormulaExcelParameter] = {}
    for parameter in _sequence(calculation.get("parameters")):
        data = _mapping(parameter)
        code = _optional_str(data.get("code"))
        if not code:
            continue
        rows[code] = FormulaExcelParameter(
            code=code,
            label=code,
            unit=_optional_str(data.get("unit")),
        )
    for item in _sequence(snapshot.get("items")):
        for code in _parameters_dict(_mapping(item).get("parameters")):
            rows.setdefault(code, FormulaExcelParameter(code=code, label=code))
    return list(rows.values())


def _parameters_dict(value: Any) -> dict[str, float]:
    if isinstance(value, dict):
        return {
            str(key): float(raw_value or 0)
            for key, raw_value in value.items()
            if str(key).strip()
        }
    rows: dict[str, float] = {}
    for item in _sequence(value):
        data = _mapping(item)
        code = _optional_str(data.get("code"))
        if not code:
            continue
        rows[code] = _float_or_default(data.get("value"), 0.0)
    return rows


def _mapping(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _sequence(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _optional_str(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value)


def _float_or_default(value: Any, default: float) -> float:
    if value in (None, ""):
        return default
    return float(value)


def _int_or_default(value: Any, default: int) -> int:
    if value in (None, ""):
        return default
    return int(value)
