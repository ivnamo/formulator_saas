from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class FormulaExportSummary:
    id: str
    name: str
    version: int
    status: str


@dataclass(frozen=True)
class FormulaExportLine:
    order_index: int
    material_code: str | None
    material_name: str
    percentage: float
    price: float | None
    currency: str
    weighted_cost: float | None


def build_formula_xlsx(
    summary: FormulaExportSummary,
    lines: list[FormulaExportLine],
    calculation: dict[str, Any],
) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _write_summary(summary_sheet, summary, calculation)
    _write_lines(workbook.create_sheet("Lines"), lines)
    _write_parameters(workbook.create_sheet("Parameters"), calculation.get("parameters", []))
    _write_warnings(workbook.create_sheet("Warnings"), calculation.get("warnings", []))

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _write_summary(
    worksheet: Worksheet,
    summary: FormulaExportSummary,
    calculation: dict[str, Any],
) -> None:
    worksheet.append(["FormulIA Cloud export"])
    worksheet.append(["Formula", summary.name])
    worksheet.append(["Formula ID", summary.id])
    worksheet.append(["Version", summary.version])
    worksheet.append(["Status", summary.status])
    worksheet.append(["Total percentage", calculation.get("total_percentage")])
    worksheet.append(["Price total", calculation.get("price_total")])
    worksheet.append(["Currency", calculation.get("currency")])
    _style_title(worksheet, "A1:B1")
    _fit_columns(worksheet)


def _write_lines(worksheet: Worksheet, lines: list[FormulaExportLine]) -> None:
    worksheet.append(
        [
            "Order",
            "Code",
            "Raw material",
            "Percentage",
            "Price",
            "Currency",
            "Weighted cost",
        ]
    )
    for line in lines:
        worksheet.append(
            [
                line.order_index + 1,
                line.material_code,
                line.material_name,
                line.percentage,
                line.price,
                line.currency,
                line.weighted_cost,
            ]
        )
    _style_header(worksheet)
    _fit_columns(worksheet)


def _write_parameters(worksheet: Worksheet, parameters: list[dict[str, Any]]) -> None:
    worksheet.append(["Code", "Value", "Unit"])
    for parameter in parameters:
        worksheet.append(
            [
                parameter.get("code"),
                parameter.get("value"),
                parameter.get("unit"),
            ]
        )
    _style_header(worksheet)
    _fit_columns(worksheet)


def _write_warnings(worksheet: Worksheet, warnings: list[dict[str, Any]]) -> None:
    worksheet.append(["Code", "Message", "Raw material ID", "Parameter code"])
    for warning in warnings:
        worksheet.append(
            [
                warning.get("code"),
                warning.get("message"),
                warning.get("raw_material_id"),
                warning.get("parameter_code"),
            ]
        )
    _style_header(worksheet)
    _fit_columns(worksheet)


def _style_title(worksheet: Worksheet, cell_range: str) -> None:
    worksheet.merge_cells(cell_range)
    cell = worksheet["A1"]
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0F766E")


def _style_header(worksheet: Worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")


def _fit_columns(worksheet: Worksheet) -> None:
    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        width = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(width + 2, 12),
            42,
        )
