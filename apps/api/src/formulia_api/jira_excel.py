from __future__ import annotations

import hashlib
import re
import uuid
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


JIRA_REVIEW_EXCEL_TYPE = "jira_review_xlsx"
JIRA_REVIEW_EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@dataclass(frozen=True)
class JiraReviewExcel:
    file_name: str
    content_type: str
    checksum_sha256: str
    size_bytes: int
    content: bytes


def build_jira_review_excel(
    snapshot: dict[str, Any],
    review_id: uuid.UUID,
) -> JiraReviewExcel:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"

    _write_summary_sheet(summary_sheet, snapshot, review_id)
    _write_composition_sheet(workbook.create_sheet("Composicion"), snapshot)
    _write_calculation_sheet(workbook.create_sheet("Calculo"), snapshot)
    _write_metadata_sheet(workbook.create_sheet("Metadatos"), snapshot, review_id)

    stream = BytesIO()
    workbook.save(stream)
    content = stream.getvalue()
    file_name = _file_name(snapshot, review_id)
    return JiraReviewExcel(
        file_name=file_name,
        content_type=JIRA_REVIEW_EXCEL_CONTENT_TYPE,
        checksum_sha256=hashlib.sha256(content).hexdigest(),
        size_bytes=len(content),
        content=content,
    )


def _write_summary_sheet(
    sheet: Worksheet,
    snapshot: dict[str, Any],
    review_id: uuid.UUID,
) -> None:
    formula = _mapping(snapshot.get("formula"))
    jira = _mapping(snapshot.get("jira"))
    calculation = _mapping(snapshot.get("latest_calculation"))
    rows = [
        ("Package", "FormulIA Jira review"),
        ("Review ID", str(review_id)),
        ("Formula ID", formula.get("id")),
        ("Formula name", formula.get("name")),
        ("Formula version", formula.get("version")),
        ("Formula status", formula.get("status")),
        ("Objective", formula.get("objective")),
        ("Total price", calculation.get("price_total") or formula.get("total_price")),
        ("Currency", calculation.get("currency") or formula.get("currency")),
        ("Jira project", jira.get("project_key")),
        ("Jira issue type", jira.get("issue_type")),
        ("Jira summary", jira.get("issue_summary")),
        ("Jira assignee", jira.get("assignee")),
        ("Notes", snapshot.get("notes")),
    ]
    _write_key_values(sheet, rows)


def _write_composition_sheet(sheet: Worksheet, snapshot: dict[str, Any]) -> None:
    _append_header(
        sheet,
        ["Order", "Code", "Raw material", "Percentage", "Quantity", "Unit", "Raw material ID"],
    )
    for item in _sequence(snapshot.get("items")):
        item_data = _mapping(item)
        sheet.append(
            [
                item_data.get("order_index"),
                item_data.get("code"),
                item_data.get("name"),
                item_data.get("percentage"),
                item_data.get("quantity"),
                item_data.get("unit"),
                item_data.get("raw_material_id"),
            ]
        )
    if sheet.max_row == 1:
        sheet.append(["", "", "No formula lines", "", "", "", ""])
    _finish_table(sheet)


def _write_calculation_sheet(sheet: Worksheet, snapshot: dict[str, Any]) -> None:
    calculation = _mapping(snapshot.get("latest_calculation"))
    _append_header(sheet, ["Type", "Code", "Value", "Unit", "Severity", "Message", "Action"])
    for parameter in _sequence(calculation.get("parameters")):
        parameter_data = _mapping(parameter)
        sheet.append(
            [
                "Parameter",
                parameter_data.get("code"),
                parameter_data.get("value"),
                parameter_data.get("unit"),
                "",
                "",
                "",
            ]
        )
    for warning in _sequence(calculation.get("warnings")):
        warning_data = _mapping(warning)
        sheet.append(
            [
                "Validation",
                warning_data.get("code"),
                "",
                "",
                warning_data.get("severity") or "warning",
                warning_data.get("message"),
                warning_data.get("recommended_action"),
            ]
        )
    if sheet.max_row == 1:
        sheet.append(["Info", "", "", "", "", "No calculation snapshot", ""])
    _finish_table(sheet)


def _write_metadata_sheet(
    sheet: Worksheet,
    snapshot: dict[str, Any],
    review_id: uuid.UUID,
) -> None:
    jira = _mapping(snapshot.get("jira"))
    rows = [
        ("Snapshot type", snapshot.get("snapshot_type")),
        ("Review ID", str(review_id)),
        ("Jira connection ID", jira.get("connection_id")),
        ("Jira base URL", jira.get("base_url")),
        ("Artifact type", JIRA_REVIEW_EXCEL_TYPE),
    ]
    _write_key_values(sheet, rows)


def _write_key_values(sheet: Worksheet, rows: list[tuple[str, Any]]) -> None:
    sheet.append(["Field", "Value"])
    for label, value in rows:
        sheet.append([label, _cell_value(value)])
    _finish_table(sheet)


def _append_header(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.alignment = Alignment(vertical="center")


def _finish_table(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
    for column_cells in sheet.columns:
        column = column_cells[0].column_letter
        max_width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column].width = min(max(max_width + 2, 12), 42)


def _file_name(snapshot: dict[str, Any], review_id: uuid.UUID) -> str:
    formula = _mapping(snapshot.get("formula"))
    name = _safe_file_part(str(formula.get("name") or "formula"))
    version = formula.get("version") or "1"
    return f"formulia_jira_review_{name}_v{version}_{str(review_id)[:8]}.xlsx"


def _safe_file_part(value: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return normalized.strip("._-")[:48] or "formula"


def _mapping(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _sequence(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _cell_value(value: Any) -> str | int | float | None:
    if value is None:
        return None
    if isinstance(value, str | int | float):
        return value
    return str(value)
